"""
ipo_listing_gainers.py — screen IPOs that gained >=50% on listing day or within
a month of listing, and map the anchor investors behind them.
================================================================================

PURPOSE
-------
Answer two linked questions in one pass:

  1. WHICH recent Indian IPOs were big listing-pops (>=50%)?
  2. WHO are the anchor investors that keep showing up in those deals?

(2) is the real payload. Anchor allocation is decided the day before an issue
opens, so a fund that repeatedly anchors IPOs that then pop is a signal you can
track forward. The frequency sheet is what surfaces that pattern.

WHAT IT DOES
------------
For every equity IPO listed in a chosen window (default: 01-Jan-2025 .. today),
compute two returns versus the issue price and flag the stock if EITHER crosses
a threshold (default +50%):

  1. Listing-day return   = (listing-day CLOSE - issue price) / issue price
  2. 30-day peak return   = (highest CLOSE within 30 calendar days of listing
                             - issue price) / issue price

Both are measured on CLOSE, not intraday high, so the numbers are achievable
rather than theoretical. EITHER crossing qualifies the IPO: a stock that opened
flat and ran up over the next month is as interesting as a listing-day pop.

Qualifying IPOs are printed as a table (biggest gain first) and written to a
four-sheet Excel `ipo_listing_gainers.xlsx` plus a tabbed dashboard
`ipo_listing_gainers.html` (both alongside this script; there is no CSV):
  * Sheet `All IPOs`           — every IPO in the window, qualified or not,
                                 newest first, with both returns and a note
                                 wherever a return could not be computed.
  * Sheet `Gainers`            — the qualifying IPOs (one row each).
  * Sheet `Anchor Investors`   — one row per anchor allocation:
                                 Company | Anchor Investor | Amount Invested ₹ |
                                 Shares Allotted | Price ₹.
  * Sheet `Investor Frequency` — how often each anchor investor recurs, keyed
                                 on the first two words of its name.
The HTML carries the same four tabs, styled like market_charts.html.

END-TO-END WORKFLOW
-------------------
Three printed stages ([1/3]..[3/3]), in this order:

  [1/3] BUILD THE IPO MASTER
        fetch_nse_past_issues() pulls the NSE past-issues API, folds every row
        into the append-only ledger `data/ipo/nse_past_issues.json`, and returns
        the LEDGER (not the live payload). collect_nse_ipos() then filters to
        securityType EQ/SME with a listing date inside --start..--end and a
        parseable issue price. With --bse-recent, fetch_bse_public_issues()
        does the same for BSE's ~20-row live window into
        `data/ipo/bse_public_issues.json`, and collect_bse_recent() adds any
        company not already matched by name against the NSE set.

  [2/3] PRICE EVERY IPO
        compute_returns() runs in a ThreadPoolExecutor (--workers, default 5).
        Each IPO's OHLCV comes from the repo's data_provider.download()
        (Angel One -> jugaad -> yfinance) through ohlcv_cache, so repeat runs
        fetch only new bars. Progress prints every 25 symbols. IPOs whose price
        history cannot be resolved are dropped with a note, not silently.
        Survivors crossing --threshold on either measure become "winners".

  [3/3] ANCHORS, FREQUENCY, WRITE
        sync_anchor_filings() downloads any missing anchor filing straight from
        NSE's archive (skip with --no-fetch). collect_anchor_investors() then
        matches files in --anchor-dir to winners by symbol or company name and
        extracts every row — name, shares, price, amount — digital text first,
        OCR fallback. build_investor_frequency() aggregates the names.
        read_existing_anchors() reads the CURRENT workbook, merge_anchors()
        unions old+new per issuer, and write_workbook() rewrites all three
        sheets; write_html() renders the same three as tabs. A per-symbol
        coverage line and a TOTAL are printed.

HOW ANCHOR FILINGS ARE OBTAINED
-------------------------------
NSE publishes every issue's anchor letter at a deterministic archive URL:

    https://nsearchives.nseindia.com/content/ipo/ANCHOR_<SYMBOL>.zip

The zip appears on anchor-allocation day (the day before the issue opens),
is retained after listing, and holds a single PDF whose INNER filename is
arbitrary — so the extractor takes whatever *.pdf it finds rather than
guessing a name. sync_anchor_filings() walks the winners plus NSE's
upcoming/current issue feeds and downloads anything not already on disk, so
the folder fills itself; dropping files in by hand still works as a fallback.

A 404 is not automatically a gap. It is cross-checked against
/api/ipo-detail: an issue whose detail payload never mentions an anchor
portion (SME issues with only a market-maker reservation) simply has no anchor
book and is reported as such, separately from one whose letter has not been
filed yet.

HOW ANCHOR EXTRACTION WORKS (the hard part)
-------------------------------------------
The complete anchor list exists only in the official "Allocation to Anchor
Investors" filing; free web sources expose just the first ~2 names. The fetcher
above puts that filing in `--anchor-dir` (default ./anchor_pdfs); anything you
place there yourself, named with the symbol or company (e.g. `MEESHO.pdf`), is
picked up too.

Both digital and SCANNED filings are handled:
  * Digital PDF — pdfplumber reads the text layer directly.
  * Scanned PDF — no text layer, so pdftoppm renders pages and tesseract OCRs
    them across four render/segmentation passes, OCR_PASSES =
    (300,psm6), (450,psm4), (450,psm6), (600,psm4). Pages scanned sideways are
    rotated upright first (tesseract's OSD mode), because on a 90-degree page
    tesseract returns confident-looking gibberish rather than failing.
  * Images (.png/.jpg/...) — same, at native resolution with psm 6/4/11.
  * OCR text is cached under .cache/ipo_gainers/ocr, keyed on the file's content
    hash, so re-runs cost a file read instead of minutes of rendering.

ARITHMETIC VERIFICATION. OCR can silently drop a table row, and a dropped row
is invisible by eye. So each pass is reconciled against the letter's own stated
total: the per-row share counts must sum to the "allocation of N Equity Shares"
figure. Before that check, _correct_shares() repairs any share cell its own row
disagrees with (amount / price), and _trim_to_total() finds the prefix of rows
that sums exactly — page furniture and closing prose sit after the table, so
the genuine rows are a prefix. The first pass that reconciles wins.

If NO pass reconciles, the names are STILL returned, flagged OCR UNVERIFIED,
and the pass kept is the one with the MOST names. Rejecting outright was the
original behaviour and it left 14 filings with no anchor data at all; ranking
the fallback on arithmetic closeness instead of name count picked pages of
letterhead over correctly-read tables. In that branch every figure is already
flagged as needing checking, so there is nothing to protect by preferring
closer arithmetic — whereas a dropped investor is exactly what must not happen.

THE MONEY COLUMNS. Shares are read per row. The PRICE is not: every anchor in
an issue is allotted at the SAME price, so the table is over-determined and
_resolve_allocation_price() lets each row vote for the price that makes its own
shares x price match its own amount cell, taking the modal winner. The Amount
column written out is then shares x that price — derived, not OCR'd, because
the amount cell is the widest and most comma-riddled field on the page and the
one tesseract mangles most often.

MERGE SEMANTICS (why re-running is safe)
----------------------------------------
Sheet 2 is MERGED on every run, never overwritten. Anchor rows already in the
workbook are preserved — including rows you added BY HAND for IPOs outside the
gainers list — with existing order kept and only genuinely new names appended
(dedupe is on alphanumerics, so punctuation/case/spacing variants do not double
up). Where an existing row has no figures, freshly parsed figures are filled in
without touching the name, so hand-corrected spellings survive. Workbooks
written before Sheet 2 became tabular (one COLUMN per symbol) are migrated
automatically, names first and figures backfilled from the filings.

Note that pandas' ExcelWriter rewrites the whole file, so this explicit
read-merge-write is what makes re-running safe; do not bypass write_workbook().

Sheets you add yourself (anything not named in MANAGED_SHEETS) are also carried
across the rewrite: their cell VALUES are snapshotted before the write and
re-appended after it. Formatting, formulas, charts and column widths on those
sheets are NOT preserved — only the values. Keep anything irreplaceable in a
separate file.

INVESTOR FREQUENCY (Sheet 3)
----------------------------
Counts how often each investor appears. Deciding that two name cells are the
same investor is the hard part: the filings are inconsistent and OCR adds row
serials, truncations and cells where two names have been run together. Names are
therefore reduced to their significant tokens — case, punctuation, accents and
structural words (Ltd, Private, Trust, Fund, PCC, Cell, AIF …) dropped — and a
shorter name is folded into a fuller one only when they share a head token, the
fuller name adds at most three tokens, that fuller name is not a one-off, and the
choice is unambiguous. So "Saint Capital Fund", "SAINT CAPITAL FUND" and "Saint
Capital Limited" collapse to one row, while HDFC's Manufacturing and Business
Cycle schemes stay apart. `--name-key words` restores the old first-N-words key
(`--freq-words` sets N), which both over-merged whole fund houses and split their
spelling variants. `Qualified IPOs` counts the issues that cleared the gain
threshold and `Total IPOs` every issue the investor anchored, so sort by
`Qualified IPOs` descending to find the genuinely repeat anchors.

DATA SOURCES
------------
* IPO master : NSE public past-issues API
    https://www.nseindia.com/api/public-past-issues?index=equities
  Covers NSE Mainboard (securityType "EQ") and NSE SME / Emerge ("SME").
  Mainboard IPOs that dual-list on BSE are included here too (they list on NSE
  as well), so BSE-mainboard is effectively covered.
* BSE supplement (--bse-recent) : BSE's public-issues JSON, which only ever
  exposes the ~20 issues open RIGHT NOW.
* Prices     : repo `data_provider.download()` (Angel One -> jugaad -> yfinance),
  fronted by `ohlcv_cache` so re-runs pull only new bars.

RESILIENCE: append-only ledgers (`data/ipo/`)
---------------------------------------------
Both exchange feeds are fragile: the NSE endpoint is undocumented and could be
blocked without notice, and BSE's only ever shows the ~20 issues open right now.
Every run therefore folds the live payload into a ledger that is never pruned,
and the LEDGER is what feeds the screen. Consequences:
  * If NSE blocks the API, past IPOs (and their issue prices, which no other
    bulk source publishes) still work; only NEW listings are lost.
  * BSE-SME-only coverage GROWS with each run, since each run banks that day's
    live window. Issues that closed before the ledger existed remain missing.

Ledger mechanics: `nse_past_issues.json` keyed `symbol|listingDate`,
`bse_public_issues.json` keyed `Scrip_cd|Start_Dt`; each record carries
`_first_seen`/`_last_seen`; records are updated in place and NEVER deleted.
Saves are atomic (write .tmp, then replace) so an interrupted run cannot leave
a truncated ledger. A run that reaches the network prints
`[nse] ledger +N new, M revised — T total`; a run that cannot prints a warning
and serves the ledger alone. This is proven behaviour, not aspiration — a BSE
timeout during testing was absorbed exactly this way.

KNOWN LIMITATIONS (stated, not hidden)
--------------------------------------
* BSE-SME-only listings come from IPOPlatform, not from an exchange. Neither
  exchange can supply them: they are absent from the NSE feed by definition,
  and BSE serves no dated history (probed — the notices floor is a hard ~5
  months and the scrip master carries no listing date or issue price). Their
  listing date and issue price are therefore third-party figures, flagged
  Exchange=BSE / Segment='BSE SME'. Anchor books are NOT available for them:
  IPOPlatform publishes anchor shares for only 18 of its 634 BSE-SME rows.
* Returns use CLOSE only; a stock that spiked and faded intraday is not caught.
* No survivorship/delisting handling — an IPO whose price history vanished is
  reported as unpriceable rather than inferred.

OUTPUTS (written next to this script)
-------------------------------------
  ipo_listing_gainers.xlsx   Gainers + Anchor Investors + Investor Frequency.
  ipo_listing_gainers.html   The same three, as tabs.
  anchor_pdfs/               Anchor letters downloaded from NSE, SYMBOL.pdf.
  data/ipo/*.json            Append-only feed ledgers (keep these; they are
                             the only durable record of past issue prices).
  .cache/ipo_gainers/        On-disk caches. Deleting them costs a ~4h rebuild,
                             not data. See CACHES below.

CACHES — WHY THE SECOND RUN TAKES ~150s INSTEAD OF ~4 HOURS
-----------------------------------------------------------
Everything under `.cache/ipo_gainers/` is keyed on the SHA of the source file,
not on its path, so the caches are machine-independent:

  anchors/      Frozen readings. A filing whose parsed share total reconciles
                EXACTLY against the stated total is final; it is never read
                again. Keyed {sha}-{ANCHOR_RESULT_SCHEMA}, deliberately WITHOUT
                the OCR version, so it survives an OCR_PIPELINE_VERSION bump.
                ~40 filings, ~176 KB.
  ocr/          Per-page OCR artefacts, keyed with OCR_PIPELINE_VERSION.
                Bumping that constant discards all of them.
  render/       pdftoppm page images, keyed with OCR_RENDER_VERSION. Bump only
                when the IMAGE changes (dpi / deskew / enhance).
  ipoplatform/  Anchor-book HTML. Settled pages are kept effectively forever
                (IPL_TTL_SETTLED); live/forthcoming pages carry a 7-day TTL.

Filings that have NOT reconciled are re-read on every run — that is the point,
so a later parser improvement can still fix them. They are cheap because their
OCR artefacts are cached. Only `--rebuild-anchors` bypasses the frozen store.

MOVING TO ANOTHER MACHINE
-------------------------
`.cache/` is gitignored, as are anchor_pdfs/ and the workbook, so a fresh copy
of this script starts fully cold: it re-downloads the filings and re-OCRs
everything once (~4h), after which that machine is warm too.

To skip that, copy the caches across — the SHA keys make them portable:
  .cache/ipo_gainers/anchors/       176 KB  → ~40 filings skipped immediately
  .cache/ipo_gainers/ocr/            29 MB  → the rest skip re-OCR
  .cache/ipo_gainers/ipoplatform/   379 MB  → no anchor-book re-fetch
Copying just `anchors/` buys most of the benefit.

RUNNING IT
----------
Prerequisites:
  * Python 3.9+ and the repo venv:   source .venv/bin/activate
  * pip install requests pandas pdfplumber openpyxl
  * For SCANNED filings only: tesseract + poppler (`brew install tesseract
    poppler`), invoked as external binaries — no pytesseract/pdf2image needed.
    The Surya tier is a separate install (`pip install surya-ocr`).
    WARNING: if these binaries are missing the run does NOT fail. _ocr_available()
    returns False, scanned filings are skipped, and you get a workbook with FEWER
    anchor names and no error. On a new machine verify the binaries BEFORE
    trusting the first run's output — "it ran fine" and "the data is right" come
    apart here.
  * Repo-local data_provider / angel_client / ohlcv_cache must be importable.

CLOSE THE WORKBOOK FIRST. If ipo_listing_gainers.xlsx is open in Excel the
write is clobbered or blocked; a stale `~$ipo_listing_gainers.xlsx` lock file
next to it is the tell.

    cd /Users/ankit.srivastava/Documents/Analysis
    source .venv/bin/activate
    python3 ipo_listing_gainers.py

Common invocations:
    python3 ipo_listing_gainers.py                     # 2025-01-01..today, >=50%
    python3 ipo_listing_gainers.py --start 2025-08-01  # custom window start
    python3 ipo_listing_gainers.py --threshold 100     # only >=100% movers
    python3 ipo_listing_gainers.py --window-days 60    # wider peak lookback
    python3 ipo_listing_gainers.py --workers 6         # price-fetch concurrency
    python3 ipo_listing_gainers.py --limit 20          # debug: first N IPOs
    python3 ipo_listing_gainers.py --bse-recent        # add BSE live supplement
    python3 ipo_listing_gainers.py --no-ocr            # text-layer PDFs only
    python3 ipo_listing_gainers.py --no-fetch          # don't download filings
    python3 ipo_listing_gainers.py --anchor-dir ./anchor_pdfs   # anchor PDFs folder

FLAGS
-----
  --start DATE        default 2025-01-01   listing window start
  --end DATE          default today        listing window end
  --threshold PCT     default 50           minimum qualifying gain
  --window-days N     default 30           peak lookback after listing
  --workers N         default 5            price-fetch threads
  --limit N           default 0 (all)      debug: first N IPOs only
  --bse-recent        off                  add BSE live-window supplement
  --anchor-dir PATH   ./anchor_pdfs        folder of anchor filings
  --name-key smart|words  default smart      how two name cells are matched
  --freq-words N      default 2            words to key on (--name-key words)
  --no-ocr            off                  skip OCR; text-layer PDFs only
  --no-fetch          off                  skip the NSE anchor-filing download
  --xlsx PATH         ./ipo_listing_gainers.xlsx
  --html PATH         ./ipo_listing_gainers.html

TYPICAL SESSION
---------------
  1. Run with no flags. Anchor letters download themselves, are parsed and
     merged, and earlier rows are preserved.
  2. Read Sheet 1 for the gainers, Sheet 2 for who anchored them.
  3. Read Sheet 3 sorted by IPOs desc to see which investors recur.
  4. For anything the fetcher reports as PENDING, re-run after the anchor
     letter is filed (the day before that issue opens).

TROUBLESHOOTING
---------------
  "0/N symbols with anchors"   anchor_pdfs/ is empty or filenames do not match
                               a symbol/company. Rename to SYMBOL.pdf.
  "OCR UNVERIFIED"             no render pass reconciled to the stated share
                               total (or the filing states no total). The NAMES
                               are still written; the shares/price/amount on
                               those rows are the suspect part, so spot-check
                               them against the filing.
  Anchor names did not update  the workbook was open in Excel during the write.
  Feed warning + normal output the ledger covered for a failed fetch. Expected.
  Fewer anchor names than      tesseract/poppler are not on PATH, so every
  a previous machine           scanned filing was silently skipped. Check with
                               `which pdftoppm tesseract`.
  Run is slow again (~hours)   OCR_PIPELINE_VERSION was bumped, or .cache/ was
                               deleted. Frozen filings are still skipped.

DEPENDENCIES: requests, pandas, pdfplumber, openpyxl (+ the repo's
data_provider / angel_client), and optionally tesseract/poppler for scans.
"""
from __future__ import annotations

import argparse
import csv
import difflib
import hashlib
import html as _html
import io
import json
import re
import shutil
import subprocess
import sys
import tempfile
import time
import unicodedata
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timedelta
from functools import lru_cache
from pathlib import Path
from urllib.parse import urlencode

import pandas as pd
import requests

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))
# Outputs land next to this script, not in Output/.
OUTPUT_DIR = SCRIPT_DIR
CACHE_DIR = SCRIPT_DIR / ".cache" / "ipo_gainers"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
# Durable, append-only copies of the exchange feeds. Deliberately NOT under
# .cache/ — that directory is disposable, this data is not recoverable.
LEDGER_DIR = SCRIPT_DIR / "data" / "ipo"

# Sheets this script owns and rewrites on every run. Anything else in the
# workbook was put there by hand and is preserved verbatim (see
# read_extra_sheets) instead of being lost to the ExcelWriter rebuild.
MANAGED_SHEETS = ("All IPOs", "Gainers", "Anchor Investors", "Investor Frequency",
                  "Tracked Investors")

UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")
NSE_HEADERS = {
    "User-Agent": UA,
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.nseindia.com/",
}
BSE_HEADERS = {
    "User-Agent": UA,
    "Accept": "application/json, text/plain, */*",
    "Referer": "https://www.bseindia.com/",
    "Origin": "https://www.bseindia.com",
}


# ───────────────────────────── data model ──────────────────────────────────
@dataclass
class IPO:
    company: str
    symbol: str                     # NSE ticker (or BSE scrip code as str)
    listing_date: date
    issue_price: float
    segment: str                    # "Mainboard" | "NSE SME" | "BSE SME"
    exchange_suffix: str = ".NS"    # ".NS" or ".BO" for yfinance fallback
    # BSE SME trades under a numeric scrip code but is READ as a ticker, so the
    # symbol shown and the symbol quoted are not the same string. Blank = same.
    quote_symbol: str = ""
    # filled in later
    listing_close: float | None = None
    listing_pct: float | None = None
    peak_close: float | None = None
    peak_date: date | None = None
    peak_pct: float | None = None
    note: str = ""


@dataclass
class Anchor:
    """One row of an anchor-allocation table.

    `amount` is DERIVED as shares x price rather than taken from the OCR'd
    amount column: the allocation price is a single figure for the whole issue,
    so the product is exact, whereas the amount column is the widest, most
    comma-riddled field on the page and the one OCR mangles most often."""
    name: str
    shares: int | None = None
    price: float | None = None
    amount: float | None = None


# ───────────────────────── tiny disk cache ─────────────────────────────────
def _cache_path(key: str) -> Path:
    safe = re.sub(r"[^A-Za-z0-9._-]", "_", key)[:200]
    return CACHE_DIR / (safe + ".json")


def cache_get(key: str, ttl_seconds: int):
    p = _cache_path(key)
    if not p.exists() or (time.time() - p.stat().st_mtime) > ttl_seconds:
        return None
    try:
        return json.loads(p.read_text())
    except Exception:
        return None


def cache_set(key: str, value) -> None:
    try:
        _cache_path(key).write_text(json.dumps(value, default=str))
    except Exception:
        pass


# ─────────────────────── append-only source ledgers ────────────────────────
# The exchange feeds are windowed or revocable: NSE's past-issues API is
# undocumented and could be blocked, and BSE's public-issue API only ever
# exposes the ~20 issues that are live right now. A plain TTL cache would let
# that history evaporate. These ledgers therefore ACCUMULATE every record ever
# seen and are never pruned, so a dead endpoint costs us only NEW issues.
NSE_LEDGER = "nse_past_issues.json"
BSE_LEDGER = "bse_public_issues.json"


def _ledger_load(name: str) -> dict:
    p = LEDGER_DIR / name
    if not p.exists():
        return {}
    try:
        data = json.loads(p.read_text())
        return data if isinstance(data, dict) else {}
    except Exception as e:
        print(f"  [ledger] {name} unreadable ({e}); starting empty.", file=sys.stderr)
        return {}


def _ledger_save(name: str, store: dict) -> None:
    # Write-then-rename: an interrupted run can never leave a half-written
    # ledger behind, which would silently destroy accumulated history.
    try:
        LEDGER_DIR.mkdir(parents=True, exist_ok=True)
        tmp = LEDGER_DIR / (name + ".tmp")
        tmp.write_text(json.dumps(store, default=str, sort_keys=True, indent=0))
        tmp.replace(LEDGER_DIR / name)
    except Exception as e:
        print(f"  [ledger] could not save {name}: {e}", file=sys.stderr)


def _ledger_merge(store: dict, rows: list[dict], key_fn) -> tuple[int, int]:
    """Fold `rows` into `store`. Records already present are updated in place
    (the feed does correct itself); records absent from `rows` are KEPT."""
    stamp = date.today().isoformat()
    added = updated = 0
    for row in rows:
        if not isinstance(row, dict):
            continue
        key = key_fn(row)
        if not key:
            continue
        prev = store.get(key)
        if prev is None:
            store[key] = {**row, "_first_seen": stamp, "_last_seen": stamp}
            added += 1
        else:
            changed = any(prev.get(k) != v for k, v in row.items())
            store[key] = {**prev, **row,
                          "_first_seen": prev.get("_first_seen", stamp),
                          "_last_seen": stamp}
            updated += int(changed)
    return added, updated


def _nse_ledger_key(r: dict) -> str:
    sym = str(r.get("symbol") or r.get("htmSym") or r.get("company") or "").strip().upper()
    return f"{sym}|{str(r.get('listingDate') or '').strip()}" if sym else ""


def _bse_ledger_key(r: dict) -> str:
    code = str(r.get("Scrip_cd") or "").strip()
    return f"{code}|{str(r.get('Start_Dt') or '').strip()}" if code else ""


# ───────────────────────────── parsing ─────────────────────────────────────
def parse_date(raw: str) -> date | None:
    if not raw or raw.strip() in ("-", ""):
        return None
    for fmt in ("%d-%b-%Y", "%d-%B-%Y", "%d-%m-%Y", "%Y-%m-%d",
                "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(raw.strip(), fmt).date()
        except ValueError:
            continue
    return None


def parse_price(raw_price: str, price_range: str = "") -> float | None:
    """Issue price as a number; fall back to the UPPER end of the band."""
    def _num(s: str):
        m = re.search(r"(\d+(?:\.\d+)?)", str(s).replace(",", ""))
        return float(m.group(1)) if m else None

    if raw_price and str(raw_price).strip() not in ("-", ""):
        v = _num(raw_price)
        if v:
            return v
    if price_range:
        nums = re.findall(r"(\d+(?:\.\d+)?)", str(price_range).replace(",", ""))
        if nums:
            return float(nums[-1])
    return None


# ────────────────────────── NSE IPO master ─────────────────────────────────
def nse_session() -> requests.Session:
    s = requests.Session()
    s.headers.update(NSE_HEADERS)
    for u in ("https://www.nseindia.com",
              "https://www.nseindia.com/market-data/all-upcoming-issues-ipo"):
        try:
            s.get(u, timeout=20)
            time.sleep(0.4)
        except Exception:
            pass
    return s


def fetch_nse_past_issues() -> list[dict]:
    """Live feed folded into the append-only ledger, and the LEDGER is returned.
    If NSE ever blocks the endpoint we still serve every record captured so far
    — crucially the issue prices, which no other bulk source publishes."""
    store = _ledger_load(NSE_LEDGER)
    fresh = cache_get("nse_past_issues", ttl_seconds=6 * 3600)
    if fresh is None:
        try:
            s = nse_session()
            r = s.get("https://www.nseindia.com/api/public-past-issues?index=equities",
                      timeout=30)
            r.raise_for_status()
            fresh = r.json()
            if not isinstance(fresh, list) or not fresh:
                raise ValueError(f"unexpected payload: {type(fresh).__name__}")
            cache_set("nse_past_issues", fresh)
        except Exception as e:
            if not store:
                raise                      # no live feed AND no history: fatal
            print(f"  [nse] live feed unavailable ({e}); serving ledger only — "
                  f"{len(store)} record(s), no new IPOs this run.", file=sys.stderr)
            fresh = None
    if fresh:
        added, updated = _ledger_merge(store, fresh, _nse_ledger_key)
        _ledger_save(NSE_LEDGER, store)
        if added or updated:
            print(f"  [nse] ledger +{added} new, {updated} revised — {len(store)} total.")
    return list(store.values())


def collect_nse_ipos(start: date, end: date) -> list[IPO]:
    out: list[IPO] = []
    for r in fetch_nse_past_issues():
        ld = parse_date(r.get("listingDate", ""))
        if ld is None or ld < start or ld > end:
            continue
        sec = (r.get("securityType") or "").upper()
        if sec not in ("EQ", "SME"):
            continue
        sym = (r.get("symbol") or "").upper().strip()
        # Drop NCDs/bonds NSE tags as SME with digit-leading symbols.
        if not sym or not sym[0].isalpha():
            continue
        # Exclude non-IPO instruments that carry a stale/partial issue price and
        # would produce bogus "gains":
        #   - partly-paid shares  (symbol ends "PP"/"PP1"/…; e.g. ADANIENPP1)
        #   - NCD-linked records  (htmSym contains "ncd")
        #   - FPOs / follow-ons   (company name contains "FPO")
        htm = (r.get("htmSym") or "").lower()
        comp = (r.get("company") or "")
        if re.search(r"PP\d*$", sym) or "ncd" in htm or "fpo" in comp.lower():
            continue
        ip = parse_price(r.get("issuePrice", ""), r.get("priceRange", ""))
        if ip is None:
            continue
        out.append(IPO(
            company=comp.strip(),
            symbol=sym,
            listing_date=ld,
            issue_price=ip,
            segment="Mainboard" if sec == "EQ" else "NSE SME",
            exchange_suffix=".NS",
        ))
    # Dedupe repeated feed records (same ticker + listing date).
    seen: set[tuple[str, date]] = set()
    deduped: list[IPO] = []
    for i in out:
        k = (i.symbol, i.listing_date)
        if k in seen:
            continue
        seen.add(k)
        deduped.append(i)
    return deduped


# ───────────────────────── BSE live supplement ─────────────────────────────
def fetch_bse_public_issues() -> list[dict]:
    """BSE only ever exposes the ~20 issues open RIGHT NOW — there is no dated
    history endpoint (`Type` is ignored; p/P/past/C all return the same rows).
    So every run snapshots that window into the ledger, which is what actually
    accumulates BSE-SME-only history over time."""
    store = _ledger_load(BSE_LEDGER)
    try:
        s = requests.Session()
        s.headers.update(BSE_HEADERS)
        s.get("https://www.bseindia.com/", timeout=20)
        r = s.get("https://api.bseindia.com/BseIndiaAPI/api/GetPublicIssue/w?flag=&Type=",
                  timeout=25)
        rows = r.json().get("Table", [])
    except Exception as e:
        print(f"  [bse] live feed unavailable: {e}", file=sys.stderr)
        rows = []
    if rows:
        added, _ = _ledger_merge(store, rows, _bse_ledger_key)
        _ledger_save(BSE_LEDGER, store)
        print(f"  [bse] live window {len(rows)} row(s); ledger +{added} new — "
              f"{len(store)} total.")
    elif store:
        print(f"  [bse] serving ledger only — {len(store)} record(s).", file=sys.stderr)
    return list(store.values())


def collect_bse_recent(start: date, end: date, have_names: set[str]) -> list[IPO]:
    """Fold in BSE-only issues from the accumulated public-issue ledger. Coverage
    starts from the first run that captured a given issue: BSE publishes no
    back-history, so anything that closed before this ledger existed is absent."""
    out: list[IPO] = []
    master = bse_scrip_master()
    by_name = {_bse_name_key(v["name"]): (c, v["id"]) for c, v in master.items()
               if v["name"]}
    for row in fetch_bse_public_issues():
        # The feed is every public issue, not every IPO: a typical window is 8
        # IPOs among 24 rows, the rest rights issues, buybacks, offers-to-buy,
        # FPOs and social-stock-exchange ZCZP paper. Only IPOs belong here.
        if str(row.get("IR_FLAG_FULL") or "").strip().upper() != "IPO":
            continue
        name = (row.get("Scrip_Name") or "").strip()
        code = str(row.get("Scrip_cd") or "").strip()
        # End_Dt is the issue close; listing is ~3-6 days later. Use it only as a
        # coarse window gate — actual listing date is resolved from price data.
        end_dt = parse_date(row.get("End_Dt", ""))
        ip = parse_price("", row.get("Price_Band", ""))
        norm = re.sub(r"\s+", " ", name.lower()).replace(" limited", "").strip()
        if (not code or ip is None or end_dt is None
                or end_dt < start or end_dt > end
                or norm in have_names):
            continue
        # `Scrip_cd` here is the feed's own issue id (4707), NOT the BSE scrip
        # code (544856) — only the scrip master can bridge the two, and only
        # once the issue has actually listed.
        scrip, ticker = by_name.get(_bse_name_key(name), ("", ""))
        out.append(IPO(
            company=name,
            symbol=ticker or scrip or code,
            listing_date=end_dt,          # provisional; refined in fetch step
            issue_price=ip,
            segment="BSE (recent)",
            exchange_suffix=".BO",
            quote_symbol=f"{scrip}.BO" if scrip else "",
        ))
    return out


# ───────────────── BSE SME back-history (via IPOPlatform) ──────────────────
@lru_cache(maxsize=1)
def bse_scrip_master() -> dict:
    """BSE scrip code -> {'id': alpha ticker, 'isin': ...}.

    SME scrips are NOT in the Equity segment this file's other BSE calls use —
    that returns 4,975 rows and none of them. They sit in MTF (12,671 rows,
    groups M and MT), which is where a numeric code becomes a readable ticker:
    544568 -> ZAPPFRESH."""
    rows = cache_get("bse_scrip_master_mtf", ttl_seconds=7 * 24 * 3600)
    if rows is None:
        try:
            s = requests.Session()
            s.headers.update(BSE_HEADERS)
            s.get("https://www.bseindia.com/", timeout=20)
            r = s.get("https://api.bseindia.com/BseIndiaAPI/api/ListofScripData/w",
                      params={"Group": "", "Scripcode": "", "industry": "",
                              "segment": "MTF", "status": "Active"}, timeout=40)
            rows = r.json()
        except Exception as e:
            print(f"  [bse] scrip master unavailable: {e}", file=sys.stderr)
            rows = []
        if isinstance(rows, list) and rows:
            cache_set("bse_scrip_master_mtf", rows)
    out: dict[str, dict] = {}
    for row in rows if isinstance(rows, list) else []:
        code = str(row.get("SCRIP_CD") or "").strip()
        if code:
            out[code] = {"id": str(row.get("scrip_id") or "").strip(),
                         "isin": str(row.get("ISIN_NUMBER") or "").strip(),
                         "name": str(row.get("Scrip_Name") or "").strip()}
    return out


def _bse_name_key(s: str) -> str:
    """BSE writes 'Yajur Fibres Ltd' where IPOPlatform writes 'Yajur Fibres
    Limited', so the suffix has to go before the two can be compared."""
    return _norm_alnum(re.sub(r"\b(limited|ltd)\b\.?\s*$", "", (s or "").strip(),
                              flags=re.I))


def collect_bse_sme(start: date, end: date, have_names: set[str]) -> list[IPO]:
    """BSE-SME-only issues, with real listing dates and issue prices.

    Neither exchange can supply these: the NSE feed never sees a BSE-only
    listing, and BSE's own public-issue API serves only the window open right
    now. IPOPlatform's index does carry them — `ipl_index()` simply drops them,
    because it keys on `nse_script_symbol` and a BSE-only issue has none."""
    master = bse_scrip_master()
    by_isin = {v["isin"]: c for c, v in master.items() if v["isin"]}
    by_id = {v["id"].upper(): c for c, v in master.items() if v["id"]}
    by_name = {_bse_name_key(v["name"]): c for c, v in master.items() if v["name"]}
    out: list[IPO] = []
    for row in _ipl_index_rows():
        if str(row.get("nse_script_symbol") or "").strip():
            continue
        if not str(row.get("exchange") or "").upper().startswith("BSE"):
            continue
        # `ipo_year` holds the listing DATE despite its name — verified against
        # first-traded bars. There is no `listing_date` field, and
        # `listing_year` is an <a> blob ("Jan 2026"), not a year.
        ld = parse_date(str(row.get("ipo_year") or "")[:10])
        if ld is None or ld < start or ld > end:
            continue
        ip = (_ipl_num(row.get("offer_price"))
              or _ipl_num(row.get("upper_price_band")))
        if not ip:
            continue
        name = _html.unescape(str(row.get("company_name") or "")).strip()
        norm = re.sub(r"\s+", " ", name.lower()).replace(" limited", "").strip()
        if not name or norm in have_names:
            continue
        isin = str(row.get("isin") or "").strip()
        code = str(row.get("bse_script_code") or "").strip()
        if not code.isdigit():
            # A few rows carry a ticker or an ISIN where the code belongs, and a
            # couple carry nothing at all.
            code = (by_id.get(code.upper())
                    or by_isin.get(isin)
                    or by_name.get(_bse_name_key(name))
                    or "")
        ticker = master.get(code, {}).get("id")
        out.append(IPO(
            company=name,
            symbol=ticker or code or isin or name,
            listing_date=ld,
            issue_price=float(ip),
            segment="BSE SME",
            exchange_suffix=".BO",
            # Quoted on the numeric code; the ticker is for reading only.
            quote_symbol=f"{code}.BO" if code else "",
        ))
    return out


# ─────────────────────────── price maths ───────────────────────────────────
def _download(symbol: str, start: date, end: date) -> pd.DataFrame | None:
    try:
        import data_provider  # type: ignore
        df = data_provider.download(symbol, start.strftime("%Y-%m-%d"),
                                    end.strftime("%Y-%m-%d"))
        if df is not None and not df.empty and "Close" in df.columns:
            return df
    except Exception:
        pass
    return None


def compute_returns(ipo: IPO, window_days: int) -> IPO:
    quote = ipo.quote_symbol or ipo.symbol
    key = f"ret_v2_{quote}_{ipo.listing_date:%Y%m%d}_{window_days}"
    cached = cache_get(key, ttl_seconds=14 * 24 * 3600)
    if cached is not None:
        ipo.listing_close = cached.get("lc")
        ipo.listing_pct = cached.get("lp")
        ipo.peak_close = cached.get("pc")
        ipo.peak_date = parse_date(cached.get("pd") or "")
        ipo.peak_pct = cached.get("pp")
        ipo.note = cached.get("note", "")
        return ipo

    reported = ipo.listing_date
    start = reported - timedelta(days=12)
    end = reported + timedelta(days=window_days + 15)
    df = _download(quote, start, end)

    # yfinance fallback (handles BSE-only via .BO and NSE gaps).
    if df is None or df.empty:
        try:
            import yfinance as yf  # type: ignore
            yf_sym = (quote if quote.endswith((".NS", ".BO"))
                      else quote + ipo.exchange_suffix)
            hist = yf.Ticker(yf_sym).history(
                start=start.strftime("%Y-%m-%d"), end=end.strftime("%Y-%m-%d"))
            if hist is not None and not hist.empty:
                hist = hist.rename(columns=str.title)
                df = hist[["Open", "High", "Low", "Close", "Volume"]]
        except Exception:
            df = None

    def _save():
        cache_set(key, {"lc": ipo.listing_close, "lp": ipo.listing_pct,
                        "pc": ipo.peak_close,
                        "pd": str(ipo.peak_date) if ipo.peak_date else None,
                        "pp": ipo.peak_pct, "note": ipo.note})

    if df is None or df.empty or "Close" not in df.columns:
        ipo.note = "no price data"
        _save()
        return ipo

    idx = pd.to_datetime(df.index)
    df = df.copy()
    df.index = idx.date

    # Reject non-IPOs: a genuine IPO has NO trading before its listing date.
    # Rights issues / FPOs / re-listings of already-listed companies (e.g. an
    # Adani rights record carrying a stale issue price) show pre-listing bars
    # and would otherwise produce bogus "gains" — exclude them.
    pre = [d for d in df.index if d < reported]
    if len(pre) >= 3:
        ipo.note = "already trading before listing date (rights/FPO/re-listing) — excluded"
        _save()
        return ipo

    # Listing-day bar = first bar on/after the reported listing date.
    on_after = [d for d in df.index if d >= reported]
    if not on_after:
        ipo.note = "no bar at/after listing date"
        _save()
        return ipo
    first_day = min(on_after)
    # Refine listing_date to the actual first-traded date (BSE provisional case).
    ipo.listing_date = first_day
    ipo.listing_close = float(df.loc[first_day, "Close"])
    ipo.listing_pct = (ipo.listing_close - ipo.issue_price) / ipo.issue_price * 100.0

    # 30-calendar-day peak CLOSE (inclusive of listing day).
    win_end = first_day + timedelta(days=window_days)
    mask = [(d >= first_day and d <= win_end) for d in df.index]
    win = df[mask]
    if not win.empty:
        pk = win["Close"].astype(float)
        ipo.peak_close = float(pk.max())
        ipo.peak_date = list(win.index)[int(pk.values.argmax())]
        ipo.peak_pct = (ipo.peak_close - ipo.issue_price) / ipo.issue_price * 100.0

    _save()
    return ipo


# ───────────────── anchor investors (parsed from local PDFs) ────────────────
# The COMPLETE anchor-investor list for an IPO lives only in the official
# "Allocation to Anchor Investors" filing (a PDF filed with BSE/NSE the day
# before the issue opens). Free web sources expose only the first ~2 names, so
# to guarantee a complete list nothing-missed, drop those PDFs into `anchor_dir`
# (default ./anchor_pdfs) and this step parses every name out of them.
ANCHOR_DIR_DEFAULT = SCRIPT_DIR / "anchor_pdfs"


def _norm_alnum(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def _company_core(company: str) -> str:
    """Company name reduced to alnum, with common corporate suffixes dropped —
    used to match a PDF (by filename or content) back to a symbol."""
    c = (company or "").lower()
    for tail in (" limited", " ltd", " private limited", " pvt ltd", " pvt. ltd.",
                 " (india)", " india"):
        c = c.replace(tail, " ")
    return re.sub(r"[^a-z0-9]", "", c)


# ── automatic download of the official filing from NSE ─────────────────────
# NSE publishes each issue's anchor letter itself, so the filings do NOT have to
# be collected by hand. Two facts make this reliable:
#
#   1. The archive URL is deterministic — ANCHOR_<SYMBOL>.zip — and it is
#      RETAINED after listing, so the whole back-catalogue can be backfilled in
#      one pass, not just the issue that is currently open.
#   2. Its presence is the "anchor book is final" signal. The row appears on
#      anchor day (T-1 before the issue opens) and not before, so polling the
#      upcoming-issues list and retrying is all the scheduling that is needed.
#
# A 404 is therefore ambiguous, and the difference matters for a script whose
# whole purpose is completeness: it means EITHER "not published yet" OR "this
# issue had no anchor book at all" (legal for SME issues, which may allocate
# only to a market maker). ipo-detail settles it — an issue with no anchor
# portion never mentions the word — so a 404 is only ever reported after that
# check, never guessed at.
NSE_ANCHOR_ZIP = "https://nsearchives.nseindia.com/content/ipo/ANCHOR_{sym}.zip"
NSE_IPO_DETAIL = "https://www.nseindia.com/api/ipo-detail?symbol={sym}&series={ser}"
NSE_UPCOMING_IPO = "https://www.nseindia.com/api/all-upcoming-issues?category=ipo"
NSE_CURRENT_IPO = "https://www.nseindia.com/api/ipo-current-issue"
# A signed, scanned anchor letter runs to a few MB; anything far beyond that is
# not the document we asked for and is refused rather than written to disk.
MAX_ANCHOR_ZIP_BYTES = 40 * 1024 * 1024


def _series_for(segment: str) -> str:
    """NSE's series code for an IPO's segment, as ipo-detail expects it."""
    return "SME" if "sme" in (segment or "").lower() else "EQ"


def fetch_ipo_detail(session: requests.Session, symbol: str,
                     series: str) -> dict:
    """NSE's per-issue detail blob, or {} if the symbol/series pair is unknown."""
    try:
        r = session.get(NSE_IPO_DETAIL.format(sym=symbol, ser=series), timeout=30)
        if r.status_code != 200:
            return {}
        d = r.json()
        return d if isinstance(d, dict) else {}
    except Exception:
        return {}


def _issue_has_anchor_portion(detail: dict) -> bool:
    """Did this issue reserve an anchor portion at all? SME issues need not."""
    return "anchor" in json.dumps(detail or {}).lower()


def download_anchor_filing(session: requests.Session, symbol: str,
                           anchor_dir: Path,
                           dest_stem: str | None = None) -> Path | None:
    """Download and unpack NSE's anchor allocation report for `symbol`.

    The zip holds exactly one PDF, but its inner filename is whatever the issuer
    happened to upload ('Company Anchor.pdf', 'intimation.pdf',
    'Anchor Investors_Company.docx.pdf', '...LETTER.PDF'), so ANY pdf member is
    taken rather than a fixed name. The bytes are read out of the archive and
    written to a path we choose, so a crafted member name cannot escape
    `anchor_dir`. `dest_stem` overrides the filename stem (default: symbol)."""
    url = NSE_ANCHOR_ZIP.format(sym=symbol)
    try:
        r = session.get(url, timeout=90)
    except Exception as e:
        print(f"  [fetch] {symbol}: {type(e).__name__}: {e}", file=sys.stderr)
        return None
    if r.status_code != 200 or not r.content.startswith(b"PK"):
        return None
    if len(r.content) > MAX_ANCHOR_ZIP_BYTES:
        print(f"  [fetch] {symbol}: archive is {len(r.content):,} bytes, refusing.",
              file=sys.stderr)
        return None
    try:
        z = zipfile.ZipFile(io.BytesIO(r.content))
        member = next((n for n in z.namelist() if n.lower().endswith(".pdf")), None)
        if member is None:
            print(f"  [fetch] {symbol}: archive holds no PDF ({z.namelist()}).",
                  file=sys.stderr)
            return None
        data = z.read(member)
    except Exception as e:
        print(f"  [fetch] {symbol}: unreadable archive ({e}).", file=sys.stderr)
        return None
    anchor_dir.mkdir(parents=True, exist_ok=True)
    dest = anchor_dir / f"{dest_stem or symbol}.pdf"
    dest.write_bytes(data)
    return dest


# ── BSE notices: the same letter, but born-digital ─────────────────────────
# NSE republishes whatever the issuer uploaded, and for most mainboard issues
# that is a photocopy — 9 of our 13 mainboard filings carry no text layer at
# all. BSE files the SAME anchor intimation as a numbered notice, "Public Issue
# of <COMPANY> - Allocation to Anchor Investors", and its attachment is the
# issuer's own born-digital document: real text, real table lines, names spelt
# correctly. Preferring it removes the OCR step entirely.
#
# The attachment path needs no guessing. The 2-page cover notice at
# /Notices/<no>/<no>.pdf states the anchor share count AND prints the
# attachment's own path, guid included, in its text.
#
# LIMIT: the notices search only reaches back about five months (nothing before
# 2026-03-16 when probed on 2026-08-07), so this helps issues from 2026 on and
# cannot repair the older backlog.
BSE_NOTICES_CSV = ("https://api.bseindia.com/BseIndiaAPI/api/"
                   "NoticesCircularDownloadcsv_New/w")
BSE_ANCHOR_SUBJECT = "Allocation to Anchor Investors"
BSE_NOTICE_PDF = "https://www.bseindia.com/downloads/UploadDocs/Notices/{no}/{no}.pdf"
_BSE_ATTACH_RE = re.compile(r"(Attach/[^\s\"'<>]+?\$[0-9a-fA-F-]{36}\.pdf)")
_CO_SUFFIX = {"LIMITED", "LTD", "PRIVATE", "PVT", "THE", "COMPANY", "CO",
              "INDIA", "AND", "OF"}


def _co_tokens(name: str) -> set[str]:
    """Significant words of a company name, legal wrapper words removed."""
    w = re.sub(r"[^A-Z0-9]+", " ", str(name or "").upper()).split()
    return {t for t in w if t not in _CO_SUFFIX and len(t) > 1}


@lru_cache(maxsize=4)
def bse_anchor_notices(days_back: int = 400) -> tuple:
    """Every 'Allocation to Anchor Investors' notice BSE still lists, as
    (notice_no, subject, company_tokens). Empty on any failure."""
    end = date.today()
    start = end - timedelta(days=days_back)
    params = {"strTxtNoticeNo": "", "strTxtDate": start.strftime("%d/%m/%Y"),
              "strTxtTodate": end.strftime("%d/%m/%Y"), "strScripcode": "",
              "strDep": "", "strSegment": "", "subject": BSE_ANCHOR_SUBJECT,
              "category": "", "containgtext": "", "str": "0"}
    try:
        r = requests.get(BSE_NOTICES_CSV, params=params, timeout=60,
                         headers={"User-Agent": UA,
                                  "Referer": "https://www.bseindia.com/"})
        if r.status_code != 200:
            return ()
        rdr = csv.DictReader(io.StringIO(r.text))
        out = []
        for row in rdr:
            no = (row.get("Notice No") or "").strip()
            sub = (row.get("Subject") or "").strip()
            m = re.search(r"Public Issue of (.+?)\s*-\s*Allocation", sub, re.I)
            if no and m:
                out.append((no, sub, frozenset(_co_tokens(m.group(1)))))
        return tuple(out)
    except Exception as e:
        print(f"  [bse] notices unavailable ({e}).", file=sys.stderr)
        return ()


def download_bse_anchor_filing(company: str, symbol: str,
                               anchor_dir: Path) -> Path | None:
    """Fetch BSE's anchor notice attachment for `company`, or None.

    Matched on company-name tokens because the notice carries no symbol. The
    match must be near-total (every significant word of one name present in
    the other) — a loose match here would file one issuer's anchor book under
    another issuer's symbol, which is worse than having no file at all."""
    want = _co_tokens(company)
    if not want:
        return None
    hit = None
    for no, sub, toks in bse_anchor_notices():
        if not toks:
            continue
        if want <= set(toks) or set(toks) <= want:
            hit = (no, sub)
            break
    # Fuzzy fallback: abbreviation mismatches ("FutureTek" vs "Future Tek")
    # break the exact token-subset test. SequenceMatcher on the joined tokens
    # catches these while staying strict enough to avoid cross-company matches.
    if hit is None:
        import difflib
        want_s = " ".join(sorted(want))
        best_ratio, best_hit = 0.0, None
        for no, sub, toks in bse_anchor_notices():
            if not toks:
                continue
            ratio = difflib.SequenceMatcher(
                None, want_s, " ".join(sorted(toks))).ratio()
            if ratio >= 0.85 and ratio > best_ratio:
                best_ratio, best_hit = ratio, (no, sub)
        hit = best_hit
    if hit is None:
        return None
    no, sub = hit
    try:
        r = requests.get(BSE_NOTICE_PDF.format(no=no), timeout=60,
                         headers={"User-Agent": UA,
                                  "Referer": "https://www.bseindia.com/"})
        if r.status_code != 200 or not r.content.startswith(b"%PDF"):
            return None
        # The attachment path is a LINK ANNOTATION, not page text: the visible
        # text only names the file ("Anchor_Intimation_Letter_22_07_2026_final")
        # and omits the guid. Read it off the raw object instead.
        m = _BSE_ATTACH_RE.search(r.content.decode("latin-1", "ignore"))
        if not m:
            print(f"  [bse] {symbol}: notice {no} names no attachment.",
                  file=sys.stderr)
            return None
        url = ("https://www.bseindia.com/downloads/UploadDocs/Notices/"
               + m.group(1))
        a = requests.get(url, timeout=90,
                         headers={"User-Agent": UA,
                                  "Referer": "https://www.bseindia.com/"})
        if a.status_code != 200 or not a.content.startswith(b"%PDF"):
            return None
        if len(a.content) > MAX_ANCHOR_ZIP_BYTES:
            return None
    except Exception as e:
        print(f"  [bse] {symbol}: {type(e).__name__}: {e}", file=sys.stderr)
        return None
    anchor_dir.mkdir(parents=True, exist_ok=True)
    dest = anchor_dir / f"{symbol}.pdf"
    dest.write_bytes(a.content)
    print(f"  [bse] {symbol}: took the born-digital attachment from notice {no}.")
    return dest


def fetch_upcoming_issues(session: requests.Session) -> list[tuple[str, str, str]]:
    """(symbol, series, company) for every issue NSE lists as open or coming.

    category=sme returns an empty payload — SME issues arrive through the same
    'ipo' category carrying series 'SME' — so only the one endpoint is polled."""
    out: dict[str, tuple[str, str, str]] = {}
    for url in (NSE_UPCOMING_IPO, NSE_CURRENT_IPO):
        try:
            r = session.get(url, timeout=30)
            rows = r.json() if r.status_code == 200 else []
        except Exception:
            rows = []
        if not isinstance(rows, list):
            continue
        for row in rows:
            sym = str(row.get("symbol") or "").strip().upper()
            if sym:
                out[sym] = (sym, str(row.get("series") or "EQ").strip().upper(),
                            str(row.get("companyName") or "").strip())
    return sorted(out.values())


def sync_anchor_filings(anchor_dir: Path, winners: list[IPO],
                        include_upcoming: bool = True) -> None:
    """Pull every missing anchor filing into `anchor_dir`, BSE first.

    BSE's copy is born-digital (real text layer, no OCR needed) and covers
    ~13 months of filings. NSE's scanned copy is used as a fallback for older
    IPOs or when BSE matching fails. Covers both directions: the back-catalogue
    (every winner that has no local filing yet) and the front edge (issues open
    or announced right now, whose letters land the day before bidding opens).
    Symbols whose PDF is already on disk are never re-downloaded, so re-runs
    cost nothing and hand-placed files are left alone."""
    anchor_dir.mkdir(parents=True, exist_ok=True)
    have: set[str] = set()
    for p in anchor_dir.glob("*.pdf"):
        s = p.stem.upper()
        have.add(s)
        for sfx in ("-NSE", "-BSE"):
            if s.endswith(sfx):
                have.add(s[: -len(sfx)])

    targets: list[tuple[str, str, str]] = [
        (w.symbol.upper(), _series_for(w.segment), w.company) for w in winners]
    session = nse_session()
    if include_upcoming:
        try:
            upcoming = fetch_upcoming_issues(session)
        except Exception as e:
            upcoming = []
            print(f"  [fetch] upcoming-issues feed unavailable ({e}).",
                  file=sys.stderr)
        seen = {t[0] for t in targets}
        for sym, ser, comp in upcoming:
            if sym not in seen:
                targets.append((sym, ser, comp))
        if upcoming:
            print(f"  [fetch] {len(upcoming)} issue(s) open or forthcoming: "
                  f"{', '.join(s for s, _, _ in upcoming)}")

    todo = [t for t in targets if t[0] not in have]
    if not todo:
        print(f"  [fetch] all {len(targets)} symbol(s) already have a local "
              "filing; nothing to download.")
        return
    print(f"  [fetch] {len(todo)} symbol(s) without a local filing; "
          "downloading (BSE preferred, NSE fallback) ...")

    got: list[str] = []
    no_book: list[str] = []
    pending: list[str] = []
    from_bse: list[str] = []
    from_nse: list[str] = []
    for sym, ser, _comp in todo:
        # BSE first: its copy is born-digital (real text, no OCR needed).
        if _comp and download_bse_anchor_filing(_comp, sym, anchor_dir):
            from_bse.append(sym)
            # Cross-filing: also grab the NSE scan as a companion so two
            # independent readings can be merged in collect_anchor_investors.
            download_anchor_filing(session, sym, anchor_dir,
                                   dest_stem=f"{sym}-nse")
            continue
        # BSE unavailable (older IPO or no match). Fall back to NSE scan.
        path = download_anchor_filing(session, sym, anchor_dir)
        if path is not None:
            from_nse.append(sym)
            continue
        # Still nothing. Ask the issue's own detail record whether an anchor
        # portion exists, so "nothing to miss" is never reported as a gap.
        detail = fetch_ipo_detail(session, sym, ser)
        if not detail:
            detail = fetch_ipo_detail(session, sym, "SME" if ser == "EQ" else "EQ")
        (pending if _issue_has_anchor_portion(detail) else no_book).append(sym)

    if from_bse:
        print(f"  [fetch] {len(from_bse)} from BSE notices (born-digital, no OCR "
              f"needed): {', '.join(from_bse)}")
    if from_nse:
        print(f"  [fetch] {len(from_nse)} from NSE (scanned): {', '.join(from_nse)}")
    if no_book:
        print(f"  [fetch] {len(no_book)} issue(s) reserved NO anchor portion, so "
              f"there is nothing to collect: {', '.join(no_book)}")
    if pending:
        print(f"  [fetch] {len(pending)} issue(s) have an anchor portion but NSE "
              f"has not published the report yet: {', '.join(pending)}\n"
              "          It appears the day before bidding opens — re-run then.")


# ── OCR for scanned filings ────────────────────────────────────────────────
# Many anchor letters are scans (photocopied, signed, stamped), so the PDF holds
# page IMAGES with no text layer and pdfplumber returns nothing at all. Those
# are rendered and OCR'd with the poppler + tesseract CLIs
# (brew install poppler tesseract).
#
# OCR accuracy is settings-dependent and a single pass CAN SILENTLY DROP a table
# row, so several passes are tried and each is reconciled against the letter's
# OWN stated total ("...finalized allocation of 14,99,600 Equity Shares..."):
# the per-row share numbers must sum to it. The first pass that reconciles wins.
# If none reconciles, the shortfall is reported loudly instead of returning a
# quietly incomplete list — a missing anchor name is worse than a visible error.
IMAGE_SUFFIXES = (".png", ".jpg", ".jpeg", ".tif", ".tiff", ".webp", ".bmp")
OCR_PASSES = ((300, "6"), (450, "4"), (450, "6"), (600, "4"))


def _to_int(s: str):
    try:
        return int(re.sub(r"[^\d]", "", s))
    except Exception:
        return None


# Numeric cells span three shapes on one row: a lakh-grouped share count
# ("37,200"), a percentage ("1.98") and a lakh-grouped amount ("1,00,44,000").
_NUM_RE = re.compile(r"\d[\d,.]*\d|\d")
# A scan routinely turns some of a grouped number's commas into full stops, so
# "14,70,686" comes back as "14,70.686". Read naively that is 1470.686 — not a
# whole number, so the share column is skipped and the row silently takes some
# other figure. A token carrying BOTH separators whose last group is three
# digits is a grouped integer, never a decimal: real decimals here are
# percentages and prices, which have two decimal places and no comma.
_MIXED_GROUPED_RE = re.compile(r"^\d{1,3}(?:[.,]\d{2,3})*[.,]\d{3}$")


def _to_num(s: str):
    t = (s or "").strip()
    if len(re.findall(r"[.,]", t)) >= 2 and _MIXED_GROUPED_RE.match(t):
        return float(re.sub(r"[^\d]", "", t))
    try:
        return float(re.sub(r"[^\d.]", "", t))
    except Exception:
        return None


def _numbers_in(text: str) -> list[float]:
    vals = (_to_num(t) for t in _NUM_RE.findall(text or ""))
    return [v for v in vals if v is not None]


# A leading "6." or "| 10, |" is the table's SERIAL number and must be dropped;
# a leading "360" is part of the investor's NAME ("360 ONE LVF Treasury
# Solutions Fund") and must be kept. Serials are 1-2 digits AND carry a trailing
# delimiter — that delimiter is the only thing that reliably separates the two.
_SERIAL_RE = re.compile(r"^\s*\d{1,2}(?!\d)\s*[.),\-]\s*")

# SEBI's minimum anchor ticket (Rs 1 crore) means a real allocation is always
# thousands of shares, so the share column is the first number on the row worth
# at least this. Taking the row's FIRST number instead silently mis-reads any
# investor whose name opens with a digit: "360 ONE LVF ... 74,400" was booked as
# 360 shares, losing 74,040 and failing the reconciliation for the whole filing.
MIN_ANCHOR_SHARES = 1000


def _split_row(core: str):
    """Split one table row into (name, allocated shares, trailing numbers).

    The share column is the first number on the row worth at least
    MIN_ANCHOR_SHARES; the name is everything BEFORE it and the remaining
    numeric columns (percentage, bid price, amount) are everything after, which
    is what the price/amount columns are recovered from.
    Cutting at that position beats matching a trailing run of numbers, which any
    stray OCR character defeats — "LRSD Securities Pvt Ltd 37,200 (1.98 270 ~——
    _'1,00,44,000" left the whole numeric tail glued to the name. It also keeps
    digits that are genuinely part of a name ("...Trust-Scheme 1"), because
    those are below the threshold and are simply skipped over."""
    for m in _NUM_RE.finditer(core):
        v = _to_num(m.group(0))
        if v is not None and v >= MIN_ANCHOR_SHARES and v.is_integer():
            return (core[:m.start()].strip(" .,:;-|"), int(v),
                    _numbers_in(core[m.end():]))
    return core.strip(" .,:;-|"), None, []


def _is_name_fragment(s: str) -> bool:
    """Is this line the wrapped remainder of the previous investor's name, or is
    it debris shed by the table's ruling lines? Debris is short, speckled with
    stray digits and seldom carries two real words ("ith 7 es"), whereas a true
    continuation reads like part of a name ("Growth Fund", "Emerging Star
    Fund")."""
    if not s or len(s) > 60 or _is_prose(s):
        return False
    toks = s.split()
    words = [t for t in toks if len(t) >= 3 and t.isalpha()]
    has_digit = any(any(c.isdigit() for c in t) for t in toks)
    return len(words) >= 2 or (len(words) == 1 and not has_digit)


def _ocr_available() -> bool:
    return bool(shutil.which("pdftoppm") and shutil.which("tesseract"))


OCR_CACHE_DIR = CACHE_DIR / "ocr"
RENDER_CACHE_DIR = CACHE_DIR / "render"
ANCHOR_RESULT_DIR = CACHE_DIR / "anchors"

# Bumped only when the IMAGE itself changes — dpi handling, deskewing,
# enhancement. A change to how that image is READ (psm, or any parser) must
# NOT bump this, or every page is re-rendered to produce a byte-identical file.
OCR_RENDER_VERSION = "r1"

_RENDER_CACHE: dict[str, list[Path]] = {}


def _render_pages(path: Path, dpi: int, enhance: bool) -> list[Path]:
    """Render a PDF to page images via pdftoppm, caching across readers.

    The images are kept on disk BETWEEN runs. Rendering is the most expensive
    step in the pipeline and its result depends only on the file and the render
    settings, so a change to how a page is read reuses them instead of paying
    pdftoppm again for the same picture."""
    sha = _file_sha(path)
    rkey = f"{sha}-{dpi}-{'enh' if enhance else 'raw'}-{OCR_RENDER_VERSION}"
    if rkey in _RENDER_CACHE:
        return _RENDER_CACHE[rkey]
    final = RENDER_CACHE_DIR / rkey
    if final.is_dir():
        cached = sorted(final.glob("pg-*.png"))
        if cached:
            _RENDER_CACHE[rkey] = cached
            return cached
    # Built in a staging directory and moved into place only once every page is
    # written, so a run killed mid-render cannot leave a half-filled directory
    # that the next run would happily read as a complete rendering.
    staging = RENDER_CACHE_DIR / f"{rkey}.part"
    shutil.rmtree(staging, ignore_errors=True)
    staging.mkdir(parents=True, exist_ok=True)
    try:
        subprocess.run(["pdftoppm", "-r", str(dpi), "-png", str(path),
                         str(staging / "raw")],
                       capture_output=True, timeout=900, check=True)
    except Exception:
        shutil.rmtree(staging, ignore_errors=True)
        _RENDER_CACHE[rkey] = []
        return []
    # Materialised before _ocr_ready runs: it writes its rotated and enhanced
    # output beside the source, which would otherwise be swept into the glob.
    raw_pages = sorted(staging.glob("raw-*.png"))
    for n, src in enumerate(raw_pages, 1):
        _ocr_ready(src, enhance, dpi).replace(staging / f"pg-{n:04d}.png")
    for leftover in staging.glob("raw-*.png"):
        leftover.unlink()
    shutil.rmtree(final, ignore_errors=True)
    staging.rename(final)
    pages = sorted(final.glob("pg-*.png"))
    _RENDER_CACHE[rkey] = pages
    return pages


# Filing name -> did this run's reading sum EXACTLY to the total the letter
# declares? That is proof of completeness, not a quality score: no investor can
# be missing from a table whose shares add up. It is what entitles a fresh
# reading to REPLACE the rows already in the workbook instead of being merged
# with them, which is the only way the debris of earlier runs ever leaves.
_RECONCILED_FILINGS: dict[str, bool] = {}


def _seal(path: Path, anchors: list, reconciled: bool) -> list:
    """Record whether `anchors` is a proven-complete reading of `path`."""
    _RECONCILED_FILINGS[path.name] = reconciled
    return anchors


# Bumped ONLY if the Anchor record gains or loses a field. Deliberately not the
# OCR version: a reading whose shares sum exactly to the declared total is
# complete, and no future change to how a page is read can improve on complete,
# so freezing it on the OCR version would throw the proof away for nothing.
ANCHOR_RESULT_SCHEMA = "a1"
_FROZEN_DISABLED = False


def _frozen_path(path: Path) -> Path:
    return ANCHOR_RESULT_DIR / f"{_file_sha(path)}-{ANCHOR_RESULT_SCHEMA}.json"


def _frozen_load(path: Path) -> "list[Anchor] | None":
    """The stored reading of `path`, if one was ever proven complete."""
    if _FROZEN_DISABLED:
        return None
    fp = _frozen_path(path)
    if not fp.exists():
        return None
    try:
        rows = json.loads(fp.read_text(encoding="utf-8"))
        return [Anchor(**r) for r in rows] or None
    except Exception:
        return None


def _frozen_save(path: Path, anchors: list) -> None:
    try:
        ANCHOR_RESULT_DIR.mkdir(parents=True, exist_ok=True)
        _frozen_path(path).write_text(
            json.dumps([asdict(a) for a in anchors]), encoding="utf-8")
    except Exception:
        pass

# Bumped whenever the IMAGE handed to tesseract changes, so that a cached
# reading taken from the old image is not silently reused. The cache is keyed
# on the PDF's content hash, which cannot notice a change in our own pipeline.
OCR_PIPELINE_VERSION = "p9"


# Minimum tesseract OSD "Orientation confidence" before we will turn a page.
# Genuine rotations in this corpus score 5-11; the false 180-degree calls that
# mirrored whole pages scored 0.06 and 0.83.
OSD_MIN_CONFIDENCE = 2.0


def _file_sha(path: Path) -> str:
    h = hashlib.sha1()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()[:16]


def _ocr_key(path: Path, dpi: int, psm: str, enhance: bool) -> str:
    """Cache key for one reading. The enhanced and unenhanced renderings of the
    same page are different readings and must never share a cache entry."""
    return (f"{_file_sha(path)}-{dpi}-{psm}-{OCR_PIPELINE_VERSION}"
            + ("-enh" if enhance else ""))


def _tesseract(img: Path, psm: str) -> str:
    try:
        r = subprocess.run(["tesseract", str(img), "-", "--oem", "1", "--psm", psm],
                           capture_output=True, text=True, timeout=300)
        return r.stdout or ""
    except Exception:
        return ""


def _deskew_page(img: Path) -> Path:
    """Rotate a sideways-scanned page upright, returning the image to OCR.

    Some letters are scanned in landscape (ANYA's whole filing is on its side).
    Tesseract does NOT auto-rotate, and on a 90-degree page it returns fluent-
    looking gibberish — "Gees / oe:ee / w — DLS" — rather than failing, so the
    page reads as an unparseable scan instead of an obviously rotated one.
    Orientation is detected with tesseract's own OSD mode; anything it cannot
    call is left exactly as rendered.

    OSD IS ONLY BELIEVED WHEN IT IS CONFIDENT. It reports a confidence score
    beside the angle, and on a dense ruled table that score collapses: CPPLUS
    renders identically at every DPI, yet at 450dpi tesseract called pages 2 and
    4 upside down at confidence 0.83 and 0.06 while calling them upright at 300
    and 600dpi at confidence 2.17 and 3.17. Obeying those two calls rotated the
    pages 180 degrees and every investor came back mirrored — "GNN4
    ONINNLOVANNVI IVWMSO IVILLOW" for MOTILAL OSWAL MANUFACTURING FUND. Real
    rotations in this corpus score 5 to 11, so a floor of 2.0 keeps them all
    and rejects the coin-flips. A page left the wrong way up is recoverable
    from the other passes; a page turned the wrong way up by us is not."""
    try:
        r = subprocess.run(["tesseract", str(img), "-", "--psm", "0"],
                           capture_output=True, text=True, timeout=120)
        out = r.stdout or ""
        m = re.search(r"^Rotate:\s*(\d+)", out, re.M)
        deg = int(m.group(1)) if m else 0
        if deg % 360 == 0:
            return img
        c = re.search(r"Orientation confidence:\s*([\d.]+)", out)
        if not c or float(c.group(1)) < OSD_MIN_CONFIDENCE:
            return img
        from PIL import Image
        out = img.with_name(img.stem + "-rot.png")
        Image.open(img).rotate(-deg, expand=True).save(out)
        return out
    except Exception:
        return img


def _remove_table_rules(a):
    """Erase long horizontal and vertical printed rules from a binarised page.

    Table borders touching glyph strokes cause tesseract to misread characters
    at cell edges.  Morphological opening isolates strokes longer than any glyph
    (kernel >= 1/15 of the page dimension), and a small dilation catches edge
    pixels that partially overlap adjacent letters.  The detected rule pixels are
    set to white (255) in the original array.

    Returns *a* unchanged on any failure or missing dependency."""
    import numpy as np
    h, w = a.shape[:2]
    min_h = max(h // 15, 20)
    min_w = max(w // 15, 20)
    ink = 255 - a
    try:
        import cv2
        h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (min_w, 1))
        v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, min_h))
        h_lines = cv2.morphologyEx(ink, cv2.MORPH_OPEN, h_kernel)
        v_lines = cv2.morphologyEx(ink, cv2.MORPH_OPEN, v_kernel)
        dilate_k = np.ones((3, 3), np.uint8)
        h_lines = cv2.dilate(h_lines, dilate_k, iterations=1)
        v_lines = cv2.dilate(v_lines, dilate_k, iterations=1)
        rules = h_lines | v_lines
        a[rules > 0] = 255
        return a
    except ImportError:
        pass
    try:
        from scipy.ndimage import binary_opening, binary_dilation
        mask = ink > 0
        h_struct = np.zeros((1, min_w), dtype=bool)
        h_struct[0, :] = True
        v_struct = np.zeros((min_h, 1), dtype=bool)
        v_struct[:, 0] = True
        h_lines = binary_opening(mask, structure=h_struct)
        v_lines = binary_opening(mask, structure=v_struct)
        dilate_s = np.ones((3, 3), dtype=bool)
        h_lines = binary_dilation(h_lines, structure=dilate_s)
        v_lines = binary_dilation(v_lines, structure=dilate_s)
        a[h_lines | v_lines] = 255
        return a
    except ImportError:
        pass
    return a


def _enhance_page(img: Path, dpi: int = 300) -> Path:
    """Clean a scanned page before OCR, the way Acrobat does on 'Recognize Text'.

    These filings are photocopies of faxes: grey and speckled. Tesseract was
    being handed the raw pdftoppm render, and a speck welded onto a letter costs
    a character that no downstream name-matching can recover.

    Pipeline (when cv2 is available):
      1. despeckle with a median filter (drops isolated specks without
         eroding strokes the way a blur would);
      2. CLAHE contrast normalisation (per-tile, so faint text in dark or
         unevenly lit regions is brought to ink-level before thresholding);
      3. adaptive Gaussian threshold (local 51×51 block instead of one global
         Otsu split — handles uneven illumination across the page);
      4. connected-component speck removal (clusters < 36 px are noise, not
         any glyph at 300 DPI).

    Falls back to Otsu when cv2 is absent.

    DELIBERATELY DOES NOT DESKEW. An earlier version corrected the half-degree
    slant of a hand-fed page here, which looked like an obvious win and was not:
    _page_grid_cells already deskews for itself via _skew_angle, and rotating
    twice — the second time on an image whose greys had been flattened to pure
    black and white — threw that estimate far enough off to turn CPPLUS upside
    down. Its investors came back as "GNN4 ONINNLOVANNVI IVWMSO IVILLOW", which
    is MOTILAL OSWAL MANUFACTURING FUND read at 180 degrees, and the filing fell
    from 49 names to 22. Rotation belongs to the reader that measures it.

    Returns the original path untouched on any failure — a missing Pillow or an
    odd image must degrade to the previous behaviour, never break the run."""
    try:
        import numpy as np
        from PIL import Image, ImageFilter, ImageOps

        scale = dpi / 300
        median_k = max(3, int(3 * scale)) | 1
        im = ImageOps.grayscale(Image.open(img)).filter(
            ImageFilter.MedianFilter(size=median_k))
        a = np.asarray(im, dtype=np.uint8).copy()

        try:
            import cv2
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
            a = clahe.apply(a)
            block = max(3, int(51 * scale)) | 1
            binarised = cv2.adaptiveThreshold(
                a, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                cv2.THRESH_BINARY, blockSize=block, C=10)
            speck_area = int(36 * scale * scale)
            n, labels, stats, _ = cv2.connectedComponentsWithStats(
                255 - binarised, connectivity=8)
            for i in range(1, n):
                if stats[i, cv2.CC_STAT_AREA] < speck_area:
                    binarised[labels == i] = 255
        except ImportError:
            # Fallback: Otsu binarisation (identical to previous behaviour).
            hist = np.bincount(a.ravel(), minlength=256).astype(np.float64)
            tot = float(hist.sum())
            if tot <= 0:
                return img
            w0 = np.cumsum(hist)
            w1 = tot - w0
            m0 = np.cumsum(hist * np.arange(256, dtype=np.float64))
            with np.errstate(divide="ignore", invalid="ignore"):
                between = (m0[-1] * w0 / tot - m0) ** 2 / (w0 * w1)
            between[~np.isfinite(between)] = 0.0
            thr = int(np.argmax(between))
            binarised = ((a > thr) * 255).astype(np.uint8)

        try:
            binarised = _remove_table_rules(binarised)
        except Exception:
            pass
        out = img.with_name(img.stem + "-enh.png")
        Image.fromarray(binarised).save(out)
        return out
    except Exception:
        return img


def _ocr_ready(img: Path, enhance: bool, dpi: int = 300) -> Path:
    """The image tesseract should actually read: upright, and optionally cleaned.

    Cleaning is OFFERED, never imposed. No single rendering is best for every
    filing: binarising sharpens CPPLUS's ruled table (3 names to 8 before the
    union, 63 to 77 after) and ELLEN (28 to 41), but it thins the faint glyphs
    MEESHO depends on and cost that filing 58 real investors including
    GOVERNMENT OF SINGAPORE and AMANSA HOLDINGS PRIVATE LIMITED. Both versions
    are therefore read and both go into the candidate pool, so the union can
    take whichever one found a given investor."""
    up = _deskew_page(img)
    return _enhance_page(up, dpi) if enhance else up


def _ocr_pdf(path: Path, dpi: int, psm: str, enhance: bool = False) -> str:
    """Render every page at `dpi` and OCR it; returns the concatenated text.

    The result is cached on disk. Rendering a 60-page filing at 600dpi and
    OCR'ing it is minutes of CPU, and the same pages get re-read on every run
    and on every parser tweak; the cache turns that into a file read. It is
    keyed on the file's content hash, so replacing a poor scan with a better one
    invalidates it automatically."""
    key = _ocr_key(path, dpi, psm, enhance)
    cached = OCR_CACHE_DIR / f"{key}.txt"
    if cached.exists():
        try:
            return cached.read_text(encoding="utf-8")
        except Exception:
            pass
    pages = _render_pages(path, dpi, enhance)
    if not pages:
        return ""
    text = "\n".join(_tesseract(p, psm) for p in pages)
    try:
        OCR_CACHE_DIR.mkdir(parents=True, exist_ok=True)
        cached.write_text(text, encoding="utf-8")
    except Exception:
        pass
    return text


def _tesseract_tsv(img: Path, psm: str) -> list[dict]:
    """One page's text lines WITH their position on the page.

    Tesseract's plain-text output throws away the one thing needed to read a
    ruled table correctly: where each line sits. Its `tsv` output keeps the
    bounding box of every word, so the words are regrouped here into lines
    carrying `top`/`bot`/`left`. See _parse_boxed_anchor_table for why that is
    not a nicety but the only reliable way to attach a wrapped name to its own
    row."""
    try:
        r = subprocess.run(["tesseract", str(img), "stdout", "--oem", "1", "--psm", psm, "tsv"],
                           capture_output=True, text=True, timeout=300)
        out = r.stdout or ""
    except Exception:
        return []
    raw = out.splitlines()
    if not raw:
        return []
    head = raw[0].split("\t")
    try:
        ix = {k: head.index(k) for k in ("block_num", "par_num", "line_num",
                                         "left", "top", "height", "conf", "text")}
    except ValueError:
        return []
    groups: dict[tuple, list] = {}
    order: list[tuple] = []
    for line in raw[1:]:
        f = line.split("\t")
        if len(f) <= ix["text"]:
            continue
        txt = f[ix["text"]].strip()
        if not txt:
            continue
        try:
            # conf is a float in tesseract 5 ("16.899977"), an int in 4.
            if float(f[ix["conf"]]) < 0:
                continue
            left = int(f[ix["left"]])
            top = int(f[ix["top"]])
            height = int(f[ix["height"]])
        except ValueError:
            continue
        key = (f[ix["block_num"]], f[ix["par_num"]], f[ix["line_num"]])
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append((left, top, height, txt))
    lines = []
    for key in order:
        ws = sorted(groups[key], key=lambda w: w[0])
        lines.append({
            "top": min(w[1] for w in ws),
            "bot": max(w[1] + w[2] for w in ws),
            "left": min(w[0] for w in ws),
            "text": " ".join(w[3] for w in ws),
        })
    lines.sort(key=lambda d: (d["top"], d["left"]))
    return lines


def _ocr_pdf_boxes(path: Path, dpi: int, psm: str,
                   enhance: bool = False) -> list[list[dict]]:
    """Every page of a filing as positioned text lines, cached like _ocr_pdf.

    Kept in a separate cache file from the plain-text OCR so that both survive;
    the key is the same content hash, so a replaced scan invalidates both."""
    key = _ocr_key(path, dpi, psm, enhance)
    cached = OCR_CACHE_DIR / f"{key}.lines.json"
    if cached.exists():
        try:
            return json.loads(cached.read_text(encoding="utf-8"))
        except Exception:
            pass
    rendered = _render_pages(path, dpi, enhance)
    if not rendered:
        return []
    pages = [_tesseract_tsv(p, psm) for p in rendered]
    try:
        OCR_CACHE_DIR.mkdir(parents=True, exist_ok=True)
        cached.write_text(json.dumps(pages), encoding="utf-8")
    except Exception:
        pass
    return pages


def _longest_run(v) -> int:
    """Length of the longest unbroken True run in a 1-D boolean array."""
    import numpy as np
    idx = np.flatnonzero(~v)
    if idx.size == 0:
        return int(v.size)
    d = np.diff(np.concatenate(([-1], idx, [v.size])))
    return int(d.max()) - 1


def _close_gaps(v):
    """Binary closing: bridge breaks of up to 4 pixels so that an anti-aliased
    or lightly broken rule still measures as one continuous run, without
    lengthening the gaps between separate glyphs."""
    d = v.copy()
    for k in (1, 2):
        d[:-k] |= v[k:]
        d[k:] |= v[:-k]
    e = d.copy()
    for k in (1, 2):
        e[:-k] &= d[k:]
        e[k:] &= d[:-k]
    return e | v


def _rule_runs(dark, axis: int) -> dict[int, int]:
    """Longest dark run on each line that could be a printed rule.

    Lines lying inside a solid dark region are left out. A logo band bridges
    into one long "run" as readily as a rule does - MEESHO's letterhead reads
    2264 pixels, wider than any rule the filing draws - so anything measuring
    itself against "the longest dark run on the page" ends up measuring the
    letterhead. What separates them is their surroundings: at that band 87% of
    the neighbouring pixels are dark, at the table's own rules 6%."""
    import numpy as np
    span = dark.shape[1] if axis == 1 else dark.shape[0]
    counts = dark.sum(axis=axis).astype(float)
    cand = np.flatnonzero(counts >= 0.15 * span)
    if not cand.size:
        return {}
    half = 2 * max(6, int(span * 0.004))
    csum = np.concatenate(([0.0], np.cumsum(counts)))
    out = {}
    for i in cand:
        i = int(i)
        lo, hi = max(0, i - half), min(counts.size, i + half + 1)
        if (csum[hi] - csum[lo]) / ((hi - lo) * span) > 0.4:
            continue
        out[i] = _longest_run(_close_gaps(
            dark[i] if axis == 1 else dark[:, i]))
    return out


def _skew_angle(arr) -> float:
    """The sub-degree rotation that straightens a scanned table.

    Ruling lines are found by projection, and projection is destroyed by skew
    far smaller than the eye notices: ELLEN page 2 is a fully ruled table, but
    at its scanned angle the longest continuous horizontal run is 35% of the
    page width, which reads as "no rules here". Rotated by a quarter of one
    degree the same page gives 76%. Tesseract's own OSD only reports multiples
    of 90, so it cannot help. The angle is searched on a downscaled copy - a
    rule stays continuous through the reduction, and it makes the search cheap
    enough to run on every page."""
    import numpy as np
    from PIL import Image
    h, w = arr.shape
    small = Image.fromarray(arr)
    if w > 1500:
        small = small.resize((1500, max(1, int(h * 1500 / w))), Image.BILINEAR)

    def _axes(im) -> tuple[float, float]:
        # Both axes, added, because the page has to serve both sets of rules
        # and the angle that flatters one can ruin the other. On MEESHO the
        # horizontal reading barely moves while the verticals go from 22% of
        # the page to 75%, so scoring on width alone - as this did - called
        # every page straight and left the grid unreadable.
        a = np.asarray(im) < 170
        return (max(_rule_runs(a, 1).values(), default=0) / a.shape[1],
                max(_rule_runs(a, 0).values(), default=0) / a.shape[0])

    # A page whose rules already run their full length needs no search.
    hh, vv = _axes(small)
    if min(hh, vv) >= 0.75:
        return 0.0
    best, best_score = 0.0, hh + vv
    for deg in [x / 4 for x in range(-8, 9)]:
        if deg == 0:
            continue
        s = sum(_axes(small.rotate(deg, resample=Image.BILINEAR,
                                   fillcolor=255)))
        if s > best_score:
            best, best_score = deg, s
    return best


def _grid_lines(arr) -> tuple[list[int], list[int]]:
    """Positions of the horizontal and vertical ruling lines on a page.

    A line is a row (or column) of pixels carrying one long unbroken dark run.
    The threshold is relative to the longest run actually found rather than to
    the page width, because an anchor table rarely spans the full page. Runs of
    adjacent hits are collapsed to their midpoint so a 3-pixel-thick rule counts
    once."""
    import numpy as np
    try:
        import cv2
        thr = cv2.threshold(arr, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
        dark = thr > 0
    except ImportError:
        dark = arr < 170
    h, w = dark.shape

    def _scan(get: str, span: int) -> list[int]:
        counts = dark.sum(axis=1) if get == "h" else dark.sum(axis=0)
        cand = np.flatnonzero(counts >= 0.15 * span)
        runs = {int(i): _longest_run(_close_gaps(
            dark[i] if get == "h" else dark[:, i])) for i in cand}
        if not runs:
            return []
        # The reference is the longest run that is plausibly a rule. Taking it
        # over everything let MEESHO's letterhead banner (2282px, 92% of the
        # page) set the bar at 1369px, just above that filing's own row rules
        # (1200-1433px), so the real table was rejected and the whole body
        # collapsed into a single cell.
        thin = _rule_runs(dark, 1 if get == "h" else 0)
        top = max(thin.values(), default=0)
        if top < 0.25 * span:
            top = max(runs.values())
        if top < 0.25 * span:
            return []
        hits = sorted(i for i, r in runs.items() if r >= max(0.6 * top,
                                                            0.25 * span))
        out, group = [], [hits[0]]
        for i in hits[1:]:
            if i - group[-1] <= 4:
                group.append(i)
            else:
                out.append(sum(group) // len(group))
                group = [i]
        out.append(sum(group) // len(group))
        return out

    return _scan("h", w), _scan("v", h)


def _tesseract_words(img: Path, psm: str) -> list[dict]:
    """Every word on a page with its bounding box, for assigning to table cells."""
    try:
        r = subprocess.run(["tesseract", str(img), "stdout", "--oem", "1", "--psm", psm, "tsv"],
                           capture_output=True, text=True, timeout=300)
        out = r.stdout or ""
    except Exception:
        return []
    raw = out.splitlines()
    if not raw:
        return []
    head = raw[0].split("\t")
    try:
        ix = {k: head.index(k) for k in ("left", "top", "width", "height",
                                         "conf", "text")}
    except ValueError:
        return []
    words = []
    for line in raw[1:]:
        f = line.split("\t")
        if len(f) <= ix["text"]:
            continue
        txt = f[ix["text"]].strip()
        if not txt:
            continue
        try:
            if float(f[ix["conf"]]) < 0:
                continue
            x, y = int(f[ix["left"]]), int(f[ix["top"]])
            dx, dy = int(f[ix["width"]]), int(f[ix["height"]])
        except ValueError:
            continue
        words.append({"cx": x + dx / 2, "cy": y + dy / 2, "left": x,
                      "top": y, "h": dy, "text": txt})
    return words


def _reading_order(ws: list[tuple]) -> str:
    """Words of one cell in reading order.

    Sorting on the raw top coordinate is not enough: within a single printed
    line the boxes of tall and short words start at different heights, so a
    two-line cell interleaves - "FINAVENUE CAPITAL TRUST- FINAVENUE GROWTH
    FUND" came back as "CAPITAL TRUST- FINAVENUE FINAVENUE GROWTH FUND". Words
    are therefore banded into lines first, using the words' own height as the
    tolerance, and only then read left to right."""
    if not ws:
        return ""
    # The printed rules themselves get read as characters ('|', '_') sitting
    # inside the cell they bound; they are not part of any name.
    ws = [w for w in ws if not re.fullmatch(r"[|_/\\:;'\"`\u2018\u2019\u201c\u201d]+", w[3])]
    if not ws:
        return ""
    ws = sorted(ws, key=lambda t: t[0])
    tol = max(4, int(sorted(w[2] for w in ws)[len(ws) // 2] * 0.6))
    lines, cur, base = [], [ws[0]], ws[0][0]
    for w in ws[1:]:
        if w[0] - base <= tol:
            cur.append(w)
        else:
            lines.append(cur)
            cur, base = [w], w[0]
    lines.append(cur)
    return " ".join(w[3] for ln in lines for w in sorted(ln, key=lambda t: t[1]))


def _page_grid_cells(img: Path, psm: str) -> list[list[str]]:
    """A ruled page read as a real table: rows of cell strings.

    This is the tier that the gap-based reader cannot reach. Where the issuer
    printed a grid, the grid says exactly which row and which column every word
    belongs to, so a wrapped name cannot bleed into its neighbour and a figure
    cannot be mistaken for a share count just because it is the first number
    over a thousand on the line. Pages with no usable grid return nothing and
    the caller falls back."""
    try:
        import numpy as np
        from PIL import Image
    except Exception:
        return []
    try:
        arr = np.asarray(Image.open(img).convert("L"))
    except Exception:
        return []
    use = img
    deg = _skew_angle(arr)
    if abs(deg) >= 0.1:
        try:
            rot = Image.open(img).convert("L").rotate(
                deg, resample=Image.BICUBIC, fillcolor=255)
            use = img.with_name(img.stem + "-fine.png")
            rot.save(use)
            arr = np.asarray(rot)
        except Exception:
            use = img
    hs, vs = _grid_lines(arr)
    if len(hs) < 3 or len(vs) < 2:
        return []
    # The outermost columns are often unruled - ELLEN's table draws no rule down
    # its left edge, so every investor NAME sat outside the detected grid and
    # the whole table read as figures only. The page edges close it off.
    if vs[0] > 10:
        vs = [0] + vs
    if vs[-1] < arr.shape[1] - 10:
        vs = vs + [arr.shape[1] - 1]
    if len(vs) < 3:
        return []
    words = _tesseract_words(use, psm)
    if not words:
        return []
    nrow, ncol = len(hs) - 1, len(vs) - 1
    cells: list[list[list[tuple]]] = [[[] for _ in range(ncol)]
                                      for _ in range(nrow)]
    import bisect
    for wd in words:
        r = bisect.bisect_right(hs, wd["cy"]) - 1
        c = bisect.bisect_right(vs, wd["cx"]) - 1
        if 0 <= r < nrow and 0 <= c < ncol:
            cells[r][c].append((wd["top"], wd["left"], wd["h"], wd["text"]))
    rows = []
    for r in cells:
        row = [_reading_order(c) for c in r]
        if any(x.strip() for x in row):
            rows.append(row)
    return rows


def _ocr_pdf_grid(path: Path, dpi: int, psm: str,
                  enhance: bool = False) -> list[list[list[str]]]:
    """Every page of a filing as ruled table cells, cached like _ocr_pdf."""
    key = _ocr_key(path, dpi, psm, enhance)
    cached = OCR_CACHE_DIR / f"{key}.grid.json"
    if cached.exists():
        try:
            return json.loads(cached.read_text(encoding="utf-8"))
        except Exception:
            pass
    rendered = _render_pages(path, dpi, enhance)
    if not rendered:
        return []
    pages = [_page_grid_cells(p, psm) for p in rendered]
    try:
        OCR_CACHE_DIR.mkdir(parents=True, exist_ok=True)
        cached.write_text(json.dumps(pages), encoding="utf-8")
    except Exception:
        pass
    return pages


_SEP_NUM = re.compile(r"(\d[\d.,]*\d)(%?)")


def _fix_digit_seps(cell: str) -> str:
    """Restore thousands separators the recogniser prints as full stops.

    It reads 998,378 as "998.378" and 110,819,958.00 as "110.819.958.00". The
    group SIZES disambiguate: a closing group of three digits can only be
    thousands, a closing group of two can only be decimals."""
    s = cell.strip().replace(" ", "")
    m = _SEP_NUM.fullmatch(s)
    if not m:
        return cell
    parts = re.split(r"[.,]", m.group(1))
    if len(parts) < 2 or not all(p.isdigit() for p in parts):
        return cell
    if any(len(p) != 3 for p in parts[1:-1]):
        return cell
    if len(parts[-1]) == 3:
        return "".join(parts) + m.group(2)
    if len(parts[-1]) == 2:
        return f'{"".join(parts[:-1])}.{parts[-1]}{m.group(2)}'
    return cell


def _surya_available() -> bool:
    """True if surya-ocr is installed and usable."""
    try:
        from surya.recognition import RecognitionPredictor  # noqa: F401
        return True
    except Exception:
        return False


_SURYA_DPI = 300
_SURYA_PREDICTOR_CACHE: dict = {}


def _surya_predictors():
    """Lazily initialise the Surya inference manager and predictors."""
    if not _SURYA_PREDICTOR_CACHE:
        from surya.recognition import RecognitionPredictor
        from surya.detection import DetectionPredictor
        from surya.inference import SuryaInferenceManager
        mgr = SuryaInferenceManager()
        _SURYA_PREDICTOR_CACHE["det"] = DetectionPredictor(mgr)
        _SURYA_PREDICTOR_CACHE["rec"] = RecognitionPredictor(mgr)
    return _SURYA_PREDICTOR_CACHE["det"], _SURYA_PREDICTOR_CACHE["rec"]


def _surya_ocr_pages(path: Path) -> list[list[dict]]:
    """Every page of a PDF as positioned text lines via Surya OCR, cached.

    Returns a list of pages, each page a list of dicts with keys
    'text', 'top', 'bot', 'left' — the same shape _parse_boxed_anchor_table
    consumes."""
    if not _surya_available():
        return []
    key = _ocr_key(path, _SURYA_DPI, "surya", False)
    cached = OCR_CACHE_DIR / f"{key}.surya.json"
    if cached.exists():
        try:
            return json.loads(cached.read_text(encoding="utf-8"))
        except Exception:
            pass
    from PIL import Image as PILImage
    with tempfile.TemporaryDirectory() as td:
        try:
            subprocess.run(["pdftoppm", "-r", str(_SURYA_DPI), "-png",
                            str(path), str(Path(td) / "pg")],
                           capture_output=True, timeout=900, check=True)
        except Exception:
            return []
        img_paths = sorted(Path(td).glob("pg*.png"))
        if not img_paths:
            return []
        try:
            det_pred, rec_pred = _surya_predictors()
        except Exception:
            return []
        all_pages: list[list[dict]] = []
        for ip in img_paths:
            try:
                pil_img = PILImage.open(ip)
                det_results = det_pred([pil_img])
                rec_results = rec_pred([pil_img], det_results)
                page_lines: list[dict] = []
                result = rec_results[0]
                for block in result.blocks:
                    if block.error or block.skipped:
                        continue
                    text = block.html or ""
                    text = re.sub(r"<[^>]+>", " ", text).strip()
                    text = _fix_digit_seps(text)
                    if not text:
                        continue
                    bbox = block.bbox
                    page_lines.append({
                        "text": text,
                        "top": bbox[1],
                        "bot": bbox[3],
                        "left": bbox[0],
                    })
                page_lines.sort(key=lambda ln: (ln["top"], ln["left"]))
                all_pages.append(page_lines)
            except Exception:
                all_pages.append([])
    try:
        OCR_CACHE_DIR.mkdir(parents=True, exist_ok=True)
        cached.write_text(json.dumps(all_pages), encoding="utf-8")
    except Exception:
        pass
    return all_pages


def _paddle_available() -> bool:
    """True if PaddleOCR is installed and usable."""
    try:
        from paddleocr import PaddleOCR  # noqa: F401
        return True
    except Exception:
        return False


_PADDLE_DPI = 300
_PADDLE_OCR_CACHE: dict = {}


def _paddle_engine():
    """Lazily initialise the PaddleOCR engine (PP-OCRv4, English)."""
    if "eng" not in _PADDLE_OCR_CACHE:
        from paddleocr import PaddleOCR
        _PADDLE_OCR_CACHE["eng"] = PaddleOCR(
            use_angle_cls=True, lang="en", show_log=False)
    return _PADDLE_OCR_CACHE["eng"]


def _paddle_ocr_pages(path: Path) -> list[list[dict]]:
    """Every page of a PDF as positioned text lines via PaddleOCR, cached.

    Returns the same shape as ``_surya_ocr_pages``: a list of pages, each page
    a list of dicts with keys 'text', 'top', 'bot', 'left'."""
    if not _paddle_available():
        return []
    key = _ocr_key(path, _PADDLE_DPI, "paddle", False)
    cached = OCR_CACHE_DIR / f"{key}.paddle.json"
    if cached.exists():
        try:
            return json.loads(cached.read_text(encoding="utf-8"))
        except Exception:
            pass
    with tempfile.TemporaryDirectory() as td:
        try:
            subprocess.run(["pdftoppm", "-r", str(_PADDLE_DPI), "-png",
                            str(path), str(Path(td) / "pg")],
                           capture_output=True, timeout=900, check=True)
        except Exception:
            return []
        img_paths = sorted(Path(td).glob("pg*.png"))
        if not img_paths:
            return []
        try:
            ocr = _paddle_engine()
        except Exception:
            return []
        all_pages: list[list[dict]] = []
        for ip in img_paths:
            try:
                result = ocr.ocr(str(ip), cls=True)
                page_lines: list[dict] = []
                if result and result[0]:
                    for line in result[0]:
                        bbox, (text, _conf) = line
                        text = _fix_digit_seps(text.strip())
                        if not text:
                            continue
                        ys = [pt[1] for pt in bbox]
                        xs = [pt[0] for pt in bbox]
                        page_lines.append({
                            "text": text,
                            "top": min(ys),
                            "bot": max(ys),
                            "left": min(xs),
                        })
                page_lines.sort(key=lambda ln: (ln["top"], ln["left"]))
                all_pages.append(page_lines)
            except Exception:
                all_pages.append([])
    try:
        OCR_CACHE_DIR.mkdir(parents=True, exist_ok=True)
        cached.write_text(json.dumps(all_pages), encoding="utf-8")
    except Exception:
        pass
    return all_pages


_TABLE_HEADER_WORDS = ("srno", "nameof", "equityshares", "sharesallocated",
                       "totalamount", "bidprice", "allocationprice",
                       "mutualfund", "investorportion")


def _is_table_header_row(row: list[str]) -> bool:
    """True if this grid row is a table's column headings.

    Once the anchor table's own heading has been consumed, the next one to
    appear belongs to the SECOND, mutual-fund-only table these letters carry,
    whose rows are anchors already counted once. MEESHO runs to serial 125 and
    then starts again at 1, and those 34 rows were being published twice. The
    Total row that used to mark the end is not dependable - at 450 dpi the scan
    loses it - but a heading is several cells of heading vocabulary at once,
    which no investor's name manages."""
    hits = sum(1 for c in row
               if any(w in _norm_alnum(c) for w in _TABLE_HEADER_WORDS))
    return hits >= 2


def _name_col_by_content(page: list[list[str]]) -> int | None:
    """Which column of a headerless page holds the investor names.

    Continuation pages carry no header, and their column indices need not match
    the header page's: MEESHO page 3 loses the rule between "Sr. No" and the
    name, so its names sit at index 0 where page 1 puts them at index 2.
    Carrying the header page's index across read the percentage column as the
    name and threw the whole filing away. Names are the one column that is
    mostly letters, so the page can be asked directly."""
    width = max((len(r) for r in page), default=0)
    best, best_n = None, 0
    for j in range(width):
        n = 0
        for r in page:
            c = r[j] if j < len(r) else ""
            letters = sum(ch.isalpha() for ch in c)
            if letters >= 6 and letters >= 2 * sum(ch.isdigit() for ch in c):
                n += 1
        if n > best_n:
            best, best_n = j, n
    return best if best_n >= 2 else None


def _parse_grid_anchor_table(pages: list[list[list[str]]], stated=None,
                             trusted: bool = False):
    """Anchor rows from ruled table cells.

    Two things become provable here that the line-based reader can only guess.
    The NAME is whatever sits in the name column of that row, so no serial digit
    and no stray figure can join it. And the SHARE COUNT is picked by finding
    the column whose values actually sum to the total the letter declares - the
    same standard of proof _trim_to_total applies to rows, applied to columns.
    Without it the reader takes the first number over a thousand, which is how
    MEESHO came to 49 billion shares against a stated 220 million."""
    name_col = None
    data: list[tuple[list[str], int]] = []
    done = False
    for page in pages:
        if done:
            break
        page_col = None
        start = 0
        for i, row in enumerate(page):
            flags = [_norm_alnum(c) for c in row]
            # The header itself wraps, so its words arrive in whatever order the
            # cell stacked them ("of Name Anchor Investors"); match on the words
            # present, not on their order. "nvestor" rather than "investor"
            # because MEESHO's header loses its leading I to the column rule.
            hit = next((j for j, f in enumerate(flags)
                        if "name" in f and ("anchor" in f or "nvestor" in f)),
                       None)
            if hit is None:
                hit = next((j for j, f in enumerate(flags)
                            if f in ("name", "nameoftheinvestor", "investorname",
                                     "nameofinvestor", "anchorinvestor")), None)
            if hit is not None:
                page_col, start = hit, i + 1
                break
        if page_col is not None:
            name_col = page_col
        else:
            page_col = _name_col_by_content(page)
            if page_col is None:
                page_col = name_col
            if page_col is None:
                continue
        for row in page[start:]:
            # The Total row ends the table for good. Reading on would pick up
            # the second, mutual-fund-only table that these letters also carry,
            # whose rows are anchors already counted once.
            # Guard: a name like "TOTAL ENERGY FUND" should NOT end parsing.
            # Real total rows carry a large number (≥50% of stated, or ≥10000).
            def _looks_like_total_row(r):
                for c in r:
                    flat = _norm_alnum(c)
                    if not flat.startswith(("total", "grandtotal")):
                        continue
                    nums = [_to_num(t) for t in re.findall(r"[\d,.]+", c)]
                    big = [n for n in nums if n and n >= (
                        stated * 0.5 if stated else 10000)]
                    if big:
                        return True
                    if len(re.findall(r"[A-Za-z]{3,}", c)) <= 2:
                        return True
                return False
            if page_col < len(row) and _norm_alnum(row[page_col]).startswith(
                    ("total", "grandtotal")) and _looks_like_total_row(row):
                done = True
                break
            if (any(_norm_alnum(c).startswith(("total", "grandtotal"))
                    for c in row[:max(1, page_col)])
                    and _looks_like_total_row(row)):
                done = True
                break
            # That second table's own heading, for the filings whose Total row
            # the scan loses.
            if _is_table_header_row(row):
                done = True
                break
            data.append((row, page_col))
    if not data:
        return []

    # Pages that found their name in a different column are shifted right so
    # that one column index means the same thing on every row.
    name_col = max(c for _, c in data)
    data = [[""] * (name_col - c) + list(r) for r, c in data]

    width = max(len(r) for r in data)
    vals: list[list[float | None]] = []
    for r in data:
        cells = list(r) + [""] * (width - len(r))
        row_v: list[float | None] = []
        for j, c in enumerate(cells):
            nums = _numbers_in(c)
            row_v.append(nums[0] if (j != name_col and len(nums) == 1) else None)
        vals.append(row_v)

    # Which rows are real investor rows, settled once. The column scoring below
    # used to run over every row it had collected, including the Total row and
    # any page furniture that the emission loop then discarded. On MEESHO that
    # single Total row carried the whole declared allocation a second time, so
    # the share column appeared to sum to twice the stated figure and was
    # rejected as mis-read - discarding a table that was in fact read perfectly.
    names = []
    for r in data:
        cells = list(r) + [""] * (width - len(r))
        nm = re.sub(r"\s+", " ", cells[name_col]).strip(" -—.")
        # The prose and letterhead tests read a NAME and guess whether it came
        # from the table, because flat OCR text cannot tell the two apart. A
        # recognised table can: furniture is not in it. Applying them anyway
        # deleted eight real MEESHO investors that had been read perfectly,
        # among them serial 34 - 128 characters long, over the 90 that mark
        # prose - and "AMUND FUNDS NEW SILK ROAD", for containing "road".
        junk = (_is_header_line(nm) if trusted else
                (_is_prose(nm) or _is_letterhead(nm) or _is_header_line(nm)))
        names.append(None if (len(nm) <= 3 or junk) else nm)

    share_col = None
    if stated:
        # Which column IS the share count is settled by arithmetic, not by
        # position or by a header word: the one whose values come nearest the
        # total the letter declares. On E2ERAIL the amount column sums to 227
        # million against a declared 1.38 million, so it cannot be mistaken for
        # the share column however it is laid out. If nothing lands anywhere
        # near the declared total the grid has been mis-read, and returning
        # nothing hands the page back to the line-based reader rather than
        # publishing figures that were never verified.
        scored = []
        for j in range(width):
            got = [v[j] for v, nm in zip(vals, names)
                   if nm and v[j] is not None
                   and v[j].is_integer() and v[j] >= MIN_ANCHOR_SHARES]
            if got:
                scored.append((abs(sum(got) - stated), j))
        if not scored:
            return []
        diff, share_col = min(scored)
        if diff > 0.25 * stated:
            return []
    else:
        # No declared total to prove it against: fall back to the widest column
        # of plausible whole share counts.
        best = -1
        for j in range(width):
            n = sum(1 for v, nm in zip(vals, names)
                    if nm and v[j] is not None and v[j].is_integer()
                    and v[j] >= MIN_ANCHOR_SHARES)
            if n > best:
                best, share_col = n, j
        if best <= 0:
            share_col = None

    rows = []
    for r, v, nm in zip(data, vals, names):
        if nm is None:
            continue
        share = None
        if share_col is not None and v[share_col] is not None \
                and v[share_col].is_integer() and v[share_col] >= MIN_ANCHOR_SHARES:
            share = int(v[share_col])
        tail = [x for j, x in enumerate(v)
                if x is not None and j != name_col
                and (share_col is None or j > share_col)]
        rows.append((nm, share, tail))
    return rows


def _stated_total_shares(text: str):
    """The share count the letter itself declares was allotted to anchors."""
    for pat in (
        r"allocat\w*\s+of\s+([\d,.]{4,})\s*(?:equity|shares)",
        r"allot\w*\s+of\s+([\d,.]{4,})\s*(?:equity|shares)",
        r"allocat\w*\s+of\s+([\d,.]{4,})\s*(?:equily|equi[lt]y)",
        r"total\s+(?:anchor|allocation)\D{0,30}?([\d,.]{4,})\s*(?:equity|shares)",
        r"([\d,.]{4,})\s*equity\s+shares?\s+(?:to|amongst|among)\s+anchor",
    ):
        m = re.search(pat, text, re.I)
        if m:
            v = _to_int(m.group(1))
            if v and v >= 1000:
                return v
    return None


def _stated_anchor_price(text: str):
    """The per-share anchor allocation price as written in the letter's prose."""
    for pat in (r"price of\s*(?:Rs\.?|INR|\u20b9)?\s*([\d,]+(?:\.\d+)?)\s*(?:/-)?\s*"
                r"per\s+(?:equity\s+)?share",
                r"anchor investor allocation price\D{0,20}?([\d,]+(?:\.\d+)?)",
                r"(?:Rs\.?|INR|\u20b9)\s*([\d,]+(?:\.\d+)?)\s*(?:/-)?\s*per\s+"
                r"(?:equity\s+)?share"):
        m = re.search(pat, text, re.I)
        if m:
            v = _to_num(m.group(1))
            if v and v > 0:
                return v
    return None


def _resolve_allocation_price(rows, stated_price=None):
    """The single per-share price the whole anchor book was allotted at.

    Every anchor in an issue pays the SAME price, so the table is massively
    over-determined: for the right price P, shares x P equals the row's own
    amount column on every row. Each row therefore votes for the P that makes
    its own arithmetic close, and the winner is the value most rows agree on.
    That beats reading the price cell directly (OCR mangles a lone 3-digit
    number readily) and beats trusting the letter's prose, where the same regex
    can just as easily catch a face value or a price band."""
    votes: dict[float, int] = {}
    for _name, shares, tail in rows:
        if not shares:
            continue
        for p in tail:
            if p <= 0 or p > 100_000:
                continue
            target = shares * p
            tol = max(1.0, target * 0.005)
            if any(abs(a - target) <= tol for a in tail):
                key = round(p, 2)
                votes[key] = votes.get(key, 0) + 1
    if votes:
        return max(votes.items(), key=lambda kv: (kv[1], kv[0]))[0]
    if stated_price:
        return stated_price
    # Nothing reconciled: fall back to the amount column over the share count,
    # taken as the MEDIAN so a single mis-read row cannot price the whole issue.
    est = sorted(max(tail) / shares for _n, shares, tail in rows
                 if shares and tail and max(tail) > shares)
    return round(est[len(est) // 2], 2) if est else None


def _finalize_anchors(rows, stated_price=None) -> list[Anchor]:
    """Turn parsed (name, shares, tail) rows into priced Anchor records."""
    price = _resolve_allocation_price(rows, stated_price)
    return [Anchor(name=name, shares=shares, price=price,
                   amount=round(shares * price, 2) if (shares and price) else None)
            for name, shares, _tail in rows]


# Header cells wrap onto several lines in a scan ("No. of Equity Shares",
# "Allocated", "Investor Portion", "(in Rs. per Equity Share)"), and they sit
# BELOW the row that opens the table — without this they get glued onto the
# first investor's name.
_HEADER_NOISE = ("allocated", "investorportion", "bidprice", "perequity",
                 "noofequity", "ofanchor", "equityshare", "sharesallocated",
                 "srno", "inrsper", "noofequio", "totalamount", "amountinrs")
# Standalone leftovers of a wrapped header cell. Matched EXACTLY, never as a
# substring — "share" as a substring would wrongly kill a real name such as
# "Share India Securities".
_HEADER_FRAGMENTS = {"share", "shares", "allocated", "portion", "name", "no",
                     "sr", "srno", "nos", "inrs", "rs", "percent", "total"}
# Once the table ends the OCR keeps going into the letter's prose; these markers
# identify a captured "name" that is really a sentence.
_PROSE_MARKERS = (" shall ", " please ", " regulation", " prospectus",
                  " website", " thanking ", " exchange board", " submitted ",
                  " as per ", " in case ", " scheme", " pension fund",
                  " required to pay", " disclosing ")


def _is_prose(name: str) -> bool:
    low = " " + name.lower() + " "
    return len(name) > 90 or any(w in low for w in _PROSE_MARKERS)


# The column headings are a stack of narrow wrapped cells, so they reach the
# parser as a run of short lines ("Anchor", "Investor", "as % of Share)",
# "N Shares Shares (= Per (in 2)"). Listing each one as noise is whack-a-mole
# and it lost: on E2ERAIL four of them survived and were glued onto the first
# investor, which came out as "N Shares Shares (< Per (in <) Anchor Investor
# SANSHI FUND-I". A line is treated as heading only when EVERY word in it is a
# heading word, which no real investor name manages - "Equity Shares Fund" is
# kept because "fund" is not in here.
_HEADER_WORDS = frozenset({
    "sr", "srno", "s", "no", "nos", "name", "names", "of", "the", "as", "in",
    "per", "and", "a", "at", "to", "on", "anchor", "anchors", "investor",
    "investors", "portion", "equity", "share", "shares", "price", "prices",
    "allocated", "allocation", "allotted", "allotment", "total", "amount",
    "amounts", "bid", "rs", "inr", "value", "face", "percent", "pct", "each",
    "n", "nc", "no1",
})


def _is_header_line(name: str) -> bool:
    words = re.findall(r"[A-Za-z]+", name or "")
    return bool(words) and all(w.lower() in _HEADER_WORDS for w in words)


def _parse_ocr_anchor_table(text: str, stated: int | None = None):
    """Pull (names, shares) out of OCR'd anchor-table text.

    Long investor names wrap across lines, and WHICH SIDE of the numbers they
    wrap onto differs by filing. The layout is therefore detected from the first
    data row rather than assumed:

      * name shares the line with the numbers -> the remainder wraps BELOW, so a
        following line without numbers continues the row just closed.
      * numbers sit alone on their line -> the name was written ABOVE and is
        reassembled from the buffered lines.

    Getting this backwards does not lose rows, it SHIFTS every wrapped name onto
    its neighbour ("Growth Fund Aarth AIF Growth Fund"), which is far more
    insidious than a missing row.

    The whole table is read; parsing does NOT stop once the allocated shares
    reach the letter's declared total. It used to, and that silently truncated
    any filing where OCR misread one share cell too large: JAINREC stopped after
    2 of ~40 anchors and QUADFUTURE after 1, because a single inflated figure
    tripped the total on the first row. Trailing junk is dealt with afterwards,
    by _trim_to_total(), which can see the whole table.

    Returns a list of (name, allocated shares, trailing numbers) triples; the
    trailing numbers are what the price and amount columns are recovered from
    once the whole table is in hand."""
    lines = []
    for raw in text.splitlines():
        line = re.sub(r"\s+", " ", raw.replace("|", " ")).strip()
        if line:
            lines.append((line, _norm_alnum(line)))

    # Where does the table begin? Normally at its "Name of the Anchor Investor"
    # header. But that header is a multi-line stack of narrow cells, and on a
    # poor scan the name cell comes back as "eee nor er" or "Namegiiiuchor a" —
    # or, on a small SME letter, the table carries no header at all. Gating on
    # the header alone therefore returned NOTHING for CPPLUS, GKENERGY and
    # BALAJIPHOS despite clean, complete text. Every such letter ends its
    # opening paragraph with "... in the following manner:" immediately above
    # the table, so that sentence is the fallback anchor point.
    start = None
    for i, (_line, flat) in enumerate(lines):
        if "nameof" in flat and ("anchor" in flat or "investor" in flat):
            start = i + 1
            break
    if start is None:
        for i, (_line, flat) in enumerate(lines):
            if "inthefollowingmanner" in flat:
                start = i + 1
                break
    if start is None:
        return []

    rows: list[list] = []
    buf: list[str] = []
    wrap_below = None          # None until the first data row reveals the layout
    for line, flat in lines[start:]:
        if flat.startswith("total") or flat.startswith("grandtotal"):
            nums = [_to_num(t) for t in re.findall(r"[\d,.]+", line)]
            big = [n for n in nums if n and n >= (
                stated * 0.5 if stated else 10000)]
            word_cnt = len(re.findall(r"[A-Za-z]{3,}", line))
            if big or word_cnt <= 2:
                break
        if flat in _HEADER_FRAGMENTS or any(t in flat for t in _HEADER_NOISE):
            # A header that wraps over several lines leaves its tail sitting in
            # the wrap buffer, which then gets glued to the first investor
            # ("Shares Investor Price HDFC MUTUAL FUND..."). Clearing it here
            # is what keeps row 1 clean.
            buf = []
            continue
        core = re.sub(r"^[^A-Za-z0-9]+", "", line)           # rule/pipe junk
        core = _SERIAL_RE.sub("", core, count=1)             # serial, not a name
        name_part, share, tail = _split_row(core)

        if share is None:                       # no numeric columns on this line
            # A line that opens with a serial number is the HEAD of the next
            # row, not the tail of the one just closed. Without this the long
            # names in CPPLUS's table ran into their neighbours
            # ("SBI CONSUMPTION OPPORTUNITIES FUND 5 HDFC MUTUAL FUND ..."),
            # because that filing wraps short names on the numeric line and
            # long ones above it.
            head = re.match(r"^\d{1,3}\s+(?=[A-Za-z])", name_part)
            if head:
                buf.append(name_part[head.end():])
            elif wrap_below and rows:
                if _is_name_fragment(name_part):
                    rows[-1][0] = f"{rows[-1][0]} {name_part}".strip()
            elif len(name_part) > 1:
                buf.append(name_part)
            continue

        if wrap_below is None:
            # Does the name live on the numeric line? Four or more letters is
            # enough to tell a real name from a stray character on a rule.
            wrap_below = len(re.sub(r"[^A-Za-z]", "", name_part)) >= 4
        nm = re.sub(r"\s+", " ", " ".join(buf + [name_part])).strip(" -—")
        buf = []
        if len(nm) > 3 and not _is_prose(nm) and not _is_letterhead(nm):
            rows.append([nm, share, tail])
    return [(r[0], r[1], r[2]) for r in rows]


def _parse_boxed_anchor_table(pages: list[list[dict]]):
    """Read the anchor table from POSITIONED lines rather than line order.

    This exists because reading order alone cannot say which row a wrapped
    investor name belongs to, and getting that wrong does not drop a row, it
    invents investors. Real examples from the filings in anchor_pdfs:

        ELLEN      "26 Sanshi Fund - I CITI GROUP GLOBAL MARKETS"
                   -> two different anchors welded into one
        E2ERAIL    "AARTH AIF GROWTH FUND AKALPYA INDIA INVESTMENT"
                   -> likewise; and "FINAVENUE CAPITAL TRUST-" was torn from
                      "FINAVENUE GROWTH FUND", which is one single anchor

    The old text parser decided the wrap direction ONCE per table from its first
    data row. That cannot work, because both directions occur inside the SAME
    table: an issuer's cell text is top-aligned while its figures are centred,
    so a two-line name puts the figures on its second line, a three-line name
    puts them in the middle, and a one-line name puts them on the only line.

    Geometry settles it, using two rules applied to each line that carries no
    figures:

      1. A line that opens with the table's next SERIAL number is the head of a
         new row, so it belongs to the figures BELOW it. The serial has to
         continue the table's own count to be believed - "360 ONE LVF Treasury
         Solutions Fund" opens with a number that is part of the name.
      2. Any other line is a continuation of whichever figure-bearing line it is
         vertically NEARER to. Within one cell the wrapped lines sit a few
         pixels apart, or even overlap the taller figure line; between rows
         there is cell padding and a rule, which is many times that gap.

    Verified against ELLEN p2 (28 rows, both wrap directions present) and
    E2ERAIL p1 (serial-headed rows), where it reproduces every name exactly."""
    flat = []
    for pno, lines in enumerate(pages):
        for ln in lines:
            txt = re.sub(r"\s+", " ",
                         (ln.get("text") or "").replace("|", " ")).strip()
            if txt:
                flat.append({"page": pno, "top": ln["top"], "bot": ln["bot"],
                             "text": txt, "flat": _norm_alnum(txt)})

    start = None
    for i, ln in enumerate(flat):
        if "nameof" in ln["flat"] and ("anchor" in ln["flat"]
                                       or "investor" in ln["flat"]):
            start = i + 1
            break
    if start is None:
        for i, ln in enumerate(flat):
            if "inthefollowingmanner" in ln["flat"]:
                start = i + 1
                break
    if start is None:
        return []

    items = []
    seq = 0
    since = 0
    for ln in flat[start:]:
        if ln["flat"].startswith("total") or ln["flat"].startswith("grandtotal"):
            break
        # These letters must also disclose, in a SECOND table, the subset of the
        # allocation that went to domestic mutual funds. Those rows are anchors
        # already read from the main table, so running into them double-counts
        # them. ELLEN came back with 41 rows for a 28-row table that way,
        # because on psm6 tesseract renders no clean "Total" line to stop at.
        if ("outofthetotal" in ln["flat"] or "domesticmutualfund" in ln["flat"]
                or "mutualfundsthrough" in ln["flat"]):
            break
        if ln["flat"] in _HEADER_FRAGMENTS or any(t in ln["flat"]
                                                  for t in _HEADER_NOISE):
            continue
        core = re.sub(r"^[^A-Za-z0-9]+", "", ln["text"])
        serial = None
        m = _SERIAL_RE.match(core)
        if m:
            serial = _to_int(m.group(0))
            core = core[m.end():]
        else:
            # A bare leading number is only the serial if it carries on the
            # table's numbering; otherwise it is the start of the name. The
            # name may open with a bracket ("8 (DIVERSIFIED POWER SECTOR FUND)").
            #
            # The window has to widen for serials the OCR destroyed, or ONE bad
            # read poisons the whole rest of the table. JAINREC: serial 22 came
            # out as "om)" and 23 as "73", so the count froze at 21 and every
            # serial from 24 to 49 was then rejected as too far ahead and left
            # glued to its name ("26 SINGULARITY EQUITY FUND I"). Allowing one
            # step per figure-bearing line seen since the last accepted serial
            # tracks the table's real position instead of its last good read,
            # while still rejecting a number that is part of the name ("360 ONE
            # LVF Treasury Solutions Fund").
            m = re.match(r"^(\d{1,3})\s+(?=[A-Za-z(])", core)
            if m and seq < int(m.group(1)) <= seq + 1 + min(since, 6):
                serial = int(m.group(1))
                core = core[m.end():]
        name, share, tail = _split_row(core)
        if serial is not None:
            seq = serial
            since = 0
        elif share is not None:
            since += 1
        items.append({**ln, "name": name, "share": share, "tail": tail,
                      "serial": serial})

    vals = [i for i, it in enumerate(items) if it["share"] is not None]
    if not vals:
        return []
    owner: dict[int, list[int]] = {v: [] for v in vals}
    for i, it in enumerate(items):
        if it["share"] is not None or not it["name"]:
            continue
        if _is_letterhead(it["name"]) or _is_header_line(it["name"]):
            continue
        if it["serial"] is None and not _is_name_fragment(it["name"]):
            continue
        # Candidates are same-page, or at most one page away (a name that
        # wraps from the bottom of page N to the top of page N+1).
        pg = it["page"]
        prev = next((v for v in reversed(vals)
                     if v < i and abs(items[v]["page"] - pg) <= 1), None)
        nxt = next((v for v in vals
                    if v > i and abs(items[v]["page"] - pg) <= 1), None)
        if it["serial"] is not None:
            pick = nxt if nxt is not None else prev
        elif prev is None or nxt is None:
            pick = nxt if prev is None else prev
        else:
            up = it["top"] - items[prev]["bot"]
            down = items[nxt]["top"] - it["bot"]
            pick = prev if up < down else nxt
        if pick is not None:
            owner[pick].append(i)

    rows = []
    for v in vals:
        above = [items[i]["name"] for i in owner[v]
                 if items[i]["top"] < items[v]["top"]]
        below = [items[i]["name"] for i in owner[v]
                 if items[i]["top"] >= items[v]["top"]]
        nm = re.sub(r"\s+", " ",
                    " ".join([*above, items[v]["name"], *below])).strip(" -—")
        if len(nm) > 3 and not _is_prose(nm) and not _is_letterhead(nm):
            rows.append((nm, items[v]["share"], items[v]["tail"]))
    return rows


# The letters run to two or more pages, and every page repeats the issuer's
# letterhead and footer. Once parsing no longer stops at the declared total,
# those lines reach the parser, and a pincode or a phone number is numerically
# indistinguishable from a share count.
_LETTERHEAD_RE = re.compile(
    r"@|www\.|http|\+\d{2}\s*\d|\b(?:tel|fax|email|e-mail|cin|website|regd|"
    r"registered\s+office|corporate\s+office|floor|road|street|marg|nagar|"
    r"tehsil|distt|district|pin|pincode|india\)|phone|mob|con[ta]{2}ct\s*no)\b"
    # "Sector" is an address word only when it numbers one ("Sector 62"). Bare,
    # it is fund vocabulary, and matching it cost a real anchor: ELLEN's row 8,
    # "RELIANCE ... (DIVERSIFIED POWER SECTOR FUND)", 4,35,009 shares, was
    # thrown away as page furniture and took the filing's reconciliation with
    # it.
    r"|\bsector\s*[-.]?\s*\d",
    re.I)


def _is_letterhead(name: str) -> bool:
    """Is this line part of the page furniture rather than the anchor table?"""
    if _LETTERHEAD_RE.search(name):
        return True
    # A postal address opens with a building or plot number and a comma
    # ("802, Suvog Center, Market Yard, Gultekadi, Pune"). No investor name
    # does, and the digits in it otherwise read as a share count.
    if re.match(r"^\d{1,4},\s", name):
        return True
    words = [w for w in re.findall(r"[A-Za-z]{3,}", name)]
    return len(words) < 1


def _trim_to_total(rows, stated):
    """Cut the parsed table down to the run of rows that actually reconciles.

    Page furniture and the letter's closing paragraphs sit AFTER the table, so
    the genuine rows are a prefix. If some prefix's share counts sum exactly to
    the declared total, that prefix is the table — proven, not guessed. Checking
    every prefix is what lets the parser read past a mis-OCR'd share cell
    instead of stopping dead at it.

    When no exact match is found, a near-match within 0.1% of the stated total
    is accepted as a fallback. A single mis-OCR'd share digit can throw the sum
    off by a handful of shares, and rejecting the entire reading for that loses
    the reconciliation safety net."""
    if not stated or not rows:
        return None
    run = 0
    best_near = None
    tol = stated * 0.001
    for i, r in enumerate(rows):
        run += r[1] or 0
        if run == stated:
            return rows[:i + 1]
        if run > stated:
            break
        if abs(run - stated) <= tol and (
                best_near is None or abs(run - stated) < abs(
                    sum(r2[1] or 0 for r2 in best_near) - stated)):
            best_near = rows[:i + 1]
    return best_near


def _shares_from_amount(tail, price):
    """The share count implied by the row's own amount cell, or None."""
    if not price or price <= 0:
        return None
    for a in tail:
        if a < price * MIN_ANCHOR_SHARES:
            continue
        cand = a / price
        if abs(cand - round(cand)) <= 0.01 and round(cand) >= MIN_ANCHOR_SHARES:
            return int(round(cand))
    return None


def _correct_shares(rows, price):
    """Repair share counts that OCR misread, using the row's own amount cell.

    Every row states shares, price and amount, and price is the same for the
    whole issue, so amount / price recovers the share count independently. A
    single mangled digit in the share column used to sink an entire filing
    (APSISAERO missed its declared total by 1,729 shares out of 915,600 and all
    8 names were thrown away); here the row simply repairs itself."""
    if not price or price <= 0:
        return rows, 0
    fixed = 0
    out = []
    for name, shares, tail in rows:
        best = _shares_from_amount(tail, price)
        # Also when the share cell is BLANK rather than wrong. amount / price
        # landing on a whole number is proof enough, and refusing to use it
        # left MEESHO 4,597,520 shares short of its declared total on rows
        # whose figures were sitting in the very next column.
        if best is not None and best != shares:
            shares = best
            fixed += 1
        out.append((name, shares, tail))
    return out, fixed


def _repair_row_from_percent(rows, stated, price):
    """Close a reconciliation gap left by ONE unverifiable share cell.

    Each row prints its share count three times over - as shares, as a
    percentage of the anchor portion, and as an amount - so a cell the amount
    column cannot vouch for still has the percentage to answer to. MEESHO's DSP
    row read 1,120 against a printed 0.23%, which implies about half a million,
    and that single cell was the whole difference between the filing
    reconciling and not.

    Rows the amount column already confirms are left alone, because the printed
    percentages are themselves rounded and sometimes mis-read - two other rows
    on the same filing disagree with theirs and are perfectly correct."""
    if not stated:
        return rows, False
    gap = stated - sum(s or 0 for _, s, _ in rows)
    if gap <= 0:
        return rows, False
    odd = []
    for i, (_, shares, tail) in enumerate(rows):
        if _shares_from_amount(tail, price) is not None:
            continue
        pct = next((x for x in tail if 0 < x < 100), None)
        if pct is None:
            continue
        implied = pct / 100 * stated
        if abs(implied - (shares or 0)) > max(2000, 0.06 * implied):
            odd.append((i, implied))
    if len(odd) != 1:
        return rows, False
    i, implied = odd[0]
    name, shares, tail = rows[i]
    mended = (shares or 0) + gap
    # The mended figure has to satisfy the percentage it was measured against,
    # or the gap belongs to some other row and this one is not the culprit.
    if abs(mended - implied) > max(2000, 0.06 * implied):
        return rows, False
    rows = list(rows)
    rows[i] = (name, mended, tail)
    return rows, True


def extract_anchors_via_ocr(path: Path) -> list[Anchor]:
    """OCR a scanned filing and return its anchor rows, reconciled against the
    letter's stated total wherever that total is readable."""
    is_pdf = path.suffix.lower() == ".pdf"
    if is_pdf and not _ocr_available():
        print(f"  [anchor] {path.name} is a scan needing OCR, but pdftoppm/"
              "tesseract are missing (brew install poppler tesseract).",
              file=sys.stderr)
        return []
    if not is_pdf and not shutil.which("tesseract"):
        print(f"  [anchor] {path.name} needs OCR, but tesseract is missing "
              "(brew install tesseract).", file=sys.stderr)
        return []

    best: list[Anchor] = []
    best_note = ""
    best_key = None
    best_reconciled = False
    best_surya = False
    best_whole = False
    added = 0
    # Every reading of every pass, kept so the winner can be topped up from the
    # names the other readings found and it did not.
    pool: list[tuple[str, list[Anchor]]] = []
    best_total: int | None = None
    # Each render pass is read TWICE for a PDF: once from the page exactly as
    # rendered, and once from a cleaned copy. Neither is reliably better - see
    # _ocr_ready - so both are offered to the selection below and both stay in
    # the pool for the union, which makes the cleaning strictly additive.
    passes = (tuple((d, p, e) for d, p in OCR_PASSES for e in (False, True))
              if is_pdf else ((0, "6", False), (0, "4", False),
                              (0, "11", False)))
    # Surya OCR reads first when available. When it reconciles, one tesseract
    # pass still runs to populate the pool for cross-engine corroboration in
    # the union step; the remaining passes are skipped.
    if is_pdf and _surya_available():
        passes = (("surya", "doc", False), *passes)
    if is_pdf and _paddle_available():
        passes = (("paddle", "doc", False), *passes)
    _surya_did_reconcile = False
    for dpi, psm, enhance in passes:
        cands = []
        surya = dpi == "surya"
        paddle = dpi == "paddle"
        if paddle:
            paddle_pages = _paddle_ocr_pages(path)
            if not paddle_pages:
                continue
            text = "\n".join(ln["text"] for pg in paddle_pages for ln in pg)
            stated0 = _stated_total_shares(text)
            cands.append(_parse_boxed_anchor_table(paddle_pages))
            if not any(cands):
                cands.append(_parse_ocr_anchor_table(text, stated0))
        elif surya:
            surya_pages = _surya_ocr_pages(path)
            if not surya_pages:
                continue
            text = "\n".join(ln["text"] for pg in surya_pages for ln in pg)
            stated0 = _stated_total_shares(text)
            cands.append(_parse_boxed_anchor_table(surya_pages))
            if not any(cands):
                cands.append(_parse_ocr_anchor_table(text, stated0))
        elif is_pdf:
            # Three readings of the same page, strongest first. The ruled-cell
            # reader knows the true row AND column of every word but needs a
            # printed grid; the positioned-line reader infers rows from vertical
            # gaps and works on borderless tables; the flat-text reader is the
            # last resort. All three are offered to the same selection below, so
            # a weaker tier can only win by reconciling or by finding more
            # investors than the stronger one did.
            pages = _ocr_pdf_boxes(path, dpi, psm, enhance)
            text = "\n".join(ln["text"] for pg in pages for ln in pg)
            stated0 = _stated_total_shares(text)
            cands.append(_parse_grid_anchor_table(
                _ocr_pdf_grid(path, dpi, psm, enhance), stated0))
            cands.append(_parse_boxed_anchor_table(pages))
            if not any(cands):
                text = _ocr_pdf(path, dpi, psm, enhance)
                cands.append(_parse_ocr_anchor_table(
                    text, _stated_total_shares(text)))
        else:
            text = _tesseract(path, psm)
            cands.append(_parse_ocr_anchor_table(text, _stated_total_shares(text)))
        if not text.strip():
            continue
        stated = _stated_total_shares(text)
        stated_price = _stated_anchor_price(text)
        for rows in cands:
            if not rows:
                continue
            pass_note = ""
            price = _resolve_allocation_price(rows, stated_price)
            # Order matters: repair the share cells first, THEN look for the run
            # of rows that reconciles. Trimming first would cut the table short
            # at the first mis-read figure.
            #
            # A table whose shares ALREADY sum to the declared allocation is
            # not repaired at all. There is nothing left to prove, and every
            # repair is a guess that can only break the proof: on HORIZONIND
            # the recogniser read all 54 rows to exactly 194,625,000, the
            # amount column then "mended" one cell that was already right, and
            # the filing landed 138,000 shares short. Having lost its
            # reconciliation it lost the ranking too, and 34 rows of tesseract
            # debris - "|CARMIGNAC EMERGENTS" among them - took its place.
            fixed = 0
            if not (stated and sum(s or 0 for _, s, _ in rows) == stated):
                rows, fixed = _correct_shares(rows, price)
                rows, mended = _repair_row_from_percent(rows, stated, price)
                fixed += 1 if mended else 0
            # No single anchor can hold more shares than the entire anchor
            # allocation, so a cell that says otherwise was mis-read and its
            # value is worth nothing. Dropping it keeps the ROW - the investor's
            # name is still good - while stopping one garbled figure from
            # dominating the arithmetic below. GKENERGY lost five real names
            # without this: its 10-name reading carried one absurd share cell,
            # which made the reading look like it had double-counted, and a
            # 5-name reading won instead.
            if stated:
                rows = [(n, (s if s is not None and s <= stated else None), t)
                        for n, s, t in rows]
            # A reading whose shares sum EXACTLY to the declared total is the
            # best figure source there is, and it is ranked first below. It is
            # NOT, however, the last word on WHO the anchors are, and this used
            # to return here on the spot. That return threw away every other
            # reading of the same pages unread. On CPPLUS an eight-name reading
            # hit the total exactly and won outright; mining the readings it had
            # suppressed put the file at 79 names with the arithmetic unchanged,
            # so the shortcut was costing 71 real investors on that filing
            # alone. Reconciling now wins the ranking instead of skipping it,
            # and the union below still runs.
            exact = _trim_to_total(rows, stated)
            reconciled = exact is not None
            # Did the reconciling run explain the WHOLE reading, or only a
            # prefix of it? The distinction decides whether the union below is
            # needed. A full-table match accounts for every row that was read,
            # so nothing is outstanding; a short prefix that happens to hit the
            # total leaves the rest of the page unexplained, and on CPPLUS just
            # such an 8-row coincidence once hid 71 real investors.
            whole = reconciled and len(exact) == len(rows)
            _engine_label = ("Surya OCR" if surya else
                             "PaddleOCR" if paddle else
                             f"OCR {dpi}dpi/psm{psm}"
                             + (" cleaned" if enhance else ""))
            _engine_tag = ("surya" if surya else
                           "paddle" if paddle else "tesseract")
            if reconciled:
                rows = exact
                pass_note = (_engine_label
                             + f" -> shares reconcile to {stated:,}"
                             + (f" at Rs {price:,.2f}/share" if price else
                                "; NO price could be resolved")
                             + (f" ({fixed} share cell(s) repaired from the "
                                f"amount column)" if fixed else ""))
            anchors = _finalize_anchors(rows, stated_price)
            got = sum(r[1] or 0 for r in rows)
            gap = abs(stated - got) if stated else None
            over = bool(stated) and got > stated * 1.02
            key = (1 if reconciled else 0, 0 if over else 1, len(anchors),
                   -(gap if gap is not None else 0),
                   1 if (surya or paddle) else 0)
            pool.append((_engine_tag, anchors))
            if best_key is None or key > best_key:
                best = anchors
                best_key = key
                best_total = stated or None
                best_reconciled = reconciled
                best_surya = surya
                best_whole = whole
                best_note = (pass_note if reconciled else
                             (f"{len(anchors)} names, shares total {got:,} vs "
                              f"stated {stated:,} (off by {stated - got:+,})"
                              if stated else
                              f"{len(anchors)} names, no stated total to check "
                              f"against"))
        if (surya or paddle) and best_reconciled:
            _surya_did_reconcile = True
            continue
        if _surya_did_reconcile and not surya and not paddle:
            break
    if best:
        # One reading is never enough. Each tier and each DPI fails on DIFFERENT
        # rows, so whichever reading wins the ranking above still drops whatever
        # IT personally mis-read. On JAINREC the winning grid reading had
        # garbled the three rows the boxed reading got clean, and SANSHI FUND-I,
        # SINGULARITY EQUITY FUND I and SOCIETE GENERALE - ODI vanished from the
        # workbook entirely. Picking a winner is therefore only half the job:
        # the winner supplies the figures, and every other reading is then
        # mined for investors it alone found.
        #
        # Not, however, when the winning reading already accounts for the whole
        # declared allocation, row for row. Completeness is proved by
        # arithmetic there — no investor can be missing from a table whose
        # shares add up, and no row is left over to suggest one was — so mining
        # the other passes can only add names the table does not contain. That
        # union is how "gna EMERGING MARKETS FUND BLACKROCK GLOBAL FUNDS Ls
        # lane" and "|CARMIGNAC EMERGENTS" reached the workbook.
        added = 0
        if not (best_reconciled and best_whole):
            best, added = _union_anchor_readings(best, pool, best_total)
        best = _split_glued_anchors(best)
        if added:
            best_note += (f"; {added} further name(s) recovered from the other "
                          f"readings of the same pages")
        if best_reconciled:
            print(f"  [anchor] {path.name}: {best_note}; {len(best)} names.")
        else:
            print(f"  [anchor] {path.name}: OCR UNVERIFIED — {best_note}. Names "
                  "are kept but the figures on this filing need checking.",
                  file=sys.stderr)
    # Only a reading that BOTH reconciled and was left alone by the union is
    # proven complete; the union deliberately adds names the arithmetic cannot
    # vouch for.
    return _seal(path, best, best_reconciled and not added)


# A recovered name has to earn its place. Anything mined out of a LOSING OCR
# reading is, by construction, text the ranking already judged less trustworthy,
# so admitting it blindly would refill the frequency table with the page
# furniture and character soup this whole exercise exists to remove — MEESHO's
# worst pass alone offers rows like "100 ARG it OGG NAD".
_NAME_JUNK = re.compile(r"[^A-Za-z0-9&,.\-/() ]")


def _plausible_investor_name(name: str) -> bool:
    """True if `name` reads like an institution rather than OCR debris.

    Deliberately strict, because the cost is asymmetric: a real investor missed
    here is still recoverable from the winning reading or from IPOPlatform,
    whereas a fake one entering the register corrupts every recurrence count
    derived from it."""
    s = re.sub(r"\s+", " ", str(name or "")).strip()
    s = re.sub(r"^\d{1,3}[\s.)\-]+", "", s).strip()      # leading row serial
    if len(s) < 12:
        return False
    words = s.split()
    if len(words) < 2:
        return False
    letters = sum(c.isalpha() for c in s)
    if letters < 10 or letters / len(s) < 0.72:
        return False
    if len(_NAME_JUNK.findall(s)) > 1:
        return False
    # OCR debris is short, mixed-case gibberish: "SEL Ea LAS SU VANTAGE",
    # "ARG it OGG NAD". Real names are dominated by proper words.
    real = [w for w in words if len(w) >= 3 and w.isalpha()]
    if len(real) < 2 or len(real) / len(words) < 0.6:
        return False
    # An institution says what it is somewhere in its name. Every anchor is a
    # fund, trust, insurer, bank or incorporated body.
    return bool(re.search(
        r"\b(FUND|TRUST|CAPITAL|INVEST\w*|SECURITIES|INSURANCE|BANK|ASSET\w*|"
        r"PARTNERS|ADVISOR\w*|HOLDINGS?|VENTURES?|EQUIT\w+|PORTFOLIO|SCHEME|"
        r"LIMITED|LTD|LLP|PLC|INC|PTE|PCC|VCC|AIF|MUTUAL|ODI|"
        r"GMBH|SA|AG|BV|NV|AB|ASA|SICAV|UCITS|FCP|SPC|SAC|"
        r"FONDS|GESTION|CONSEIL|MANAGEMENT|FINANCE|FINANCIAL|"
        r"SOVEREIGN|PENSION|ENDOWMENT|FOUNDATION|GROUP|"
        r"CORP|CORPORATION|WEALTH|CREDIT|TRADING)\b", s, re.I))


def _union_anchor_readings(backbone: list[Anchor],
                           pool: list[tuple[str, list[Anchor]]],
                           stated: int | None) -> "tuple[list[Anchor], int]":
    """Add to `backbone` every plausible investor that only the other readings
    found. Returns (merged, how_many_added).

    Identity is `investor_key`, the same key the merge and the frequency table
    use, so a name recovered here cannot show up beside the spelling already
    present — completeness is bought without creating duplicates.

    Figures are carried over only while the arithmetic still permits it. Once
    the running total reaches the filing's declared allocation, further
    recovered rows keep the NAME and drop the share count: an unverifiable
    figure is worse than a blank one, and blank is already handled downstream.

    Pool entries are tagged with their engine name (``"surya"``,
    ``"paddle"``, or ``"tesseract"``). Cross-engine agreement (2+ distinct
    engines see the same name) is trusted with just 2 readings;
    same-engine-only agreement requires 3+, because same-engine readings
    share systematic OCR biases."""
    have = {investor_key(a.name) for a in backbone}
    have.discard(())
    # Second identity, for the case canonical_investor_tokens cannot see: OCR
    # dropping the spaces inside a name, which turns "SBI GENERAL INSURANCE"
    # into the single token "SBIGENERAL INSURANCE" and defeats token matching.
    squashed = {squash_investor_name(a.name) for a in backbone}
    squashed.discard("")
    merged = list(backbone)
    running = sum(a.shares or 0 for a in backbone)
    added = 0
    # CORROBORATION. A recovered name is admitted only if a SECOND, independent
    # reading of the same pages saw the same investor. This is the test that
    # keeps completeness and tidiness in one fix instead of trading them off: a
    # real table row is seen by several passes because it is really there, while
    # OCR debris is random and comes out differently every time — "ARG it OGG
    # NAD" is a one-off, "SINGULARITY EQUITY FUND I" is not. Without this the
    # register bloated from 596 investors to 838.
    #
    # Agreement is measured on shared significant tokens, not on the whole name,
    # because two passes rarely cut a cell the same way: GKENERGY's PineBridge
    # row arrives once as "PINEBRIDGE INDIA EQUITY FUND" and once as "PINEBRIDGE
    # GLOBAL FUNDS PINEBRIDGE INDIA EQUITY FUND". Demanding an identical token
    # tuple scored both as unique and threw the investor away; demanding two
    # tokens in common recognises them as one fund. Two is also enough to reject
    # coincidence — names sharing a single generic token do not corroborate.
    engines: dict[int, str] = {}
    token_readings: dict[str, set[int]] = {}
    for i, (engine, reading) in enumerate(pool):
        engines[i] = engine
        for a in reading:
            for t in _key_split(investor_key(a.name))[0]:
                token_readings.setdefault(t, set()).add(i)

    def _corroborated(k: tuple) -> bool:
        hits: dict[int, int] = {}
        for t in _key_split(k)[0]:
            for i in token_readings.get(t, ()):
                hits[i] = hits.get(i, 0) + 1
        supporters = [i for i, n in hits.items() if n >= 2]
        if len(supporters) < 2:
            return False
        eng_set = {engines[i] for i in supporters}
        if len(eng_set) >= 2:
            return True
        return len(supporters) >= 3

    for _engine, reading in pool:
        for a in reading:
            k = investor_key(a.name)
            q = squash_investor_name(a.name)
            if not k or q in squashed:
                continue
            if not _corroborated(k) or not _plausible_investor_name(a.name):
                continue
            # Already held — as the same name, as the same name minus its
            # series ordinal, or as a fragment of it cut differently by another
            # pass ("INDIA VALUE FUND" against "... A/C AXIS INDIA VALUE
            # FUND"). An ambiguous match, where the candidate straddles two
            # investors already held, is debris and is dropped for the same
            # reason.
            if _investor_matches(k, list(have)):
                continue
            have.add(k)
            squashed.add(q)
            shares = a.shares
            if stated and shares and running + shares > stated:
                shares = None
            if shares:
                running += shares
            merged.append(Anchor(name=re.sub(r"^\d{1,3}[\s.)\-]+", "",
                                             a.name).strip(),
                                 shares=shares, price=a.price,
                                 amount=(shares * a.price)
                                 if (shares and a.price) else None))
            added += 1
    return merged, added


# Anchor tables number their rows, so a row serial appearing in the MIDDLE of a
# name cell means two table rows were read as one: GKENERGY produced "PINEBRIDGE
# INDIA EQUITY FUND 3 HSBC FLEX! CAP FUND 4 3P INDIA EQUITY FUND IM", which is
# three investors, and JAINREC produced "SOCIETE GENERALE - ODI SANSHI FUND-I
# ... GENERAL INSURANCE". Left glued they are three fake entities instead of
# three real ones — the exact failure this rewrite is meant to end.
#
# The serial must sit between two word-like tokens to count; a trailing digit is
# part of the name ("Meru Investment Fund PCC-Cell 1", "Nexus Growth Fund Sch 2")
# and roman numerals are never treated as serials at all.
_INTERIOR_SERIAL = re.compile(r"(?<=[A-Za-z)\]])\s+(\d{1,3})\s+(?=[A-Za-z]{3})")


def _split_glued_anchors(anchors: list[Anchor]) -> list[Anchor]:
    """Break name cells that hold more than one investor into separate rows.

    The split only happens when EVERY resulting fragment reads like a real
    institution, so a name that merely contains a number survives intact. The
    share is read from one row of the table, and guessing which of the glued
    names it belonged to would be worse than leaving the others blank.

    No de-duplication happens here. It used to, and it silently deleted real
    investors: PINEBRIDGE INDIA EQUITY FUND was dropped from GKENERGY because a
    fragment produced earlier in the list happened to reduce to the same tokens.
    Splitting and de-duplicating are separate jobs, and the callers above have
    already done the second one."""
    out: list[Anchor] = []
    for a in anchors:
        parts = [p.strip(" .-") for p in _INTERIOR_SERIAL.split(str(a.name or ""))]
        # re.split keeps the captured serial; drop those.
        parts = [p for p in parts if p and not p.isdigit()]
        if len(parts) < 2 or not all(_plausible_investor_name(p) for p in parts):
            parts = [str(a.name or "")]
        for i, p in enumerate(parts):
            sh = a.shares if i == 0 else None
            out.append(Anchor(name=p, shares=sh, price=a.price,
                              amount=(sh * a.price) if (sh and a.price)
                              else (a.amount if i == 0 else None)))
    return out


def extract_anchors_from_file(path: Path, use_ocr: bool = True) -> list[Anchor]:
    """Anchor rows from an anchor-allocation filing, PDF or image."""
    frozen = _frozen_load(path)
    if frozen is not None:
        # Re-asserted so a reading served from the store still REPLACES the
        # workbook's rows rather than merging with them.
        _RECONCILED_FILINGS[path.name] = True
        print(f"  [anchor] {path.name}: reusing the stored reading "
              f"({len(frozen)} names) — it reconciled exactly, so it is final.")
        return frozen
    if path.suffix.lower() in IMAGE_SUFFIXES:
        return extract_anchors_via_ocr(path) if use_ocr else []
    return extract_anchors_from_pdf(path, use_ocr=use_ocr)


def _serial_gaps(seq: list[int]) -> set[int]:
    """Row numbers the table reader skipped, judged per table.

    An anchor letter usually carries a second table listing the mutual-fund
    subset, whose serials restart at 1. Measuring gaps over the combined
    sequence let that restart fill the holes in the first table and hid them,
    so the sequence is cut into ascending runs and each is checked alone."""
    gaps: set[int] = set()
    run: list[int] = []
    for s in seq:
        if run and s < run[-1]:
            gaps |= set(range(min(run), max(run) + 1)) - set(run)
            run = []
        run.append(s)
    if run:
        gaps |= set(range(min(run), max(run) + 1)) - set(run)
    return gaps


def _is_table_continuation(tbl: list, name_col: int) -> bool:
    """Is this header-less table the rest of the anchor table from the page
    before, rather than one of the little header fragments pdfplumber also
    reports ('Total Amount', '% of', 'Sr.')?

    Judged on the cells themselves: most rows must carry something name-shaped
    in the column the header said held the name. Matching the column count
    alone is not enough, because the stray fragments sometimes match it."""
    named = 0
    for row in tbl:
        if name_col >= len(row):
            return False
        if len(re.sub(r"[^A-Za-z]", "", str(row[name_col] or ""))) >= 8:
            named += 1
    return named >= max(1, (len(tbl) + 1) // 2)


def extract_anchors_from_pdf(path: Path, use_ocr: bool = True) -> list[Anchor]:
    """Parse an anchor-allocation PDF and return EVERY anchor investor row in
    document order, de-duplicated. Tries table extraction first (the anchor
    table has a 'Name of the Anchor Investor' column), then a numbered-line
    text fallback, and finally OCR for scans that carry no text layer."""
    try:
        import pdfplumber  # type: ignore
    except Exception:
        print("  [anchor] pdfplumber not installed; run: pip install pdfplumber",
              file=sys.stderr)
        return []

    rows: list[tuple[str, int | None, list[float]]] = []
    seen: set[str] = set()
    text_all: list[str] = []
    _SKIP = {"name", "nameoftheanchorinvestor", "anchorinvestor", "total",
             "grandtotal", "nameofanchorinvestor", "nameoftheinvestor",
             "investorname", "nameofinvestor", "srno", "sno", "particulars"}

    def _add(raw: str, nums: list[float]) -> None:
        n = re.sub(r"\s+", " ", (raw or "").strip()).strip(" .")
        if not n:
            return
        key = _norm_alnum(n)
        # A continuation table's row 0 is data, but a letter that restarts the
        # table part-way (Groww, INDO MIM) reprints the header there, and
        # "Name of Mutual Fund / Insurance Company" is not an investor.
        if not key or key in _SKIP or key in seen or key.startswith("nameof"):
            return
        # drop pure numbers / amounts / percentages
        if re.fullmatch(r"[\d,\.\s%()/-]+", n):
            return
        seen.add(key)
        shares = next((int(v) for v in nums
                       if v >= MIN_ANCHOR_SHARES and v.is_integer()), None)
        tail = nums[nums.index(float(shares)) + 1:] if shares is not None else []
        rows.append((n, shares, tail))

    try:
        with pdfplumber.open(str(path)) as pdf:
            table_hit = False
            # An anchor table runs over many pages but the COLUMN HEADER is
            # printed only once, on the first. pdfplumber therefore hands back
            # a continuation table whose row 0 is already data ('17', '46',
            # '80'...). Requiring a name column on every table silently threw
            # all of those away: INDO MIM kept 15 rows of 109 and discarded
            # 131, so the (perfectly clean) text reading failed to reconcile
            # and lost to the OCR mess. The last seen header is carried
            # forward instead.
            # ncols -> the name column seen in a header of that shape. Keyed
            # on the column count rather than kept as a single "last header",
            # because a letter can interleave two differently shaped tables
            # (Groww has an 8-column anchor table and an 18-column mutual-fund
            # one) and a single slot let the wrong shape evict the right one.
            header_cols: dict[int, int] = {}
            serial_seq: list[int] = []
            for page in pdf.pages:
                for tbl in (page.extract_tables() or []):
                    if not tbl:
                        continue
                    header = [_norm_alnum(c or "") for c in tbl[0]]
                    name_col = None
                    for i, h in enumerate(header):
                        if "name" in h and ("anchor" in h or "investor" in h):
                            name_col = i
                            break
                    if name_col is None:
                        for i, h in enumerate(header):
                            if h in ("name", "nameoftheinvestor", "investorname",
                                     "nameofinvestor", "nameoftheanchorinvestor"):
                                name_col = i
                                break
                    if name_col is not None:
                        header_cols[len(tbl[0])] = name_col
                        if len(tbl) < 2:
                            continue
                        body = tbl[1:]
                    else:
                        name_col = header_cols.get(len(tbl[0]))
                        if name_col is None or not _is_table_continuation(
                                tbl, name_col):
                            continue
                        body = tbl
                    table_hit = True
                    for row in body:
                        if name_col >= len(row):
                            continue
                        if name_col != 0:
                            s0 = re.fullmatch(r"\s*(\d{1,3})[\.\)]?\s*",
                                              str(row[0] or ""))
                            if s0:
                                serial_seq.append(int(s0.group(1)))
                        nums = _numbers_in(" ".join(
                            str(c or "") for i, c in enumerate(row)
                            if i != name_col))
                        _add(row[name_col] or "", nums)
            for page in pdf.pages:
                text_all.append(page.extract_text() or "")

            def _line_pass(only_serials: set[int] | None = None,
                           budget: int | None = None) -> None:
                for line in "\n".join(text_all).splitlines():
                    m = re.match(r"^\s*(\d{1,3})[\.\)]?\s+(.+?)\s+[\d,]{3,}",
                                 line)
                    if not m:
                        continue
                    if (only_serials is not None
                            and int(m.group(1)) not in only_serials):
                        continue
                    # Strip exactly the serial that was matched. The generic
                    # serial regex is not safe here: it would turn the real
                    # name "360 ONE FLEXICAP FUND" into "ONE FLEXICAP FUND".
                    nm, shares, tail = _split_row(
                        line[m.end(1):].lstrip(" .)\t"))
                    if shares is None:
                        _add(m.group(2), [])
                        continue
                    # A gap serial can match a line in the mutual-fund
                    # sub-table as well as the anchor table, and that sub-table
                    # re-lists investors already counted. The declared
                    # allocation is the arbiter: a candidate that would push
                    # the running total past it is not a missing row.
                    if budget is not None:
                        if shares > budget:
                            continue
                        budget -= shares
                    _add(nm, [float(shares), *tail])

            # The table reader is the better of the two — it splits columns
            # exactly — but pdfplumber's ruled-region detection drops a row
            # wherever the table straddles a page break: INDO MIM's row 16,
            # "SBI CHILDREN'S FUND - SAVINGS PLAN", is plainly in the page-2
            # text yet falls outside the detected table, and one row is lost
            # at every page boundary that way.
            #
            # Those losses show up as HOLES IN THE SERIAL SEQUENCE, so the
            # line reader is asked only for the missing serials rather than
            # being unioned wholesale. Turning it loose on the whole document
            # took INDO MIM from 90 names to 208: it re-reads the mutual-fund
            # sub-table (a subset of the anchor table, not new investors) and
            # splits wrapped names into fragments that do not de-duplicate.
            if not table_hit:
                _line_pass()
            else:
                stated_now = _stated_total_shares("\n".join(text_all))
                if (not stated_now
                        or sum(r[1] or 0 for r in rows) != stated_now):
                    _line_pass(_serial_gaps(serial_seq),
                               (stated_now - sum(r[1] or 0 for r in rows))
                               if stated_now else None)
    except Exception as e:
        print(f"  [anchor] failed to parse {path.name}: {e}", file=sys.stderr)

    text = "\n".join(text_all)
    if not rows:
        # No text layer at all (a pure scan) — or a text layer too mangled to
        # yield a single name. Fall back to OCR.
        return extract_anchors_via_ocr(path) if use_ocr else []

    anchors = _finalize_anchors(rows, _stated_anchor_price(text))
    stated = _stated_total_shares(text)
    got = sum(a.shares or 0 for a in anchors)
    if not use_ocr or (stated and got == stated):
        return _seal(path, anchors, bool(stated) and got == stated)
    # The text layer yielded SOMETHING, but not a table that reconciles. That
    # used to end it, and it silently lost investors: SACHEEROME's text layer
    # gives 2 names where OCR reads 5 (Bharat Venture Opportunities, Sanshi,
    # Finavenue, Shine Star, Chartered Finance & Leasing), because half the
    # letter is a scanned image pasted into a text PDF. Neither source is
    # trustworthy on its own, so read both and keep the one that reconciles —
    # or, failing that, the one that found more investors.
    ocr = extract_anchors_via_ocr(path)
    if not ocr:
        return _seal(path, anchors, False)
    if stated and sum(a.shares or 0 for a in ocr) == stated:
        return ocr
    # A REAL TABLE in a REAL TEXT LAYER beats OCR even when it falls short of
    # the declared total, and it must not be judged on row count. Checked name
    # by name on INDO MIM, Groww, Urban Company and Kissht: every name OCR
    # "found" that the text reading lacked was already in the text layer under
    # a garbled spelling — HSBC SMALL CAP, ENAM, Goldman Sachs, Kedaara,
    # Nippon Small Cap, Edelweiss all sit there in plain text. OCR contributed
    # no investor at all, only noise ("ON SNORT EUND Oe 100", "ATOR
    # LONG-SHORTEUND", "Bangalore, Bangalore South, Karnataka, India"), and
    # winning on count is exactly how that noise used to get into the workbook.
    if table_hit and len(anchors) >= 10:
        if len(ocr) > len(anchors):
            print(f"  [anchor] {path.name}: OCR read {len(ocr)} row(s) against "
                  f"{len(anchors)} from the text layer — keeping the text "
                  "layer; the surplus is OCR noise, not investors.")
        return _seal(path, anchors, False)
    # Counted on names that read like institutions, not on rows. A 30-row OCR
    # reading that is half debris used to beat a clean 12-row text reading on
    # OPTIMYSTIX, PATIL, SACHEEROME and KISSHT purely on volume.
    good_txt = sum(1 for a in anchors if _plausible_investor_name(a.name))
    good_ocr = sum(1 for a in ocr if _plausible_investor_name(a.name))
    if good_ocr > good_txt:
        print(f"  [anchor] {path.name}: text layer read {len(anchors)} name(s) "
              f"of which {good_txt} read as investors, OCR read {len(ocr)} of "
              f"which {good_ocr} do — keeping the OCR reading.")
        return _seal(path, ocr, False)
    return _seal(path, anchors, False)


# ─────────────────── IPOPlatform: anchor books without OCR ──────────────────
# Everything above this line reads the anchor table off a scanned PDF, and no
# amount of image work makes that reliable: the filings are photocopies of
# faxes, and a mis-read character silently becomes a fake investor. IPOPlatform
# (IntelliFin, a Chittorgarh company) publishes the same books as structured
# HTML, free and unpaywalled, and — the part that matters — every anchor row
# links to a canonical /anchor-investor/<slug>/<id> page. That numeric id is a
# real identity key, so "SANSHI FUND-I" and "Sanshi Fund-I" are provably the
# same investor instead of probably the same investor. No fuzzy matching, no
# name-key heuristics, no invented entities.
#
# THE LIMIT, stated plainly: they carry anchor NAMES for SME issues only. Their
# database has the mainboard anchor TOTAL (`anchor_investor_shares_offered` is
# populated for Meesho, Groww, Jain Resource and others) but the name list is
# absent from every mainboard page checked — 0 names across all 13 mainboard
# issues in this workbook, including listed ones. Their own master list is
# titled "SME IPO Anchor Investors List" and their investor pages count only
# SME participations. Mainboard therefore still comes from the PDFs above.
IPL_BASE = "https://www.ipoplatform.com"
IPL_CACHE_DIR = CACHE_DIR / "ipoplatform"
IPL_TTL_HOURS = 24 * 7
# An anchor book is FINAL once the issue has listed — allocations cannot change
# afterwards — so its page is fetched once and then served from disk forever.
# Without this the weekly TTL re-downloads all ~414 in-window pages on every run
# a week apart, which is a full reload of data that is known not to have moved.
IPL_TTL_SETTLED = 24 * 365 * 20
IPL_DELAY = 0.8
# The IPO index is a Laravel DataTables endpoint that returns the whole row, so
# one request with a large `length` replaces walking 1,520 paginated pages.
IPL_INDEX_URL = IPL_BASE + "/main-board/index"
# The anchor register and each investor's own deal list, both DataTables feeds.
# `/invested-ipos` dates every participation, which is what makes a count over
# a window possible without reading one filing per issue.
IPL_PARENT_URL = IPL_BASE + "/parent/datatable"
IPL_INVESTED_URL = IPL_BASE + "/invested-ipos"


def _ipl_get(url: str, params: dict | None = None,
             ttl_hours: float = IPL_TTL_HOURS) -> str:
    """Cached, paced GET. Returns "" on any failure, which callers treat as
    'no data' and fall back to the PDF reader.

    ONLY successful responses are cached. An earlier version wrote the empty
    body of a failed request to the cache file too, which meant one transient
    5xx silently disabled this source for the whole TTL — a week of quietly
    reverting to OCR with no error to see. Now a failure leaves the last good
    copy in place and returns it, so a site outage degrades to stale-but-real
    data first and to the PDF reader only when there is nothing cached at
    all."""
    IPL_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    key = url + ("?" + urlencode(sorted(params.items())) if params else "")
    fp = IPL_CACHE_DIR / (hashlib.sha1(key.encode()).hexdigest() + ".html")
    stale = ""
    if fp.exists():
        stale = fp.read_text(encoding="utf-8")
        if time.time() - fp.stat().st_mtime < ttl_hours * 3600:
            return stale
    headers = {"User-Agent": UA, "Accept-Language": "en-US,en;q=0.9"}
    if params is not None:
        headers["X-Requested-With"] = "XMLHttpRequest"
        headers["Referer"] = IPL_BASE + "/ipo-insights/exchange-wise-ipos"
    try:
        r = requests.get(url, headers=headers, params=params, timeout=90)
    except requests.RequestException as exc:
        print(f"  [ipoplatform] {url}: {exc}"
              f"{' — serving the cached copy' if stale else ''}",
              file=sys.stderr)
        return stale
    time.sleep(IPL_DELAY)
    if r.status_code != 200:
        print(f"  [ipoplatform] {url}: HTTP {r.status_code}"
              f"{' — serving the cached copy' if stale else ''}",
              file=sys.stderr)
        return stale
    try:
        fp.write_text(r.text, encoding="utf-8")
    except OSError:
        pass
    return r.text


def _html_cells(row_html: str) -> list[str]:
    """Visible text of every cell in one <tr>, whitespace collapsed."""
    cs = re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", row_html, re.S | re.I)
    return [re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", c)).strip() for c in cs]


def _html_rows(html: str, table_id: str | None = None):
    """Yield the <tr> fragments of each <table>, optionally only the one whose
    id matches."""
    for t in re.findall(r"<table[^>]*>.*?</table>", html, re.S | re.I):
        if table_id is not None:
            m = re.search(r'id="([^"]+)"', t[:300])
            if not m or m.group(1) != table_id:
                continue
        yield re.findall(r"<tr[^>]*>(.*?)</tr>", t, re.S | re.I)


_IPL_INV_ID = re.compile(r'href="[^"]*?/anchor-investor/[^"/]+/(\d+)"')


def _ipl_num(text: str) -> float | None:
    """'₹ 1,07,730.00' / '798,000.00' -> float. None if there is no number."""
    s = re.sub(r"[^\d.]", "", str(text or ""))
    if not s or s.count(".") > 1:
        s = s.split(".")[0] if s else ""
    try:
        return float(s)
    except ValueError:
        return None


@lru_cache(maxsize=1)
def _ipl_index_rows() -> tuple:
    """Every raw index row IPOPlatform serves, both segments (~2,300).

    Kept separate from `ipl_index` because that view is keyed on the NSE symbol
    and so cannot represent the 600-odd BSE-SME-only issues, which have none."""
    rows: list[dict] = []
    for seg in ("SME", "Mainboard"):
        body = _ipl_get(IPL_INDEX_URL, {
            "draw": 1, "start": 0, "length": 5000, "ipo_type": seg,
            "selected_exchange": "all", "search[value]": "",
            "search[regex]": "false"})
        if not body:
            continue
        try:
            data = json.loads(body).get("data") or []
        except json.JSONDecodeError:
            print(f"  [ipoplatform] index for {seg} was not JSON", file=sys.stderr)
            continue
        for row in data:
            row["_seg"] = seg
            rows.append(row)
    return tuple(rows)


@lru_cache(maxsize=1)
def ipl_index() -> dict:
    """NSE symbol -> {id, url, seg, anchor_shares} for every SME and mainboard
    IPO IPOPlatform knows about (~2,300).

    Keyed on `nse_script_symbol` rather than the company name on purpose: the
    name arrives as "Onemi Technology Solutions" where the exchange says
    KISSHT, and as "Urbanclap Technologies India" where it says URBANCO. The
    symbol matched all 54 issues in this workbook on the first attempt; name
    matching would have needed a synonym table."""
    out: dict[str, dict] = {}
    for row in _ipl_index_rows():
        sym = str(row.get("nse_script_symbol") or "").strip().upper()
        if not sym:
            continue
        m = re.search(r'href="([^"]+)"', str(row.get("company_link") or ""))
        out[sym] = {
            "id": row.get("id"),
            "url": m.group(1) if m else f"{IPL_BASE}/ipo/x/{row.get('id')}",
            "seg": row.get("_seg"),
            "company": row.get("company_name"),
            "anchor_shares": row.get("anchor_investor_shares_offered"),
        }
    return out


@lru_cache(maxsize=1)
def ipl_bse_index() -> dict:
    """`_bse_name_key(company)` -> the same record shape `ipl_index` returns, for
    the ~670 BSE-SME issues that have no NSE symbol to be keyed on.

    Needed because both of `ipl_resolve`'s paths read `ipl_index`, so a BSE-only
    issue resolved to None and `_apply_ipoplatform` skipped it before it even
    reached the `no_book` list — 0 of 22 BSE gainers got an anchor book and
    nothing in the log said so. The name is a safe key here in a way it is not
    against the exchanges: this is IPOPlatform's own spelling matched back to
    IPOPlatform, and it is the very string `collect_bse_sme` copied into
    `IPO.company`.

    A key two issues share is dropped rather than guessed at, for the reason
    `ipl_resolve` gives: the wrong issuer's book is worse than no book."""
    out: dict[str, dict] = {}
    clash: set[str] = set()
    for row in _ipl_index_rows():
        if str(row.get("nse_script_symbol") or "").strip():
            continue
        if not str(row.get("exchange") or "").upper().startswith("BSE"):
            continue
        name = _html.unescape(str(row.get("company_name") or ""))
        key = _bse_name_key(name)
        if len(key) < 5 or key in clash:
            continue
        if key in out:
            del out[key]
            clash.add(key)
            continue
        m = re.search(r'href="([^"]+)"', str(row.get("company_link") or ""))
        out[key] = {
            "id": row.get("id"),
            "url": m.group(1) if m else f"{IPL_BASE}/ipo/x/{row.get('id')}",
            "seg": "BSE SME",
            "company": name,
            "anchor_shares": row.get("anchor_investor_shares_offered"),
        }
    return out


@lru_cache(maxsize=1)
def ipl_parents() -> tuple:
    """(parent id, canonical name, all-time IPO count) for every anchor
    investor IPOPlatform tracks.

    Replaces a scrape of /anchor-investors that returned nothing. That page
    used to ship its rows as HTML, one /anchor-investor/<slug>/<id> link each;
    it now renders client-side from /parent/datatable and carries no rows and
    no links at all. The scrape therefore yielded an EMPTY register, and since
    `_apply_ipoplatform` treats an empty register as 'unavailable' and returns,
    this entire source had been silently switched off.

    The register is also now keyed on PARENT entities rather than individual
    vehicles, so these ids do NOT share a number space with the investor ids in
    an IPO page's anchor links — looking one up with the other returns a
    different investor's name."""
    body = _ipl_get(IPL_PARENT_URL, {"draw": 1, "start": 0, "length": 5000})
    try:
        rows = json.loads(body or "").get("data") or []
    except ValueError:
        return ()
    out = []
    for r in rows:
        name = re.sub(r"\s+", " ",
                      re.sub(r"<[^>]+>", " ", str(r.get("name") or ""))).strip()
        if name and r.get("id") is not None:
            out.append((str(r["id"]), name, int(_ipl_num(r.get("total_ipos")) or 0)))
    return tuple(out)


def ipl_parent_ipos(parent_id: str) -> tuple:
    """(company name, listing date) for every IPO one investor has anchored.

    Both categories are requested because the endpoint reads anything that is
    not the literal 'sme' as mainboard, so a single call silently returns the
    mainboard book alone — 26 of Rajasthan Global's 187."""
    out: list[tuple[str, str]] = []
    for cat in ("sme", "mainboard"):
        body = _ipl_get(IPL_INVESTED_URL,
                        {"parent_id": str(parent_id), "ipo_category": cat,
                         "draw": 1, "start": 0, "length": 5000})
        try:
            rows = json.loads(body or "").get("data") or []
        except ValueError:
            continue
        for r in rows:
            comp = str(r.get("company_name") or "").strip()
            when = str(r.get("ipo_year") or "")[:10]
            if comp and re.fullmatch(r"\d{4}-\d{2}-\d{2}", when):
                out.append((comp, when))
    return tuple(out)


def _ipl_label_matches(pkey: tuple, ours: dict) -> set:
    """Our labels for one of IPOPlatform's parent names.

    A match is a full prefix in either direction: their "Rajasthan Global
    Securities" against our "RAJASTHAN GLOBAL SECURITIES LIMITED", and their
    bare "Finavenue" against our "FINAVENUE CAPITAL TRUST GROWTH FUND". Their
    register names the HOUSE, so most of them are one or two brand words and
    demanding a longer overlap matched almost none of them.

    Requiring the whole of the shorter key to match is what keeps this safe:
    HDFC BANK and HDFC MUTUAL FUND share a first token and still do not meet,
    and neither do GOVERNMENT OF SINGAPORE and GOVERNMENT OF INDIA."""
    hits: set = set()
    for k, labels in ours.items():
        n = min(len(k), len(pkey))
        if n and k[:n] == pkey[:n]:
            hits |= labels
    return hits


def ipl_participation(anchor_map: dict, label, start: str, end: str) -> dict:
    """Our investor label -> every company it anchored in the window, per
    IPOPlatform, as normalised company cores.

    This is the denominator the workbook cannot see. Anchor filings are only
    ever downloaded for issues that CLEARED the gain threshold, so counting
    participations from our own books would divide winners by winners and hand
    every investor a hit rate near 100%. IPOPlatform dates every participation,
    so the full count is a filter rather than 300-odd more filings to read.

    Counted under OUR grouping, not theirs: their register rolls vehicles up to
    a parent entity while ours clusters spellings, and the two disagree. Several
    of their parents can land on one of our labels, so the companies are unioned
    and an issue both name is still counted once."""
    ours: dict[tuple, set] = {}
    for anchors in anchor_map.values():
        for a in anchors:
            nm = str(a.name).strip()
            lab = label(nm)
            if not lab:
                continue
            k = investor_key(nm)
            if k:
                ours.setdefault(k, set()).add(lab)
    parents = ipl_parents()
    if not parents:
        print("  [ipoplatform] register unavailable — no participation counts.",
              file=sys.stderr)
        return {}
    # Only parents that match something of ours are fetched; the rest would be
    # two requests each for a name the workbook never mentions.
    wanted = [(pid, nm) for pid, nm, _ in parents
              if _ipl_label_matches(investor_key(nm), ours)]
    print(f"  [ipoplatform] {len(wanted)} of {len(parents)} register entries "
          f"match an investor in the workbook; reading their deal lists ...")
    out: dict[str, set] = {}
    for n, (pid, nm) in enumerate(wanted, 1):
        labels = _ipl_label_matches(investor_key(nm), ours)
        for comp, when in ipl_parent_ipos(pid):
            if start <= when <= end:
                core = _company_core(comp)
                if core:
                    for lab in labels:
                        out.setdefault(lab, set()).add(core)
        if n % 50 == 0:
            print(f"            {n}/{len(wanted)} ...")
    print(f"  [ipoplatform] participation counts for {len(out)} investor(s).")
    return out


def ipl_resolve(symbol: str, company: str = "") -> "dict | None":
    """IPOPlatform's record for an issue, by NSE symbol or failing that by name.

    The index is keyed on `nse_script_symbol`, so anything this script filed
    under a label of its own — a filename-derived key, or the Company cell of a
    workbook row — never resolves on the symbol alone. Poojaa Precision is filed
    here as POOJAAPRECISION and there as PPEL, Merritronix as MRTX; both have a
    full anchor book that the symbol lookup simply could not see.

    A name match is only accepted when exactly ONE issue matches. Filing one
    issuer's anchor book under another issuer's symbol is worse than leaving the
    figures blank, so an ambiguous name is treated as no match at all."""
    idx = ipl_index()
    hit = idx.get(str(symbol or "").strip().upper())
    if hit:
        return hit
    # BSE's exact whole-name match is tried before the NSE prefix match, which
    # is the weaker evidence of the two and was matching BSE issues onto
    # unrelated NSE rows that merely began with the same words.
    bse = ipl_bse_index()
    for cand in (company, symbol):
        hit = bse.get(_bse_name_key(cand))
        if hit:
            return hit
    for cand in (company, symbol):
        k = _norm_alnum(cand)
        if len(k) < 5:
            continue
        named = [i for i in idx.values()
                 if _norm_alnum(i.get("company") or "").startswith(k)]
        if len(named) == 1:
            return named[0]
    return None


def ipl_anchors(symbol: str) -> "list[tuple[Anchor, str]]":
    """(Anchor, investor_id) for one NSE symbol, or [] if IPOPlatform has no
    anchor list for it."""
    info = ipl_index().get(str(symbol or "").strip().upper())
    return _ipl_anchors_for(info) if info else []


def _ipl_deslug(row_html: str) -> str:
    """Full investor name out of the /anchor-investor/<slug>/<id> link.

    The slug is generated from the untruncated name, so it carries the words
    the display cell drops."""
    m = re.search(r'/anchor-investor/([^/"\']+)/\d+', row_html)
    return re.sub(r"[-_]+", " ", m.group(1)).strip().upper() if m else ""


def _ipl_anchors_for(info: dict,
                     ttl_hours: float = IPL_TTL_HOURS) -> "list[tuple[Anchor, str]]":
    """(Anchor, investor_id) for an already-resolved index record.

    The name on the IPO page is TRUNCATED for display — "SageOne India
    Opportunity" for what is really "SageOne India Opportunity Trust" — so the
    missing words are recovered from the slug in the row's own link. This used
    to be a lookup in the anchor register, which no longer publishes investor
    ids at all; worse, the ids it publishes now belong to parent entities, so
    the same lookup would silently attach another investor's name.

    The slug only WINS where it extends the display string. Equal readings keep
    the display text, which has the punctuation and case the slug flattened."""
    body = _ipl_get(info["url"], ttl_hours=ttl_hours)
    if not body:
        return []
    out: list[tuple[Anchor, str]] = []
    for rows in _html_rows(body):
        if not rows:
            continue
        head = _html_cells(rows[0])
        if not head or "anchor investor" not in head[0].lower():
            continue
        for row in rows[1:]:
            c = _html_cells(row)
            if len(c) < 4 or not c[0]:
                continue
            m = _IPL_INV_ID.search(row)
            inv_id = m.group(1) if m else ""
            name = c[0].strip()
            full = _ipl_deslug(row)
            if full:
                dk, fk = investor_key(name), investor_key(full)
                if len(fk) > len(dk) and fk[:len(dk)] == dk:
                    name = full
            shares = _ipl_num(c[1])
            price = _ipl_num(c[2])
            amount = _ipl_num(c[3])
            out.append((Anchor(
                name=name,
                shares=int(shares) if shares else None,
                price=price or None,
                # Consistent with the PDF path: shares x price is exact, the
                # amount column is only a fallback.
                amount=(shares * price) if (shares and price) else amount,
            ), inv_id))
        break
    return out


# ── the tracked watchlist ──────────────────────────────────────────────────
# Spelling variants of one investor are grouped under a canonical name. Only
# forms that denote the SAME party are grouped: word-order and punctuation
# variants of a person, and funds of a single house. Distinct people are kept
# apart even where they are related or share a firm — Mukul and Asha Agrawal,
# Madhusudan and Madhuri Kela, and the three Kedias each stand alone, and
# Param Capital and Cohesion are listed as themselves rather than folded into
# the individual behind them. Merge any of those by moving a line.
TRACKED_INVESTORS: "dict[str, tuple[str, ...]]" = {
    "Ajay Kumar Aggarwal": ("AJAY KUMAR AGGARWAL",),
    "Ajay Upadhyaya": ("AJAY UPADHYAYA", "UPADHYAYA AJAY",
                       "UPADHYAYA AJAY SHIV NARAYAN"),
    "Akash Bhanshali": ("AKASH BHANSHALI",),
    "Ankit Vijay Kedia": ("Ankit Vijay Kedia",),
    "Ankush Kedia": ("ANKUSH KEDIA", "ANKUSH  KEDIA"),
    "Vijay Krishanlal Kedia": ("Vijay Krishanlal Kedia",),
    "Kedia Securities": ("Kedia Secuirities Private Limited",),
    "Ashish Kacholia": ("ASHISH KACHOLIA", "ASHISH RAMESH KACHOLIA",
                        "ASHISH RAMESHCHANDRA KACHOLIA", "KACHOLIA ASHISH"),
    "Bengal Finance & Investment": ("BENGAL FIN. & INV. PVT. LTD",
                                    "BENGAL FINANCE & INVESTMENT PRIVATE LIMITED"),
    "Suryavanshi Commotrade": ("SURYAVANSHI COMMOTRADE PVT LTD",
                               "Suryavanshi Commotrade Private Limited",
                               "SURYA VANSHI COMMOTRADE PVT. LTD."),
    "Himalaya Finance & Investment": ("HIMALAYA FINANCE & INV. CO",
                                      "HIMALAYA FINANCE & INVESTMENT COMPANY",
                                      "HIMALAYA FINANCE AND INVESTMENT CO"),
    "Lucky Investment Managers": ("LUCKY INVESTMENT MANAGERS PRIVATE LIMITED",),
    "R.B.A. Finance & Investment": ("R.B.A. FINANCE ## INVESTMENT CO.",
                                    "R.B.A.FINANCE & INVT. CO"),
    "Suresh Kumar Agarwal": ("Suresh Kumar Agarwal",),
    "Goldman Sachs": ("GOLDMAN SACHS (SINGAPORE) PTE",
                      "GOLDMAN SACHS (SINGAPORE) PTE.- ODI",
                      "GOLDMAN SACHS COLLECTIVE TRUST - EMERGING MARKETS EQUITY EX CHINA FUND",
                      "GOLDMAN SACHS COLLECTIVE TRUST - EMERGING MARKETS EQUITY EX. CHINA FUND",
                      "GOLDMAN SACHS FDS GOLDMAN SACHS INDIA EQ PORTFOLIO",
                      "GOLDMAN SACHS FUNDS  GOLDMAN SACHS INDIA EQUITY PORTFOLIO",
                      "GOLDMAN SACHS FUNDS - GOLDMAN SACHS INDIA EQUITY PORTFOLIO",
                      "GOLDMAN SACHS FUNDS GOLDMAN SACHS INDIA EQUITY PORTFOLIO",
                      "GOLDMAN SACHS FUNDS-GOLDMAN SACHS ASIA EQUITY PORTFOLIO",
                      "GOLDMAN SACHS INDIA LIMITED",
                      "GOLDMAN SACHS INVESTMENT (MAURITIUS) I LTD",
                      "GOLDMAN SACHS INVESTMENTS (MAURITIUS) I LIMITED",
                      "GOLDMAN SACHS INVESTMENTS HOLDINGS ASIA LIMITED",
                      "GOLDMAN SACHS INVESTMENTS MAURITIUS  I LIMITED",
                      "GOLDMAN SACHS INVESTMENTS MAURITIUS  I LTD",
                      "GOLDMAN SACHS INVESTMENTS MAURITIUS I LIMITED",
                      "GOLDMAN SACHS TRUST II - GOLDMAN SACHS GQG PARTNERS INTERNATIONAL OPPORTUNITIES FUND",
                      "GOLDMANSACHS FUNDS GOLDMANSACHS INDIA EQUITY PORTFOLIO"),
    "India Equity Fund 1": ("INDIA EQUITY FUND 1",),
    "Madhuri Madhusudan Kela": ("MADHURI MADHUSUDAN KELA",),
    "Madhusudan Murlidhar Kela": ("Madhusudan Murlidhar Kela",),
    "Cohesion MK Best Ideas": ("COHESION MK BEST IDEAS SUB-TRUST",),
    "Founders Collective Fund": ("FOUNDERS COLLECTIVE FUND",),
    "Singularity": ("SINGULARITY EQUITY FUND I", "SINGULARITY LARGE VALUE FUND I",
                    "SINGULARITY LARGE VALUE FUND II",
                    "SINGULARITY LARGE VALUE FUND III"),
    "Chartered Finance & Leasing": ("Chartered Finance & Leasing Limited",
                                    "CHARTERED FINANCE & LEASI NG LIMITED"),
    "Mona Laroia": ("LAROIA MONA", "MONA LAROIA"),
    "Bijal Pritesh Vora": ("BIJAL PRITESH VORA",),
    "Malabar India Fund": ("MALABAR INDIA FUND LIMITED",),
    "Massachusetts Institute of Technology": ("MASSACHUSETTS INSTITUTE OF TECHNOLOGY",),
    "Manish Grover": ("MANISH GROVER",),
    "Rohan Gupta": ("ROHAN GUPTA",),
    "Nalanda": ("NALANDA INDIA EQUITY FUND LIMITED", "NALANDA INDIA FUND LIMITED"),
    "NAV Capital": ("NAV CAPITAL VCC - NAV CAPITAL EMERGING STAR FUND",),
    "Rajasthan Global Securities": ("RAJASTHAN GLOBAL SECURITIES PVT.LTD",
                                    "RAJASTHAN GLOBAL SECURITIES PRIVATE LIMITED"),
    "Finavenue": ("FINAVENUE CAPITAL TRUST-FINAVENUE GROWTH FUND",
                  "Finavenue Capital Trust - Finavenue Growth Fund",
                  "Finavenue Capital trust Finavenue Growth Fund",
                  "Finavenue Capital Trust - Finavenue Strategic Fund"),
    "Saint Capital Fund": ("SAINT CAPITAL FUND",),
    "Meru Investment Fund": ("MERU INVESTMENT FUND PCC-CELL 1",
                             "MERU INVESTMENT FUND PCC - CELL 1",
                             "MERU INVESTMENT FUND"),
    "Vikasa": ("VIKASA INDIA EIF I FUND-INCUBE GLOBAL OPPORTUNITIES",
               "VIKASA INDIA EIF I FUND - INCUBE GLOBAL OPPORTUNITIES",
               "Vikasa India EIF I Fund- Share ClassP",
               "Vikasa India EIF I Fund - Share Class P",
               "Vikasa Global Fund PCC - Eubilia Capital Partners Fund - I",
               "VIKASA INDIA EIF I FUND - pte OPPORTUNITIES",
               "VIKASA CAPITAL INC"),
    "LRSD Securities": ("LRSD SECURITIES PRIVATE LIMITED", "LRSD SECURITIES PVT.LTD",
                        "LRSD SECURITIES PVT LTD"),
    "Tiger Strategies Fund": ("TIGER STRATEGIES FUND - 1", "Tiger Strategies Fund-I",
                              "Tiger Strategies Fund - I"),
    "Evergrow Capital Opportunities": ("EVERGROW CAPITAL OPPORTUNITIES FUND",),
    "SageOne": ("SAGEONE-FLAGSHIP GROWTH OE FUND", "Sageone - Flagship Growth OE Fund",
                "SageOne India Opportunity Trust"),
    "Mint Focused Growth Fund": ("Mint Focused Growth Fund PCC- CELL 1",
                                 "Mint Focused Growth Fund-PCC Cell 1",
                                 "Mint Focused Growth Fund PCC- Cell",
                                 "MINT FOCUSED GROWTH FUND"),
    "Religo Commodities Ventures": (
        "Religo Commodities Venture Trust-Religo Commodities Ventures Fund",
        "RELIGO COMMODITIES VENTURES TRUST - RELIGO COMMODITIES VENTURES FUND",
        "Religo Commeadition Ventures Trust Religo Commodities Ventures Fund"),
    "Hem Growth Opportunities Fund": ("Hem Growth Opportunities Fund",),
    "RGSL Investment Fund": ("RGSL INVESTMENT FUND - RGSL INVESTMENT LVF 1",),
    "Ritu Bapna": ("RITU BAPNA",),
    "Sandeep Singh": ("SANDEEP SINGH",),
    "Mukul Mahavir Agrawal": ("Mukul Mahavir Agrawal",),
    "Asha Mukul Agrawal": ("Asha Mukul Agrawal",),
    "Sanshi Fund-I": ("SANSHI FUND-I",),
    "Param Capital": ("PARAM CAPITAL",),
    "Shalu Aggarwal": ("SHALU  AGGARWAL",),
    "Vanaja Sundar Iyer": ("VANAJA SUNDAR IYER",),
    "Venkata Nagaraju Padala": ("VENKATA NAGARAJU PADALA",),
    "Vinod Kumar": ("VINOD  KUMAR",),
    "ValueQuest": ("Valuequest S C A L E Fund", "VQ FASTERCAP FUND",
                   "VALUEQUEST INVESTMENT ADVISORS PVT LTD"),
}


def _track_norm(s: str) -> str:
    """Upper-case, alphanumerics and single spaces only — the same reading
    ipo_anchor_tracker applies to its watchlist, so both agree on a name."""
    s = unicodedata.normalize("NFKD", str(s or "")).upper()
    return re.sub(r"\s+", " ", re.sub(r"[^A-Z0-9]+", " ", s)).strip()


# Words that name a vehicle or a legal form rather than a house. An overlap
# built only from these identifies nobody: "STRATEGIC" sits inside Finavenue's
# strategic fund and inside a hundred others, and "INDIA EQUITY FUND 1" is a
# substring of "3P INDIA EQUITY FUND 1M", which is a different manager. A match
# must therefore carry at least one word from outside this set.
_TRACK_GENERIC = {
    "FUND", "FUNDS", "CAPITAL", "TRUST", "TRUSTEE", "INDIA", "INDIAN", "EQUITY",
    "EQUITIES", "GROWTH", "OPPORTUNITY", "OPPORTUNITIES", "STRATEGIC", "STRATEGY",
    "EMERGING", "STAR", "LIMITED", "LTD", "PVT", "PRIVATE", "PCC", "CELL", "VCC",
    "SECURITIES", "INVESTMENT", "INVESTMENTS", "PORTFOLIO", "ASSET", "ASSETS",
    "MANAGERS", "MANAGEMENT", "ADVISORS", "ADVISERS", "COMPANY", "CO", "AND",
    "OF", "THE", "LLP", "INC", "FINANCE", "FINANCIAL", "MULTI", "LARGE", "SMALL",
    "MID", "CAP", "VALUE", "FOCUSED", "FLAGSHIP", "GLOBAL", "SUB", "ODI", "PTE",
    "PLAN", "SCHEME", "SERIES", "CLASS", "I", "II", "III", "IV", "V", "1", "2",
    "3", "4", "5", "A", "B", "C", "NEW", "GENERAL", "SPECIAL", "SELECT", "PLUS",
}


def _track_tokens(s: str) -> frozenset:
    """The words in a name that actually identify someone.

    Substring tests cannot do this job: "BENGAL FINANCE & INVESTMENT LIMITED"
    is not a substring of "...INVESTMENT PRIVATE LIMITED" because one word is
    inserted in the middle, so containment silently loses a real investor.
    Comparing sets of words survives insertions, reordering and punctuation,
    and comparing WORDS rather than characters also stops "PARAM" reaching
    inside "PARAMOUNT"."""
    return frozenset(w for w in _track_norm(s).split()
                     if w not in _TRACK_GENERIC and len(w) > 1)


@lru_cache(maxsize=1)
def _tracked_index() -> tuple:
    """(identifying words, canonical name) for every watchlist spelling."""
    out = {(_track_tokens(v), canon)
           for canon, variants in TRACKED_INVESTORS.items() for v in variants}
    return tuple((t, c) for t, c in out if t)


def tracked_match(name: str) -> "str | None":
    """The tracked investor an anchor name denotes, or None.

    Two ways to match, and they are deliberately not symmetric. A watchlist
    identity wholly present in the anchor name is a match at any size, so the
    single word BENGAL still finds its company. The reverse — a fragment of a
    watchlist name — needs two words, because one word plucked out of a long
    entry identifies nobody: a bare MAURITIUS or INTERNATIONAL would otherwise
    land on Goldman's Mauritius and GQG vehicles.

    The widest overlap wins, which is what decides a garbled row naming two
    investors at once."""
    t = _track_tokens(name)
    if not t:
        return None
    best_n, best = 0, None
    for toks, canon in _tracked_index():
        if toks <= t:
            n = len(toks)
        elif t <= toks and len(t) >= 2:
            n = len(t)
        else:
            continue
        if n > best_n:
            best_n, best = n, canon
    return best


def ipl_anchor_scan(ipos: "list[IPO]") -> "dict[str, dict]":
    """Tracked investor -> the issues it anchored, over EVERY IPO in the window.

    Our own anchor filings cannot answer this. They are downloaded only for
    issues that cleared the gain threshold, so they cover 88% of winners and
    under 4% of everything else; counting participations from them would report
    a fraction of a tracked investor's real activity and would do it with a bias
    towards the deals that worked. IPOPlatform publishes the anchor book of
    every issue, so reading all of them gives a count that does not know or care
    what the price did next.

    Pages for issues that have already listed are cached permanently — an
    allocation is final once the issue lists — so a second run re-reads them
    from disk and only genuinely new issues are fetched."""
    today = date.today()
    hits: dict[str, dict] = {}
    todo = [i for i in ipos if i.listing_date]
    fetched = 0
    print(f"  [ipoplatform] reading anchor books for {len(todo)} issue(s) "
          "(already-listed pages are served from cache) ...")
    for n, ipo in enumerate(todo, 1):
        info = ipl_resolve(ipo.symbol, ipo.company)
        if not info:
            continue
        settled = ipo.listing_date < today
        rows = _ipl_anchors_for(info,
                               ttl_hours=IPL_TTL_SETTLED if settled else IPL_TTL_HOURS)
        if rows:
            fetched += 1
        for anchor, _ in rows:
            canon = tracked_match(anchor.name)
            if not canon:
                continue
            d = hits.setdefault(canon, {"issues": {}, "names": set()})
            d["names"].add(str(anchor.name).strip())
            d["issues"][ipo.symbol] = ipo
        if n % 50 == 0:
            print(f"            {n}/{len(todo)} ...")
    print(f"  [ipoplatform] anchor books read for {fetched} issue(s); "
          f"{len(hits)} tracked investor(s) found.")
    return hits


def build_tracked_investors(hits: "dict[str, dict]",
                            win_syms: "set[str]") -> pd.DataFrame:
    """One row per tracked investor: how many issues it anchored in the window.

    Every name on the watchlist gets a row even at zero, because "did not
    participate" is the answer to the question being asked and a missing row
    reads as an oversight."""
    rows = []
    for canon in TRACKED_INVESTORS:
        d = hits.get(canon) or {"issues": {}, "names": set()}
        issues = d["issues"]
        qualified = sorted(s for s in issues if s in win_syms)
        won = len(qualified)
        rows.append({
            "Investor": canon,
            "Total IPOs": len(issues),
            "Qualified IPOs": won,
            "Hit Rate %": (round(100.0 * won / len(issues), 1) if issues else None),
            "Symbols": ", ".join(sorted(issues)),
            "Qualified Symbols": ", ".join(qualified),
        })
    df = pd.DataFrame(rows)
    for col in ("Total IPOs", "Qualified IPOs"):
        df[col] = df[col].astype("Int64")
    return df.sort_values(["Total IPOs", "Investor"],
                          ascending=[False, True]).reset_index(drop=True)


# Boilerplate words in a filing's filename that are not part of the issuer's
# name, stripped when a filing has to be keyed by its filename.
_FILING_NOISE = {"anchor", "anchors", "investor", "investors", "allocation",
                 "allotment", "intimation", "letter", "list", "final", "ipo",
                 "nse", "bse", "ltd", "limited", "pdf", "scan", "signed"}


def _filing_key(path: Path) -> str:
    """Column name for a filing that matches no current winner: the filename
    with the boilerplate stripped, upper-cased ('anawil-wire-anchor-investor'
    -> 'ANAWIL WIRE'). Rename the column to the NSE symbol once it lists."""
    words = [w for w in re.split(r"[^A-Za-z0-9]+", path.stem) if w]
    kept = [w for w in words if w.lower() not in _FILING_NOISE]
    return (" ".join(kept).upper()[:60] or path.stem.upper()[:60]).strip()


def collect_anchor_investors(anchor_dir: Path, winners: list[IPO],
                             use_ocr: bool = True, use_ipoplatform: bool = True,
                             extra_keys: "tuple[str, ...]" = (),
                             universe: "list[IPO] | None" = None):
    """Match every anchor filing in `anchor_dir` (PDF or image) to one of the
    winning IPOs and extract its full anchor table. Matching order: filename
    contains the NSE symbol -> filename contains the company name -> first-page
    text contains the company name. Returns (symbol -> [Anchor],
    unmatched_file_names, symbols_without_a_filing).

    `extra_keys` are anchor labels the workbook already holds that no filing in
    this run produced. They carry no names of their own here; they are passed on
    so IPOPlatform is asked about them too, which is the only way a row that was
    saved as a bare name ever acquires its figures.

    Where IPOPlatform publishes the book (SME issues), its list REPLACES the
    OCR reading rather than being merged into it. Merging sounds safer and is
    not: OCR failures are corrupt names, not missing ones, so a union would
    keep "GROWTH OE FT]ND" alongside the real fund and inflate the frequency
    table with entities that do not exist. A source with canonical investor ids
    is strictly better than a photocopy, so it wins outright."""
    mapping: dict[str, list[Anchor]] = {}
    unmatched: list[str] = []
    # Keys whose filing was read to a table that reconciles exactly. Collected
    # so the caller can let those readings REPLACE what the workbook holds.
    verified: set[str] = set()
    if not anchor_dir.exists():
        return mapping, unmatched, [w.symbol for w in winners], verified

    cores = [(w, _company_core(w.company), w.symbol.lower()) for w in winners]

    def _register(sym: str, anchors: list[Anchor]) -> None:
        cur = mapping.setdefault(sym, [])
        have = {_norm_alnum(a.name) for a in cur}
        for a in anchors:
            k = _norm_alnum(a.name)
            if k and k not in have:
                cur.append(a)
                have.add(k)

    files = sorted(p for p in anchor_dir.iterdir()
                   if p.is_file()
                   and p.suffix.lower() in ((".pdf",) + IMAGE_SUFFIXES))
    for pdf in files:
        stem = _norm_alnum(pdf.stem)
        match: IPO | None = None
        # 1) filename contains the symbol
        for w, core, sym in cores:
            if sym and sym in stem:
                match = w
                break
        # 2) filename contains the company core (prefer the longest match)
        if match is None:
            cands = [(w, core) for w, core, _ in cores if core and core in stem]
            if cands:
                match = max(cands, key=lambda t: len(t[1]))[0]
        # 3) content match against the first page (OCR it if it's a scan)
        if match is None:
            head = ""
            if pdf.suffix.lower() == ".pdf":
                try:
                    import pdfplumber  # type: ignore
                    with pdfplumber.open(str(pdf)) as pf:
                        if pf.pages:
                            head = _norm_alnum(
                                (pf.pages[0].extract_text() or "")[:2000])
                except Exception:
                    head = ""
            if not head and use_ocr and _ocr_available():
                head = _norm_alnum(
                    (_ocr_pdf(pdf, 200, "6") if pdf.suffix.lower() == ".pdf"
                     else _tesseract(pdf, "6"))[:4000])
            cands = [(w, core) for w, core, _ in cores if core and core in head]
            if cands:
                match = max(cands, key=lambda t: len(t[1]))[0]

        if match is None:
            # No winner owns this filing — normally an IPO that has not listed
            # yet (so it cannot have a gain), or one that missed the threshold.
            # Parse it ANYWAY and file it under its filename-derived key.
            # Discarding it would leave a silently incomplete anchor book, and
            # that is the one outcome this script exists to prevent.
            anchors = extract_anchors_from_file(pdf, use_ocr)
            if anchors:
                key = _filing_key(pdf)
                _register(key, anchors)
                if _RECONCILED_FILINGS.get(pdf.name):
                    verified.add(key)
                    _frozen_save(pdf, anchors)
                print(f"  [anchor] {pdf.name}: no matching gainer — kept under "
                      f"'{key}' ({len(anchors)} names). Rename the row label to "
                      "the NSE symbol once it lists.")
            else:
                unmatched.append(pdf.name)
            continue
        anchors = extract_anchors_from_file(pdf, use_ocr)
        _register(match.symbol, anchors)
        if _RECONCILED_FILINGS.get(pdf.name):
            verified.add(match.symbol)
            _frozen_save(pdf, anchors)

    if use_ipoplatform:
        _apply_ipoplatform(mapping, winners, tuple(extra_keys), universe)

    _repair_ocr_spellings(mapping)
    _repair_multichar_confusables(mapping)
    _dictionary_fuzzy_repair(mapping)
    _symspell_repair(mapping)

    missing = [w.symbol for w in winners if not mapping.get(w.symbol)]
    return mapping, unmatched, missing, verified


# Glyph pairs tesseract genuinely confuses, because they are the pairs that
# LOOK ALIKE once a filing has been faxed, photocopied and rendered to a
# bitmap. This list is the whole basis of _repair_ocr_spellings and it is
# deliberately short: it exists to separate a MISREADING from a DIFFERENT
# INVESTOR, and those two cases are otherwise indistinguishable.
#
# "HDEC LIFE INSURANCE" and "HDFC LIFE INSURANCE" differ by one character, and
# so do "ITI MULTI CAP FUND" and "UTI MULTI CAP FUND". Every similarity score
# ever tried rates them the same, so any rule built on edit distance alone
# either leaves HDEC in the register or merges two real asset managers. The
# difference is not statistical, it is physical: E and F differ by one stroke
# and tesseract confuses them constantly, while I and U share no strokes at all
# and it does not. So E/F is repairable and I/U is not, and ITI survives.
_OCR_CONFUSABLE = [set(p) for p in (
    "EF", "CG", "CO", "OQ", "OD", "GQ", "IL", "IT", "JI", "SB", "PR",
    "VY", "UV", "MN", "KX", "ZS", "HN", "DO", "BR", "EB", "FP", "TY",
)]


def _confusable_edit(a: str, b: str) -> bool:
    """True if `a` could be `b` mis-read once, on strings of equal length.

    Exactly one position may differ, and that position must be a documented
    glyph confusion. Digits are never repairable: AXIS MAX LIFE's ULIF00225 and
    ULIF00625 are two different unit-linked funds, not one fund read twice, and
    the same goes for 3P INDIA EQUITY 1M against 2M."""
    diff = [(x, y) for x, y in zip(a, b) if x != y]
    if len(diff) != 1:
        return False
    x, y = diff[0]
    if x.isdigit() or y.isdigit():
        return False
    return {x, y} in _OCR_CONFUSABLE


def _indel_edit(a: str, b: str) -> bool:
    """True if `a` is `b` with one letter dropped or one letter added.

    Covers the other half of what OCR does to a name — swallowing a repeated
    letter or doubling one: ICIC for ICICI, AGGRESIVE for AGGRESSIVE, MULT for
    MULTI, EXX for EX. Both names must be long enough that one letter cannot be
    the whole difference between two real institutions, and neither edit may
    touch a digit."""
    if abs(len(a) - len(b)) != 1 or min(len(a), len(b)) < 8:
        return False
    short, long = (a, b) if len(a) < len(b) else (b, a)
    i = 0
    while i < len(short) and short[i] == long[i]:
        i += 1
    if i < len(long) and long[i].isdigit():
        return False
    return short[i:] == long[i + 1:]


def _repair_key(name: str) -> str:
    """Squashed identity of `name` with any leading table serial removed.

    OCR routinely welds the row number onto the name — "19 HDEC Life Insurance
    company Limited", "2 ICIC] PRUDENTIAL LIFE INSURANCE", "10 NUVAMA MULT]
    ASSET" — and squash_investor_name keeps it, so those rows never line up
    with the clean spelling of the same investor.

    Stripping a leading number is NOT safe on its own, because 360 ONE FLEXICAP
    FUND and 3P INDIA EQUITY FUND are real names that begin with digits. It is
    made safe by what the caller does with the result: a stripped key is only
    ever acted on when it lands on a spelling that ALREADY EXISTS in the
    filings. "ONEFLEXICAPFUND" matches nothing, so 360 ONE is left alone."""
    return squash_investor_name(re.sub(r"^\s*[A-Za-z]?\d{1,3}[\s.)\-\]]+", "",
                                       str(name or "")))


def _repair_ocr_spellings(mapping: "dict[str, list[Anchor]]") -> None:
    """Rewrite one-off mis-spellings onto the spelling the filings agree on.

    NOTHING IS EVER DROPPED. Every row survives; only the name cell changes, so
    the count of investors per issue cannot fall because of this pass. That is
    the point: the register was carrying HDFC LIFE INSURANCE six times and HDEC
    LIFE INSURANCE once, CITIGROUP GLOBAL MARKETS six times beside IL CITIGROUP
    GLOBAL and IZ CITIGROUP GLOBAL — one investor showing up as three.

    A spelling is only repaired when it is a one-off, when the two differ by a
    single confusable glyph or one letter, when the spelling it would move to
    is the tidier of the two, and when exactly one such target exists. An
    ambiguous case is left alone rather than guessed."""
    seen: dict[str, dict[str, int]] = {}
    for anchors in mapping.values():
        for a in anchors:
            q = _repair_key(a.name)
            if len(q) >= 8:
                nm = str(a.name).strip()
                seen.setdefault(q, {})
                seen[q][nm] = seen[q].get(nm, 0) + 1
    total = {q: sum(n.values()) for q, n in seen.items()}

    def _best(q: str) -> str:
        return max(seen[q].items(),
                   key=lambda kv: (kv[1], -_label_penalty(kv[0]), len(kv[0])))[0]

    def _tidier(src: str, tgt: str) -> bool:
        # The move has to be an improvement, decided the same way the frequency
        # table picks a group's label: corroboration first, then how mangled the
        # spelling is. Without this the repair could run downhill and rename the
        # good spelling to the bad one.
        return ((total[tgt], -_label_penalty(_best(tgt)))
                > (total[src], -_label_penalty(_best(src))))

    fixes: dict[str, str] = {}
    for q in total:
        if total[q] != 1:
            continue
        hits = [t for t in total
                if t != q and _tidier(q, t)
                and ((len(t) == len(q) and _confusable_edit(q, t))
                     or _indel_edit(q, t))]
        if len(hits) == 1:
            fixes[q] = _best(hits[0])
    if not fixes:
        return
    changed = 0
    for anchors in mapping.values():
        for a in anchors:
            tgt = fixes.get(_repair_key(a.name))
            if tgt and str(a.name).strip() != tgt:
                a.name = tgt
                changed += 1
    print(f"  [anchor] repaired {changed} mis-spelled name(s) onto the spelling "
          f"other filings agree on ({len(fixes)} distinct); no row was dropped.")


def _token_fuzzy_score(ocr_tokens: "tuple[str, ...]",
                       dict_tokens: "tuple[str, ...]") -> float:
    """Score how well OCR tokens match a dictionary entry's tokens.

    Each OCR token is matched to the best dictionary token by: exact match,
    confusable edit, indel edit, or SequenceMatcher ratio >= 0.80.  The score
    is matched_count / max(len(ocr), len(dict)) so a short garbage string
    cannot score high against a long dictionary name."""
    if not ocr_tokens or not dict_tokens:
        return 0.0
    used: set[int] = set()
    matched = 0
    for ot in ocr_tokens:
        best_j = -1
        best_sim = 0.0
        for j, dt in enumerate(dict_tokens):
            if j in used:
                continue
            if ot == dt:
                best_j, best_sim = j, 1.0
                break
            if len(ot) == len(dt) and _confusable_edit(ot, dt):
                if 0.95 > best_sim:
                    best_j, best_sim = j, 0.95
                continue
            if _indel_edit(ot, dt):
                if 0.90 > best_sim:
                    best_j, best_sim = j, 0.90
                continue
            ratio = difflib.SequenceMatcher(None, ot, dt).ratio()
            if ratio >= 0.80 and ratio > best_sim:
                best_j, best_sim = j, ratio
        if best_j >= 0 and best_sim >= 0.80:
            used.add(best_j)
            matched += 1
    denom = max(len(ocr_tokens), len(dict_tokens))
    return matched / denom if denom else 0.0


_MULTICHAR_CONFUSABLE = [
    ("rn", "m"), ("m", "rn"),
    ("cl", "d"), ("d", "cl"),
    ("vv", "w"), ("w", "vv"),
    ("fl", "fi"), ("fi", "fl"),
    ("li", "h"), ("h", "li"),
    ("ll", "U"), ("U", "ll"),
    ("lI", "U"),
]


def _repair_multichar_confusables(mapping: "dict[str, list[Anchor]]") -> None:
    """Fix multi-character OCR confusions like rn→m, cl→d, vv→w.

    For each name, tries every substitution and checks whether the result
    matches a spelling already seen across filings. Only applies when exactly
    one substitution produces a known spelling."""
    known: dict[str, str] = {}
    for anchors in mapping.values():
        for a in anchors:
            k = squash_investor_name(a.name)
            if k and k not in known:
                known[k] = str(a.name).strip()
    changed = 0
    for anchors in mapping.values():
        for a in anchors:
            orig_k = squash_investor_name(a.name)
            if not orig_k:
                continue
            hits: list[str] = []
            for bad, good in _MULTICHAR_CONFUSABLE:
                if bad not in str(a.name):
                    continue
                fixed = str(a.name).replace(bad, good, 1)
                fk = squash_investor_name(fixed)
                if fk != orig_k and fk in known and known[fk] != str(a.name).strip():
                    hits.append(known[fk])
            if len(hits) == 1:
                a.name = hits[0]
                changed += 1
    if changed:
        print(f"  [anchor] repaired {changed} multi-char OCR confusion(s) "
              f"(rn/m, cl/d, vv/w, fi/fl).")


def _dictionary_fuzzy_repair(mapping: "dict[str, list[Anchor]]") -> None:
    """Rewrite OCR-damaged names onto known-investor canonical spellings.

    NOTHING IS EVER DROPPED. Like _repair_ocr_spellings, only the name cell
    changes; the count of investors per issue cannot fall."""
    dict_entries: list[tuple[str, "tuple[str, ...]"]] = []
    seen_keys: set["tuple[str, ...]"] = set()
    try:
        for _pid, name, _count in ipl_parents():
            toks = canonical_investor_tokens(name)
            if len(toks) >= 2 and toks not in seen_keys:
                dict_entries.append((name, toks))
                seen_keys.add(toks)
    except Exception:
        pass
    for anchors in mapping.values():
        if len(anchors) < 5:
            continue
        for a in anchors:
            if not _plausible_investor_name(a.name):
                continue
            toks = canonical_investor_tokens(a.name)
            if len(toks) >= 2 and toks not in seen_keys:
                dict_entries.append((str(a.name).strip(), toks))
                seen_keys.add(toks)
    if not dict_entries:
        return
    changed = 0
    for anchors in mapping.values():
        for a in anchors:
            ocr_toks = canonical_investor_tokens(a.name)
            if len(ocr_toks) < 2:
                continue
            if ocr_toks in seen_keys:
                continue
            if _plausible_investor_name(a.name) and _label_penalty(a.name) == 0:
                continue
            scores: list[tuple[float, str]] = []
            for canon_name, dict_toks in dict_entries:
                sc = _token_fuzzy_score(ocr_toks, dict_toks)
                if sc >= 0.70:
                    scores.append((sc, canon_name))
            if not scores:
                continue
            scores.sort(reverse=True)
            if len(scores) == 1 or scores[0][0] - scores[1][0] >= 0.10:
                new_name = scores[0][1]
                if str(a.name).strip() != new_name:
                    a.name = new_name
                    changed += 1
    if changed:
        print(f"  [anchor] dictionary-matched {changed} name(s) to known "
              f"investors; no row was dropped.")


def _symspell_delete_variants(word: str, max_dist: int = 2) -> set[str]:
    """Generate all delete-only variants within edit distance max_dist."""
    variants: set[str] = {word}
    queue = [word]
    for _d in range(max_dist):
        nxt: list[str] = []
        for w in queue:
            for i in range(len(w)):
                v = w[:i] + w[i + 1:]
                if v not in variants:
                    variants.add(v)
                    nxt.append(v)
        queue = nxt
    return variants


def _symspell_repair(mapping: "dict[str, list[Anchor]]") -> None:
    """Fast edit-distance correction using the Symmetric Delete algorithm.

    Builds a delete-neighbourhood index over all known investor names (from
    ipl_parents and from the mapping itself), then looks up each OCR'd name.
    Only names within edit distance 2 of exactly ONE dictionary entry are
    corrected — ambiguous matches are left alone."""
    dict_names: dict[str, str] = {}
    try:
        for _pid, name, _count in ipl_parents():
            key = _norm_alnum(name)
            if key and key not in dict_names:
                dict_names[key] = name.strip()
    except Exception:
        pass
    for anchors in mapping.values():
        if len(anchors) < 5:
            continue
        for a in anchors:
            if not _plausible_investor_name(a.name):
                continue
            key = _norm_alnum(str(a.name))
            if key and key not in dict_names:
                dict_names[key] = str(a.name).strip()
    if not dict_names:
        return
    delete_index: dict[str, list[str]] = {}
    for key in dict_names:
        for v in _symspell_delete_variants(key, 2):
            delete_index.setdefault(v, []).append(key)
    changed = 0
    for anchors in mapping.values():
        for a in anchors:
            ocr_key = _norm_alnum(str(a.name))
            if not ocr_key or ocr_key in dict_names:
                continue
            candidates: set[str] = set()
            for v in _symspell_delete_variants(ocr_key, 2):
                for dk in delete_index.get(v, ()):
                    if dk == ocr_key:
                        continue
                    d = abs(len(dk) - len(ocr_key))
                    if d <= 2:
                        candidates.add(dk)
            if len(candidates) == 1:
                winner = candidates.pop()
                new_name = dict_names[winner]
                if str(a.name).strip() != new_name:
                    a.name = new_name
                    changed += 1
    if changed:
        print(f"  [anchor] symspell-corrected {changed} name(s) to known "
              f"investors; no row was dropped.")


def _apply_ipoplatform(mapping: dict, winners: list[IPO],
                       extra_keys: "tuple[str, ...]" = (),
                       universe: "list[IPO] | None" = None) -> None:
    """Overwrite `mapping[key]` with IPOPlatform's book wherever it has one.

    Reports what it did per symbol against the OCR reading it replaced, because
    the interesting cases are the disagreements: where IPOPlatform returns MORE
    names the PDF reader lost investors, and where it returns FEWER the PDF
    reader invented them out of page furniture and address lines.

    Runs over every anchor key, not only the winners. Gating on the winners list
    meant an issue that had not listed yet, or had not cleared the gain
    threshold, was never asked about — MILLWORKS sat in the workbook as nine
    names with no shares, no price and no amount while IPOPlatform had all three
    for every one of them. Nothing about an issue's listing performance changes
    who its anchors were."""
    idx = ipl_index()
    if not idx:
        print("  [ipoplatform] index unavailable — keeping the OCR readings.",
              file=sys.stderr)
        return
    # No register check any more: full names now come from each row's own link,
    # so the register being down no longer costs us the untruncated spelling.
    # (key to write under, symbol to look up, company name to fall back on)
    targets = [(w.symbol, w.symbol, w.company) for w in winners]
    # Sheet 2 labels a winner's rows with its COMPANY NAME, so those labels come
    # back as extra keys on the next run. Enriching one of them would file the
    # same anchor book a second time under a key of its own, and the issuer then
    # prints as one block with every name twice — 233 such rows across 33
    # issuers. An extra key is only a new issue if no winner already covers it.
    # Drawn from every IPO in the window, not just the winners: an issue that
    # stops qualifying keeps its rows on the sheet, and if its two spellings
    # stop being recognised as one issue they split into two blocks.
    covered = universe if universe is not None else winners
    known = {i.symbol for i in covered} | {i.company for i in covered}
    for key in list(mapping) + list(extra_keys):
        if key not in known:
            known.add(key)
            targets.append((key, key, key))
    took = gained = lost = 0
    no_book: list[str] = []
    degraded: list[str] = []
    for key, sym, company in targets:
        info = ipl_resolve(sym, company)
        if not info:
            continue
        rows = _ipl_anchors_for(info)
        if not rows:
            # Mainboard always lands here; so does an SME issue too recent for
            # them to have keyed in. Either way the PDF reading stands.
            no_book.append(f"{key}/{info['seg'][:3]}")
            continue
        before = len(mapping.get(key) or [])
        stated = _ipl_num(info.get("anchor_shares"))
        got = sum(a.shares or 0 for a, _ in rows)
        reconciles = bool(stated) and abs(got - stated) <= 1
        # A half-rendered page must never silently delete investors the PDF
        # reader did find. Returning somewhat FEWER names is normal and good —
        # that is OCR furniture (GSTIN lines, pincodes, address fragments)
        # being dropped — but a collapse to under half the OCR count, without
        # the shares reconciling to the declared allocation to prove the short
        # list is complete, is a broken response rather than a cleaner one.
        if before and len(rows) * 2 < before and not reconciles:
            print(f"  [ipoplatform] {key}: returned only {len(rows)} name(s) "
                  f"against {before} from the PDF and does not reconcile — "
                  "treating as a degraded response and keeping the OCR reading.",
                  file=sys.stderr)
            degraded.append(key)
            continue
        mapping[key] = [a for a, _ in rows]
        took += 1
        gained += max(0, len(rows) - before)
        lost += max(0, before - len(rows))
        via = "" if info is idx.get(sym.strip().upper()) else \
            f" [matched on name -> {info.get('company')}]"
        if stated and not reconciles:
            print(f"  [ipoplatform] {key}: {len(rows)} names "
                  f"(OCR had {before}); shares {got:,} vs declared "
                  f"{stated:,.0f} (off by {got - stated:+,.0f}).{via}")
        else:
            print(f"  [ipoplatform] {key}: {len(rows)} names "
                  f"(OCR had {before})"
                  f"{' — reconciles exactly' if stated else ''}.{via}")
    print(f"  [ipoplatform] used for {took} issue(s): +{gained} name(s) the PDF "
          f"reader had missed, -{lost} it had invented. No published book for "
          f"{len(no_book)}: {', '.join(no_book) or 'none'}")
    if degraded:
        print(f"  [ipoplatform] {len(degraded)} issue(s) fell back to OCR: "
              f"{', '.join(degraded)}", file=sys.stderr)


# ──────────────── investor frequency table (Sheet 3) ───────────────────────
# Anchor books repeat: the same funds seed issue after issue. Counting that
# reliably means deciding when two name cells are the same investor, and the
# filings make that hard — the same fund arrives as "Saint Capital Fund",
# "SAINT CAPITAL FUND" and "Saint Capital Limited", while OCR adds leading row
# serials, truncations and rows where two names have been glued together.
#
# The default key ("smart") canonicalises each name to its significant tokens
# and then merges a name into a fuller one. The legacy key — first N words of
# the raw string — is still available via --name-key words; it both over-merged
# (every "HDFC MUTUAL FUND - HDFC <scheme>" collapsed onto "HDFC MUTUAL") and
# under-merged (case and OCR variants split apart).
def investor_key(name: str, words: int = 2) -> str:
    """First `words` words of a name, upper-cased and stripped of punctuation."""
    w = re.sub(r"[^A-Z0-9 ]+", " ", str(name or "").upper()).split()
    return " ".join(w[:words])


# Tokens that describe a legal wrapper or a fund structure rather than the
# investor itself. Removing them is what collapses "... Fund", "... Limited"
# and "... Trust" onto one key. Deliberately conservative: words that carry
# identity (CAPITAL, SECURITIES, GROWTH, INVESTMENT) are NOT in here, because
# they are what separates one scheme of a house from another.
_NAME_NOISE = frozenset({
    "LTD", "LIMITED", "PVT", "PRIVATE", "LLP", "PLC", "INC", "CORP",
    "CORPORATION", "CO", "COMPANY", "COMPANIES",
    "THE", "AND", "OF", "FOR", "ON", "BEHALF", "ITS", "THROUGH", "VIA",
    "TRUST", "TRUSTEE", "TRUSTEES", "TRUSTEESHIP",
    "FUND", "FUNDS", "SCHEME", "SCHEMES", "MUTUAL", "MF",
    "PCC", "CELL", "VCC", "AIF", "IFSC", "CATEGORY", "CLASS", "SERIES", "SUB",
})
# How many extra tokens a fuller name may add and still be treated as the same
# investor. Loose enough for "Shine Star" -> "Shine Star Build Cap", tight
# enough that a name cell holding two glued investors is not a bridge between
# them.
_MAX_EXTRA_TOKENS = 3
# A figure that leaked into the name cell from the amount/price column beside
# it: a decimal, or a bare integer of three digits or more. Single digits are
# left alone because they belong to real names ("PCC-CELL 1", "FUND-I").
_LEAKED_FIGURE = re.compile(r"^\d[\d,]*\.\d+$|^\d{3,}$")
# Characters no filing actually uses; their presence means OCR invented them.
_CLEAN_CHARS = re.compile(r"[^A-Za-z0-9 &.,:\-()'/]")


def _label_penalty(name: str) -> int:
    """How mangled a spelling looks. Lower is a better display label.

    Scores the three ways OCR corrupts a name cell: it drags in the figure from
    the column beside it ("TIGER STRATEGIES FUND-I 600 4.18"), it invents glyphs
    ("Shine Star Build Cap Private Limited —«*"), and it re-reads a fragment it
    has already read ("EDELWEISS TRUSTEESHIP CO LTD AC - EDELWEISS MF AC-
    EDELWEISS RECENTLY LISTED IPO FUND")."""
    words = re.sub(r"[^A-Za-z0-9. ]+", " ", str(name)).split()
    figures = sum(2 for w in words if _LEAKED_FIGURE.match(w))
    glyphs = len(_CLEAN_CHARS.findall(str(name)))
    upper = [w.upper() for w in words]
    repeats = len(upper) - len(set(upper))
    return figures + glyphs + repeats


# How many identifying tokens a displayed investor name is shortened to. The
# filed names run to 80 characters and carry OCR debris, which makes the tables
# unreadable; three tokens is enough to tell one investor from another.
_LABEL_TOKENS = 3


def _short_labels(clean: "dict") -> "dict":
    """Shorten each group's chosen spelling to its first identifying tokens.

    Purely cosmetic — grouping and counts are already settled. A label is
    widened one token at a time whenever it would collide with another group's,
    because two groups sharing a label would silently overwrite each other in
    the profile dictionary. Anything that still collides keeps its full
    spelling, which is unique because a filed name belongs to exactly one
    group."""
    toks = {k: canonical_investor_tokens(v) for k, v in clean.items()}
    width = {k: _LABEL_TOKENS for k in clean}
    out = {k: (str(clean[k]).strip() if not t else " ".join(t[:_LABEL_TOKENS]))
           for k, t in toks.items()}
    for _ in range(max((len(t) for t in toks.values()), default=0)):
        out = {k: (str(clean[k]).strip() if not t else " ".join(t[:width[k]]))
               for k, t in toks.items()}
        seen: dict[str, list] = {}
        for k, lab in out.items():
            seen.setdefault(lab, []).append(k)
        clashing = [k for ks in seen.values() if len(ks) > 1 for k in ks]
        if not clashing:
            return out
        widened = False
        for k in clashing:
            if width[k] < len(toks[k]):
                width[k] += 1
                widened = True
        if not widened:
            break
    seen = {}
    for k, lab in out.items():
        seen.setdefault(lab, []).append(k)
    for ks in seen.values():
        if len(ks) > 1:
            for k in ks:
                out[k] = str(clean[k]).strip()
    return out


def _is_ordinal_token(t: str) -> bool:
    """Does this token number a series ("VI", "3") rather than name it?"""
    if t in _ROMAN_ORDINAL:
        return True
    return t.isdigit() and 0 < int(t) <= _MAX_ORDINAL


def canonical_investor_tokens(name: str) -> "tuple[str, ...]":
    """Reduce an investor name to its identifying tokens, in order.

    Upper-cases, strips accents and punctuation, drops the leading row serial
    OCR picks up from the filing's numbered table, removes structural words and
    de-duplicates. "FINAVENUE CAPITAL TRUST-FINAVENUE GROWTH FUND" and
    "6 Finavenue Capital Trust Finavenue Growth Fund" both reduce to
    (FINAVENUE, CAPITAL, GROWTH)."""
    s = unicodedata.normalize("NFKD", str(name or "")).encode("ascii", "ignore").decode()
    words = re.sub(r"[^A-Za-z0-9]+", " ", s).upper().split()
    while words and words[0].isdigit():
        words = words[1:]
    out, seen = [], set()
    for i, t in enumerate(words):
        if len(t) < 2 or t.isdigit() or t in seen:
            continue
        if t in _NAME_NOISE:
            # "SERIES" is structural only where it NUMBERS something ("SERIES
            # VI"), and the ordinal is preserved separately. Where a fund is
            # actually called that, it is as much part of the name as any other
            # word: dropping it made MEESHO's FIDELITY SERIES EMERGING MARKETS
            # FUND identical to its FIDELITY EMERGING MARKETS FUND, so two
            # anchors holding 2,408,535 and 828,225 shares merged into one row
            # and the smaller of them left the workbook.
            if not (t == "SERIES"
                    and not _is_ordinal_token(words[i + 1]
                                              if i + 1 < len(words) else "")):
                continue
        seen.add(t)
        out.append(t)
    return tuple(out)


# Legal suffixes, spelled without the spaces that OCR sometimes swallows.
# Longest first so "PRIVATELIMITED" is peeled whole rather than leaving "PRIVATE".
_GLUED_SUFFIX = ("PRIVATELIMITED", "PRIVATELTD", "PVTLIMITED", "PVTLTD",
                 "LIMITED", "PRIVATE", "LLP", "LTD", "PVT", "INC", "PLC",
                 "CORPORATION", "CORP", "COMPANY")


def squash_investor_name(name: str) -> str:
    """Collapse a name to bare alphanumerics with legal suffixes peeled off.

    Exists because OCR sometimes loses the spaces in a filing's investor cell:
    TECHNOCRAF yielded "LRSDSecuritiesPrivateLimited", a single token that
    canonical_investor_tokens cannot split, so it clustered on its own and LRSD
    was reported as two investors of 4 and 1 IPOs instead of one of 5.
    Squashing both spellings gives LRSDSECURITIES either way."""
    s = unicodedata.normalize("NFKD", str(name or "")).encode("ascii", "ignore").decode()
    s = re.sub(r"[^A-Za-z0-9]+", "", s).upper()
    peeled = True
    while peeled:
        peeled = False
        for suf in _GLUED_SUFFIX:
            # Never peel away so much that nothing identifying is left.
            if s.endswith(suf) and len(s) - len(suf) >= 4:
                s, peeled = s[:-len(suf)], True
                break
    return s


_ROMAN_ORDINAL = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6,
                  "VII": 7, "VIII": 8, "IX": 9, "X": 10, "XI": 11, "XII": 12}
# A series/cell/tranche number. Anything larger is not an ordinal but a figure
# that leaked in from the column beside the name cell.
_MAX_ORDINAL = 20


def _fold_token(t: str) -> str:
    """Fold a plural onto its singular so VENTURES and VENTURE key alike.

    Crude on purpose. The result is never displayed, only compared, so it has
    to be consistent rather than correct: SECURITIES and SECURITIE both key as
    SECURITIE, which is all the comparison needs."""
    return t[:-1] if t.endswith("S") and len(t) > 4 else t


def investor_key(name: str) -> "tuple[str, ...]":
    """The identity two spellings must share to be the same investor.

    `canonical_investor_tokens` plus the two things that were splitting real
    rows in half:

      * plurals fold, so "RELIGO COMMODITIES VENTURES TRUST" and "Religo
        Commodities Venture Trust-..." stop being two investors;
      * a series ordinal is normalised whether it is written in roman or in
        arabic, and is KEPT as a trailing "#n" rather than silently dropped.
        Both halves matter: keeping it is what stops SB OPPORTUNITIES II from
        being absorbed into SB OPPORTUNITIES, and normalising it is what lets
        "TIGER STRATEGIES FUND - 1" and "Tiger Strategies Fund-I" collapse onto
        one key instead of one being dropped and one kept.

    The ordinal is only ever read off the LAST token. A series number is written
    at the end of a name; a number anywhere else is an interior row serial
    ("STEADVIEW CAPITAL MASTER 6 FUND LTD") or a figure that leaked in from the
    column beside it ("SBI CONSUMPTION OPPORTUNITIES FUND 2,02;222 2.58%
    675.00"), and reading either as a series split those investors in two."""
    s = unicodedata.normalize("NFKD", str(name or "")).encode("ascii", "ignore").decode()
    words = re.sub(r"[^A-Za-z0-9]+", " ", s).upper().split()
    while words and words[0].isdigit():
        words = words[:0] + words[1:]       # row serial from the filing's table
    ordinal = ""
    if words:
        last = words[-1]
        if last in _ROMAN_ORDINAL:
            ordinal = str(_ROMAN_ORDINAL[last])
            words = words[:-1]
        elif last.isdigit() and 1 <= int(last) <= _MAX_ORDINAL:
            ordinal = str(int(last))
            words = words[:-1]
    out: list[str] = []
    seen: set[str] = set()
    for i, t in enumerate(words):
        if t.isdigit() or t in _ROMAN_ORDINAL:
            continue
        if len(t) < 2:
            continue
        if t in _NAME_NOISE:
            # "SERIES" is structural only where it NUMBERS something, and any
            # such ordinal has already been taken off the end above. Where a
            # fund is actually called that, the word is part of its name:
            # dropping it made MEESHO's FIDELITY SERIES EMERGING MARKETS FUND
            # identical to its FIDELITY EMERGING MARKETS FUND, so two anchors
            # holding 2,408,535 and 828,225 shares merged into a single row and
            # the smaller of them left the workbook.
            if not (t == "SERIES"
                    and not _is_ordinal_token(words[i + 1]
                                              if i + 1 < len(words) else "")):
                continue
        t = _fold_token(t)
        if t in seen:
            continue
        seen.add(t)
        out.append(t)
    return tuple(out) + ((f"#{ordinal}",) if ordinal and out else ())


def _key_split(k: "tuple[str, ...]") -> "tuple[tuple[str, ...], str]":
    return (k[:-1], k[-1]) if k and k[-1].startswith("#") else (k, "")


def _investor_matches(key: "tuple[str, ...]",
                      candidates: "list[tuple[str, ...]]") -> list:
    """Every candidate key that plausibly names the same investor as `key`.

    Four tests, tried in order and never mixed — the first that hits any
    candidate settles the answer, so a weaker test can never override the exact
    match that a stronger one already found:

      1. identical.
      2. same base, one side carrying no ordinal. IPOPlatform files "Upsurge
         Opportunities Fund 1" where the filing says "Upsurge Opportunities
         Fund". Test 1 runs first, so where a list holds BOTH the bare name and
         an ordinal-bearing one the two series can never be merged together.
      3. one base is a subset of the other, same head token, at most
         _MAX_EXTRA_TOKENS more, and the extra tokens carry no ordinal — a name
         one reading cut short. Deliberately symmetric: the old guard only ever
         discarded the shorter spelling, which meant whichever reading happened
         to be the backbone decided the outcome.
      4. identical once spaces and legal suffixes are squeezed out, for the
         glued OCR cell ("LRSDSecuritiesPrivateLimited") that has no separate
         tokens to compare at all.

    The head-token test in 3 is what stops a cell holding two run-together
    investors from acting as a bridge between them."""
    if not key:
        return []
    base, ordv = _key_split(key)
    exact = [c for c in candidates if c == key]
    if exact:
        return exact
    loose = []
    for c in candidates:
        cb, co = _key_split(c)
        if cb == base and (not ordv or not co):
            loose.append(c)
    if loose:
        return loose
    frag = []
    for c in candidates:
        cb, co = _key_split(c)
        if not cb or not base or cb[0] != base[0]:
            continue
        short, long_ = (base, cb) if len(base) < len(cb) else (cb, base)
        if set(short) < set(long_) and len(long_) - len(short) <= _MAX_EXTRA_TOKENS:
            frag.append(c)
    if frag:
        return frag
    # Folded after squashing, not before: the glued spelling is one token that
    # never had a plural to fold, so "LRSDSecuritiesPrivateLimited" squashes to
    # LRSDSECURITIES while the spaced spelling is already LRSD + SECURITIE.
    sq = _fold_token(squash_investor_name("".join(base)))
    if not sq:
        return []
    return [c for c in candidates
            if _fold_token(squash_investor_name("".join(_key_split(c)[0]))) == sq]


def match_investor_key(key: "tuple[str, ...]",
                       candidates: "list[tuple[str, ...]]"):
    """The ONE candidate that means the same investor as `key`, else None.

    Ambiguity is refused rather than guessed: filing one investor's allocation
    onto another's row is a worse error than leaving two rows that need a human
    eye."""
    hits = _investor_matches(key, candidates)
    return hits[0] if len(hits) == 1 else None


def cluster_investor_names(counts: "dict[str, int]") -> "dict[str, str]":
    """Map every raw investor name onto the label of the investor it belongs to.

    Two passes. First, names sharing an identical canonical token list are one
    investor — that alone absorbs case, punctuation and legal-suffix variants.
    Second, a shorter name is absorbed into a fuller one, but only when all of
    the following hold, each guarding against a real failure seen in the data:

      * the HEAD token matches. Without it a cell reading "SHINE STAR BUILD CAP
        PVT LTD -_ GALAXY NOBLE GLOBAL" — two investors run together by OCR —
        becomes a bridge that merges Galaxy Noble Global into Shine Star.
      * the fuller name adds at most _MAX_EXTRA_TOKENS.
      * the fuller name is not a one-off. A spelling seen once is far more
        likely to be an OCR artefact than a real fuller name, and letting one
        act as a merge target is what fused HDFC Manufacturing Fund with HDFC
        Business Cycle Fund through a single run-together cell.
      * the choice is unambiguous. Where several fuller names qualify the most
        common wins; on a tie the name is left on its own rather than guessed.

    The label of a group is its tidiest spelling — most common, then least
    mangled (see _label_penalty), then longest — shortened to its first few
    identifying tokens by _short_labels. The filed spellings are kept and shown
    on the Investor Profile tab."""
    buckets: dict[tuple, dict] = {}
    for name, n in counts.items():
        toks = investor_key(name)
        if not toks:
            continue
        b = buckets.setdefault(toks, {"names": {}, "n": 0})
        b["names"][name] = b["names"].get(name, 0) + n
        b["n"] += n

    keys = sorted(buckets, key=len)
    parent = {k: k for k in keys}

    def root(k):
        while parent[k] != k:
            k = parent[k]
        return k

    def union(a, b):
        ra, rb = root(a), root(b)
        if ra != rb:
            parent[ra] = rb

    # Pass 0: buckets that are the same name once spaces and legal suffixes are
    # taken out. This is the only thing that can rejoin a run-together spelling
    # such as LRSDSecuritiesPrivateLimited to LRSD Securities Pvt Ltd, because
    # the glued form is a single token with no head token in common.
    by_squash: dict[str, list] = {}
    for k in keys:
        sq = _fold_token(squash_investor_name("".join(_key_split(k)[0])))
        if sq:
            by_squash.setdefault(sq, []).append(k)
    for group in by_squash.values():
        for other in group[1:]:
            union(other, group[0])

    # An ordinal tells two series apart inside ONE anchor book, where both are
    # listed side by side. The pooled register has no such context — one filing
    # writes "RMS Growth Fund Scheme 1" and another just "RMS GROWTH" — so here
    # a bare name and its ordinal-bearing twin are taken as one investor.
    bare = {k for k in keys if not _key_split(k)[1]}
    for k in keys:
        base, ordv = _key_split(k)
        if ordv and base in bare:
            union(k, base)

    for k in keys:
        ks = set(k)
        cands = [o for o in keys
                 if len(o) > len(k)
                 and o[0] == k[0]
                 and len(o) - len(k) <= _MAX_EXTRA_TOKENS
                 and buckets[o]["n"] > 1
                 and ks < set(o)]
        if not cands:
            continue
        cands.sort(key=lambda o: (-buckets[o]["n"], len(o)))
        if len(cands) > 1 and buckets[cands[0]]["n"] == buckets[cands[1]]["n"]:
            continue                       # ambiguous — refuse to guess
        union(k, cands[0])

    grouped: dict[tuple, dict] = {}
    for k in keys:
        g = grouped.setdefault(root(k), {})
        for name, n in buckets[k]["names"].items():
            g[name] = g.get(name, 0) + n

    label_of: dict[str, str] = {}
    clean = {k: max(names, key=lambda n: (names[n], -_label_penalty(n), len(n)))
             for k, names in grouped.items()}
    short = _short_labels(clean)
    for k, names in grouped.items():
        for name in names:
            label_of[name] = short[k]
    return label_of


def investor_labeller(anchor_map: "dict[str, list[Anchor]]",
                      words: int = 2, smart: bool = True):
    """Return a function mapping a raw anchor name to its investor label."""
    if not smart:
        return lambda name: investor_key(name, words)
    counts: dict[str, int] = {}
    for anchors in anchor_map.values():
        for a in anchors:
            nm = str(a.name).strip()
            if nm:
                counts[nm] = counts.get(nm, 0) + 1
    label_of = cluster_investor_names(counts)
    return lambda name: label_of.get(str(name).strip(), "")


def build_investor_frequency(anchor_map: dict[str, list[Anchor]],
                             words: int = 2,
                             smart: bool = True,
                             remap: "dict[str, str] | None" = None,
                             qualified: "dict[str, bool] | None" = None,
                             company_of: "dict[str, str] | None" = None,
                             participation: "dict[str, set] | None" = None) -> pd.DataFrame:
    """Frequency table over every anchor name in `anchor_map`.

    Three counts are reported on purpose:
      * Total IPOs     — every issue the investor anchored in the window.
      * Qualified IPOs — those that cleared the gain threshold.
      * Hit Rate %     — the second as a share of the first.

    The two IPO counts come from different places by necessity. Anchor filings
    are downloaded only for issues that CLEARED the threshold, so our own books
    can supply the numerator but not the denominator — counting participations
    from them alone would divide winners by winners. `participation` carries
    IPOPlatform's dated deal list, which covers issues that went nowhere too.

    Where an investor is absent from that register the total is left BLANK
    rather than filled from our own books: blank reads as unknown, whereas a
    number there would assert a 100% hit rate we have no evidence for.

    `remap` folds together groups that Screener has since proved to be one
    investor (see _split_investor_remap). It is passed in rather than derived
    here so this sheet and the Investor Profile tab group identically."""
    label = investor_labeller(anchor_map, words, smart)
    remap = remap or {}
    company_of = company_of or {}
    occ: dict[str, int] = {}
    syms: dict[str, list[str]] = {}
    for sym, anchors in anchor_map.items():
        for a in anchors:
            if not str(a.name).strip():
                continue
            k = label(a.name)
            if not k:
                continue
            k = remap.get(k, k)
            occ[k] = occ.get(k, 0) + 1
            bucket = syms.setdefault(k, [])
            if sym not in bucket:
                bucket.append(sym)
    if not occ:
        return pd.DataFrame(columns=["Investor", "Total IPOs",
                                     "Qualified IPOs", "Hit Rate %", "Symbols"])
    rows = []
    for k in occ:
        got = syms[k]
        hits = (sum(1 for s in got if qualified.get(s)) if qualified is not None
                else len(got))
        seen = participation.get(k) if participation else None
        total = None
        if seen is not None:
            # Our own issues are folded in as well: a filing we read that their
            # register happens not to list is still a participation.
            mine = {_company_core(company_of.get(s, s)) for s in got}
            total = len(seen | {c for c in mine if c})
        rows.append({"Investor": k,
                     "Total IPOs": total,
                     "Qualified IPOs": hits,
                     "Hit Rate %": (round(100.0 * hits / total, 1)
                                    if total else None),
                     "Symbols": ", ".join(got)})
    df = pd.DataFrame(rows)
    # Nullable ints: a count is a whole number, and a plain int column cannot
    # hold the blank that stands for 'IPOPlatform does not list this investor'.
    for col in ("Total IPOs", "Qualified IPOs"):
        df[col] = df[col].astype("Int64")
    return df.sort_values(["Qualified IPOs", "Total IPOs", "Investor"],
                          ascending=[False, False, True]).reset_index(drop=True)


# Sheet 2's column contract. Column A is the ISSUER, then one row per anchor.
ANCHOR_COLUMNS = ("Company", "Anchor Investor", "Amount Invested ₹",
                  "Shares Allotted", "Price ₹")


def read_existing_anchors(out_xlsx: Path) -> dict[str, list[Anchor]]:
    """Anchor rows already saved in the workbook, keyed by their Company cell.

    Sheet 2 gets curated by hand between runs (rows typed in from filings this
    script can't reach, and the handful of names OCR spells wrong), so a re-run
    must NEVER silently discard it.

    Workbooks written before Sheet 2 became tabular are still readable: they
    held one COLUMN per symbol with names running down and no figures at all,
    so they are transposed into name-only rows here rather than being dropped."""
    if not out_xlsx.exists():
        return {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(out_xlsx, read_only=True, data_only=True)
        if "Anchor Investors" not in wb.sheetnames:
            wb.close()
            return {}
        rows = list(wb["Anchor Investors"].iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        print(f"  [anchor] could not read existing workbook ({e}); "
              "continuing WITHOUT merge.", file=sys.stderr)
        return {}
    if not rows:
        return {}

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: dict[str, list[Anchor]] = {}

    if headers[:2] == list(ANCHOR_COLUMNS[:2]):
        for r in rows[1:]:
            cells = list(r) + [None] * (len(ANCHOR_COLUMNS) - len(r))
            key = str(cells[0]).strip() if cells[0] is not None else ""
            name = str(cells[1]).strip() if cells[1] is not None else ""
            if not key or not name:
                continue
            out.setdefault(key, []).append(Anchor(
                name=name,
                shares=_to_int(str(cells[3])) if cells[3] is not None else None,
                price=_to_num(str(cells[4])) if cells[4] is not None else None,
                amount=_to_num(str(cells[2])) if cells[2] is not None else None))
        return out

    for ci, h in enumerate(headers):                       # legacy wide layout
        if not h:
            continue
        col = [Anchor(name=str(r[ci]).strip()) for r in rows[1:]
               if ci < len(r) and r[ci] is not None and str(r[ci]).strip()]
        if col:
            out[h] = col
    if out:
        print(f"  [anchor] migrated {sum(len(v) for v in out.values())} name(s) "
              "from the previous wide Sheet-2 layout (figures were not stored "
              "there, so they will be filled in from the filings).")
    return out


def merge_anchors(existing: list[Anchor], fresh: list[Anchor]) -> list[Anchor]:
    """Existing rows first (preserving hand-curated order), then genuinely new
    ones. De-duplicated on `investor_key`, so a second source that spells a name
    differently fills the existing row in rather than sitting beside it.

    That key, not the bare alphanumerics, is what makes this work across
    sources. IPOPlatform writes "LRSD Securities Private Limited", "Inti Capital
    VCC-Inti Capital 1" and "Meru Investment Fund Pcc" for rows the filing spells
    "LRSD SECURITIES PVT LTD", "INTI CAPITAL VCC - INTI CAPITAL I" and "MERU
    INVESTMENT FUND PCC - CELL 1"; on exact matching each pair became two rows
    for one allocation, one of them permanently without figures, and then
    double-counted in the frequency sheet.

    Where an existing row has no figures — hand-typed, or carried over from the
    old wide layout — the freshly parsed figures are filled in. Two existing
    spellings that key alike are folded the same way rather than the first one
    simply winning: Urban Company held both "PGIM INDIA SMALL CAP FUND" with no
    figures and "PGIM INDIA TRUSTEE PRIVATE LIMITED A/C - PGIM INDIA SMALL CAP
    FUND" with 789,235 shares, and keeping the first would have thrown the
    allocation away."""
    fresh_keyed = [(investor_key(a.name), a) for a in fresh]
    by_key: dict = {}
    for k, a in fresh_keyed:
        if k:
            by_key.setdefault(k, a)

    def _fold(keep: Anchor, other: Anchor) -> Anchor:
        """One row out of two spellings of one investor: blanks are filled from
        whichever row has the figure, and the tidier spelling is displayed."""
        better = other if (_label_penalty(other.name),
                           -len(other.name)) < (_label_penalty(keep.name),
                                                -len(keep.name)) else keep
        return Anchor(name=better.name,
                      shares=keep.shares if keep.shares is not None else other.shares,
                      price=keep.price if keep.price is not None else other.price,
                      amount=keep.amount if keep.amount is not None else other.amount)

    out: list[Anchor] = []
    at: dict = {}
    claimed: set = set()
    for a in existing:
        # Debris no key can identify ("SERIES III A") is kept rather than
        # dropped, on its literal spelling, so it can never match anything else.
        k = investor_key(a.name) or ("\0", _norm_alnum(a.name))
        if k == ("\0", ""):
            continue
        if k in at:
            out[at[k]] = _fold(out[at[k]], a)
            continue
        src_key = match_investor_key(k, [c for c in by_key if c not in claimed])
        at[k] = len(out)
        if src_key is None:
            out.append(a)
            continue
        claimed.add(src_key)
        at[src_key] = at[k]
        out.append(_fold(a, by_key[src_key]))
    for k, a in fresh_keyed:
        if not k:
            continue
        if k in at:
            out[at[k]] = _fold(out[at[k]], a)
            continue
        at[k] = len(out)
        out.append(a)
    return out


def read_extra_sheets(out_xlsx: Path) -> "dict[str, list[list]]":
    """Capture the raw cell values of any sheet this script does not manage.

    pandas' ExcelWriter rebuilds the workbook from scratch, so a sheet the user
    added by hand would silently vanish on the next run. We snapshot those
    sheets here and re-append them after the managed ones are written. Values
    only — formulas are resolved to their last-calculated result and cell
    formatting is not preserved."""
    if not out_xlsx.exists():
        return {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(out_xlsx, read_only=True, data_only=True)
    except Exception as exc:
        print(f"  [warn] could not scan workbook for extra sheets: {exc}")
        return {}
    extra: dict[str, list[list]] = {}
    try:
        for name in wb.sheetnames:
            if name in MANAGED_SHEETS:
                continue
            rows = [[c.value for c in row] for row in wb[name].iter_rows()]
            while rows and all(v is None for v in rows[-1]):
                rows.pop()
            if rows:
                extra[name] = rows
    finally:
        wb.close()
    return extra


def build_merged_anchors(winners: list[IPO], anchor_map: dict[str, list[Anchor]],
                         existing: dict[str, list[Anchor]],
                         rebuild: bool = False,
                         universe: "list[IPO] | None" = None,
                         verified: "set[str] | None" = None
                         ) -> dict[str, list[Anchor]]:
    """Fold this run's parsed anchors into whatever the workbook already held.

    Keys stay INTERNAL (NSE symbol, or a filename-derived key for a filing that
    matches no winner) because Sheet 3 reports symbols; the Company column is
    only how a key is rendered. Existing rows are looked up under both forms so
    a workbook written before Sheet 2 was keyed by company still merges.

    `rebuild` drops the stored rows for any symbol this run re-read from its
    filing, instead of merging them. Merging assumes the stored names are at
    least as good as the fresh ones, and after a parser fix they are not: a
    corrupt name like "26 Sanshi Fund - I CITI GROUP GLOBAL MARKETS" matches no
    fresh name, so it survives ALONGSIDE the two correct rows that replace it.
    Symbols with no filing to re-read are left alone either way, so anything
    hand-entered is kept.

    `verified` does the same thing per symbol, but on evidence rather than on a
    flag: those filings were read to a table whose shares sum EXACTLY to the
    declared allocation, which proves no investor is missing from it. Merging
    such a reading can only re-admit the debris of earlier runs — MEESHO read
    125 names, reconciled to the share, and still published 296 rows because
    the previous run's "gna EMERGING MARKETS FUND BLACKROCK GLOBAL FUNDS Ls
    lane" matched nothing fresh and so survived."""
    disp = {i.symbol: i.company for i in (universe if universe is not None
                                          else winners)}
    used: set[str] = set()
    proven = verified or set()

    def _prior(key: str, fresh: list[Anchor]) -> list[Anchor]:
        for cand in (disp.get(key, key), key):
            if cand in existing:
                used.add(cand)
                if fresh and (rebuild or key in proven):
                    return []
                return existing[cand]
        return []

    merged: dict[str, list[Anchor]] = {}
    for w in winners:
        fresh = anchor_map.get(w.symbol, [])
        merged[w.symbol] = merge_anchors(_prior(w.symbol, fresh), fresh)
    # Filings that matched no winner (typically an IPO that has not listed yet)
    # are keyed in anchor_map by filing name rather than symbol. Without this
    # they would be parsed and then thrown away at the write step.
    for key, anchors in anchor_map.items():
        if key not in merged:
            merged[key] = merge_anchors(_prior(key, anchors), anchors)
    for key, anchors in existing.items():
        if key not in used and anchors:
            merged[key] = anchors
    # Last line of defence: two keys that render to the same Company label would
    # print as a single block holding every name twice. Fold them into one.
    by_label: dict[str, str] = {}
    for key in list(merged):
        first = by_label.setdefault(disp.get(key, key), key)
        if first != key:
            merged[first] = merge_anchors(merged[first], merged.pop(key))
    return merged


def anchors_to_frame(merged: dict[str, list[Anchor]],
                     winners: list[IPO],
                     universe: "list[IPO] | None" = None) -> pd.DataFrame:
    """Sheet 2 as a flat table: one row per (issuer, anchor investor)."""
    disp = {i.symbol: i.company for i in (universe if universe is not None
                                          else winners)}
    rows = [{ANCHOR_COLUMNS[0]: disp.get(key, key),
             ANCHOR_COLUMNS[1]: a.name,
             ANCHOR_COLUMNS[2]: a.amount,
             ANCHOR_COLUMNS[3]: a.shares,
             ANCHOR_COLUMNS[4]: a.price}
            for key, anchors in merged.items() for a in anchors
            if str(a.name).strip()]
    return pd.DataFrame(rows, columns=list(ANCHOR_COLUMNS))


def write_workbook(all_df: pd.DataFrame, gainers_df: pd.DataFrame,
                   merged: dict[str, list[Anchor]],
                   anchors_df: pd.DataFrame, freq_df: pd.DataFrame,
                   out_xlsx: Path,
                   tracked_df: "pd.DataFrame | None" = None) -> None:
    """Write the managed workbook.

      Sheet 1 'All IPOs'           — every IPO in the window, qualified or not.
      Sheet 2 'Gainers'            — qualifying IPOs, one row each.
      Sheet 3 'Anchor Investors'   — one row per anchor: issuer, investor,
                                     amount invested, shares, price.
      Sheet 4 'Investor Frequency' — how often each investor recurs.
      Sheet 5 'Tracked Investors'  — watchlist participation, outcome-blind.

    Sheet 3 is MERGED upstream, never overwritten, so rows already in the file
    keep their place and their hand-typed corrections.

    Any OTHER sheet found in the file — one you added by hand — is snapshotted
    before the rewrite and restored afterwards (values only, no formatting)."""
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment

    wrap = Alignment(wrap_text=True, vertical="top")
    extra_sheets = read_extra_sheets(out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    # Build in a temp file and swap in only once it is complete. pandas saves
    # the workbook on the way out of the `with` block even when an exception is
    # in flight, so writing in place means a crash midway leaves a file that has
    # the managed sheets but has LOST every user sheet.
    tmp = out_xlsx.with_name(out_xlsx.stem + ".tmp.xlsx")
    with pd.ExcelWriter(tmp, engine="openpyxl") as xw:
        managed = [("All IPOs", all_df), ("Gainers", gainers_df),
                   ("Anchor Investors", anchors_df),
                   ("Investor Frequency", freq_df)]
        if tracked_df is not None:
            managed.append(("Tracked Investors", tracked_df))
        for sheet, d in managed:
            d.to_excel(xw, sheet_name=sheet, index=False)
        for sheet, d in managed:
            ws = xw.sheets[sheet]
            for i, col in enumerate(d.columns):
                # len(str(v)) per cell rather than .astype(str).map(len): a
                # numeric column that is entirely empty stays float NaN through
                # astype(str) and len() then blows up.
                body = d[col].map(lambda v: len(str(v))).max() if not d.empty else 0
                width = min(48, max(12, int(max(body, len(str(col))) + 2)))
                letter = get_column_letter(i + 1)
                ws.column_dimensions[letter].width = width
                # Anything past the width cap would otherwise run off under the
                # neighbouring cells, so it is wrapped inside its own column.
                if body > 48:
                    for r in range(2, len(d) + 2):
                        ws[f"{letter}{r}"].alignment = wrap
        ws = xw.sheets["Anchor Investors"]
        for col, fmt in ((3, "#,##0"), (4, "#,##0"), (5, "#,##0.00")):
            letter = get_column_letter(col)
            for r in range(2, len(anchors_df) + 2):
                ws[f"{letter}{r}"].number_format = fmt
        for name, rows in extra_sheets.items():
            ws = xw.book.create_sheet(title=name[:31])
            for row in rows:
                ws.append(row)
        if extra_sheets:
            print(f"  [xlsx] preserved {len(extra_sheets)} user sheet(s): "
                  f"{', '.join(extra_sheets)}.")
    tmp.replace(out_xlsx)


# ─────────────────────── Screener.in investor profiles ─────────────────────
# Screener has no way to look an investor up by name — there is no investor
# index, and its search covers companies only. The only route to "everything
# this fund owns" is therefore backwards:
#
#   company page  ->  /api/3/<companyId>/investors/<class>/<period>/
#                     returns {investor name: {quarter: pct, ...,
#                              "setAttributes": {"data-person-url": "/people/..."}}}
#   that person URL -> /people/<id>/<slug>/
#                     a full cross-company holdings table for that investor
#
# So: resolve each gainer to a Screener company id, pull its named institutional
# and public shareholders, and keep the person URLs that come back. That gives a
# name -> person-page index built from exactly the companies our anchors bought
# into, which is where they are most likely to be found. Anchors that have since
# exited every one of those companies, or that sit below Screener's 1% naming
# threshold, will not resolve — that is a limit of the source, not a bug.
SCREENER_BASE = "https://www.screener.in"
SCREENER_CACHE_DIR = CACHE_DIR / "screener"
SCREENER_TTL_HOURS = 24 * 7
# Anchor books are institutions and large non-institutional funds. Promoters and
# government holdings are never anchors, so those two classifications are skipped.
SCREENER_CLASSES = ("foreign_institutions", "domestic_institutions", "public")
# Pacing. Screener bans by IP, silently and durably, so this is set slow on
# purpose — a full cold run is a few minutes, every later run is served from
# .cache/ipo_gainers/screener/ and costs nothing.
SCREENER_DELAY = 1.1          # seconds between successful requests
SCREENER_BACKOFF = 20.0       # seconds to wait after a refused connection
SCREENER_MAX_FAILS = 4        # consecutive failures before abandoning the stage
# A holdings header cell: "Mar 2016", "Jun 2026". Distinguishes the holdings
# table from the deal tables, which a person can have more than one of.
_QUARTER_LABEL = re.compile(r"^([A-Z][a-z]{2})\s+(\d{4})$")
_MONTHS = {m: i for i, m in enumerate(
    ("Jan", "Feb", "Mar", "Apr", "May", "Jun",
     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"), 1)}
_SCREENER_FAILS = 0


_PERSON_ID = re.compile(r"^/people/(\d+)/")


def _person_id(url: str) -> str:
    """The numeric id in /people/<id>/<slug>/, which is the real identifier.

    The slug is cosmetic and Screener redirects any slug to the canonical one,
    so the same person arrives under different spellings depending on which
    company page linked to them — LRSD's id 167947 is linked both as
    lrsd-securites-pvt-ltd and lrsd-securities-pvt-ltd. Keying on the id stops
    one record being fetched and merged twice."""
    m = _PERSON_ID.match(str(url or ""))
    return m.group(1) if m else str(url or "")


def _quarter_sort_key(label: str) -> "tuple[int, int]":
    """Order 'Mar 2016' before 'Jun 2026'. Unparseable labels sort last."""
    m = _QUARTER_LABEL.match(str(label).strip())
    if not m:
        return (9999, 99)
    return (int(m.group(2)), _MONTHS.get(m.group(1), 99))


def _deal_sort_key(date_text: str) -> "tuple[int, int, int]":
    """Order '06 Aug, 2026' style deal dates, newest first when negated."""
    m = re.match(r"(\d{1,2})\s+([A-Za-z]{3})\w*,?\s+(\d{4})", str(date_text).strip())
    if not m:
        return (0, 0, 0)
    return (int(m.group(3)), _MONTHS.get(m.group(2).title(), 0), int(m.group(1)))


def _screener_cache(key: str, ttl_hours: float = SCREENER_TTL_HOURS) -> str | None:
    """Return cached body for `key` if it is present and still fresh."""
    fp = SCREENER_CACHE_DIR / (hashlib.sha1(key.encode()).hexdigest() + ".json")
    if not fp.exists():
        return None
    try:
        blob = json.loads(fp.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None
    if time.time() - blob.get("t", 0) > ttl_hours * 3600:
        return None
    return blob.get("body")


def _screener_store(key: str, body: str) -> None:
    SCREENER_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    fp = SCREENER_CACHE_DIR / (hashlib.sha1(key.encode()).hexdigest() + ".json")
    try:
        fp.write_text(json.dumps({"t": time.time(), "key": key, "body": body}),
                      encoding="utf-8")
    except OSError:
        pass


def screener_session() -> "tuple[requests.Session, bool]":
    """Return a Screener session and whether the login succeeded.

    A session is always returned, even unauthenticated: cached responses are
    served without touching the network, so an expired login or a rate-limit
    block still leaves the profile tab populated from the last good run."""
    import os
    s = requests.Session()
    s.headers.update({"User-Agent": UA, "Accept-Language": "en-US,en;q=0.9"})
    user, pwd = os.environ.get("SCREENER_USER"), os.environ.get("SCREENER_PASS")
    if not user or not pwd:
        return s, False
    try:
        page = s.get(f"{SCREENER_BASE}/login/", timeout=20)
        m = re.search(r'name="csrfmiddlewaretoken"[^>]*?value="([^"]+)"', page.text)
        if not m:
            m = re.search(r'value="([^"]+)"[^>]*?name="csrfmiddlewaretoken"', page.text)
        if not m:
            return s, False
        r = s.post(f"{SCREENER_BASE}/login/",
                   data={"username": user, "password": pwd, "next": "/",
                         "csrfmiddlewaretoken": m.group(1)},
                   headers={"Referer": f"{SCREENER_BASE}/login/"},
                   timeout=25)
        return s, r.status_code < 400
    except requests.RequestException:
        return s, False


def _screener_get(session: requests.Session, path: str,
                  ttl_hours: float = SCREENER_TTL_HOURS) -> str | None:
    """Cached, rate-limited GET against Screener. `path` is site-relative.

    Screener blocks a client at the TCP level — connection refused, not 429 —
    once it decides the request rate is abusive, and the block outlives the
    process. Hence the deliberate pacing, the backoff on connection errors, and
    the circuit breaker: after a run of hard failures the whole enrichment gives
    up rather than hammering a host that has already shut the door. Everything
    fetched before that point stays cached, so a later run resumes instead of
    starting over."""
    global _SCREENER_FAILS
    hit = _screener_cache(path, ttl_hours)
    if hit is not None:
        return hit or None
    if _SCREENER_FAILS >= SCREENER_MAX_FAILS:
        return None
    # /api/ paths are XHR endpoints; ordinary pages are not, and asking for one
    # with an XHR header invites a fragment instead of the page.
    headers = {"X-Requested-With": "XMLHttpRequest"} if path.startswith("/api/") else {}
    for attempt in range(3):
        try:
            r = session.get(SCREENER_BASE + path, timeout=25, headers=headers)
        except requests.RequestException:
            time.sleep(SCREENER_BACKOFF * (attempt + 1))
            continue
        _SCREENER_FAILS = 0
        time.sleep(SCREENER_DELAY)
        if r.status_code == 429:
            time.sleep(SCREENER_BACKOFF)
            continue
        if r.status_code != 200:
            _screener_store(path, "")      # cache the miss; 404s do not heal
            return None
        _screener_store(path, r.text)
        return r.text
    _SCREENER_FAILS += 1
    if _SCREENER_FAILS == SCREENER_MAX_FAILS:
        print("  [screener] connection repeatedly refused — Screener is rate "
              "limiting this IP. Skipping the rest; cached data is kept and the "
              "next run resumes from it.")
    return None


def _company_search(session: requests.Session, q: str) -> list:
    """One /api/company/search/ call, decoded, never raising."""
    body = _screener_get(session, f"/api/company/search/?q={requests.utils.quote(q)}")
    if not body:
        return []
    try:
        hits = json.loads(body)
    except json.JSONDecodeError:
        return []
    return hits if isinstance(hits, list) else []


def screener_company_id(session: requests.Session, name: str) -> int | None:
    """Resolve an issuer name to a Screener company id via its search API.

    Screener's search matches on a prefix of the registered name, so the full
    legal name from the NSE list often returns nothing at all: "Manas Polymers
    and Energies Limited" misses because Screener spells it with an ampersand,
    and "Sat Kartar Shopping Limited" misses because the company now trades as
    Sat Kartar Life. The query is therefore shortened a word at a time until
    something comes back, and a hit is only accepted while its leading words
    still agree with ours, so shortening cannot drift onto another company."""
    base = re.sub(r"\s*\b(limited|ltd\.?|private|pvt\.?)\b\s*$", "", name,
                  flags=re.I).strip()
    words = base.split()
    tried: set[str] = set()
    for variant in (base, base.replace("&", "and"), base.replace(" and ", " & ")):
        for cut in range(len(words), 0, -1):
            q = " ".join(variant.split()[:cut])
            if len(q) < 3 or q in tried:
                continue
            tried.add(q)
            hits = _company_search(session, q)
            if not hits:
                continue
            want = _norm_alnum(base)
            for h in hits:
                nm = re.sub(r"\s*\bltd\.?$", "", str(h.get("name", "")), flags=re.I)
                if _norm_alnum(nm) == want:
                    return h.get("id")
            # No exact name; accept the single best hit only if it still starts
            # with the words we searched on.
            qk = _norm_alnum(q)
            near = [h for h in hits if _norm_alnum(str(h.get("name", ""))).startswith(qk)]
            if len(near) == 1:
                return near[0].get("id")
            if cut == len(words) and hits:
                return hits[0].get("id")    # full name matched something
    return None


def screener_people_for_company(session: requests.Session,
                                cid: int) -> "list[tuple[str, str]]":
    """Named shareholders of one company -> [(display name, person URL)].

    Only investors that are STILL named holders appear here; Screener drops a
    shareholder from this API once it falls below the disclosure threshold."""
    found: list[tuple[str, str]] = []
    for cls in SCREENER_CLASSES:
        body = _screener_get(session, f"/api/3/{cid}/investors/{cls}/quarterly/")
        if not body:
            continue
        try:
            payload = json.loads(body)
        except json.JSONDecodeError:
            continue
        if not isinstance(payload, dict):
            continue
        for disp, rec in payload.items():
            if not isinstance(rec, dict):
                continue
            url = (rec.get("setAttributes") or {}).get("data-person-url")
            if url:
                found.append((str(disp).strip(), url))
    return found


def screener_people_in_trades(session: requests.Session,
                              cid: int) -> "list[tuple[str, str]]":
    """Everyone who has dealt in one company -> [(display name, person URL)].

    The companion to screener_people_for_company, and not optional. Screener
    has no investor search and a person page is reachable only by its numeric
    id, so the only way to learn an investor's id is to find it linked from a
    company we already know. The shareholder API lists CURRENT holders alone,
    which is why LRSD Securities was invisible on Exim Routes — 5 people there
    against 19 on this page — while the deals it did are recorded here."""
    body = _screener_get(session, f"/trades/company-{cid}/")
    if not body:
        return []
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return []
    found: list[tuple[str, str]] = []
    for a in BeautifulSoup(body, "html.parser").find_all("a", href=True):
        href = a["href"]
        if href.startswith("/people/"):
            nm = a.get_text(" ", strip=True)
            if nm:
                found.append((nm, href))
    return found


def _table_heading(table) -> str:
    """The section heading a table sits under: Bulk Deals, Shareholding, ..."""
    node = table
    for _ in range(6):
        node = node.parent
        if node is None:
            break
        hd = node.find(["h1", "h2", "h3", "h4"])
        if hd:
            return hd.get_text(" ", strip=True)
    return ""


def screener_person_holdings(session: requests.Session,
                             person_url: str) -> "dict | None":
    """Scrape one /people/<id>/<slug>/ page in full.

    A person page carries two KINDS of table and can have several of each, so
    they are told apart by their section heading and header row rather than by
    position — Rajasthan Global Securities has three, Bulk Deals then Block
    Deals then Shareholding:

      * Shareholding — first header cell empty, the rest quarter labels. Rows
        are the companies the investor is named in, cells the holding
        percentage. A blank cell means 'not a named holder that quarter',
        which is how entries and exits read off this table.
      * Bulk / Block / Insider deals — header Company / (action) / Quantity /
        Price, with the company in an <a> and the deal date in a <span>. The
        heading is kept per row because the same deal can be disclosed under
        two regimes and the distinction matters.

    Returns {"name", "url", "periods", "rows", "trades"}; trades are
    [company, date, action, quantity, price, kind]. Nothing is trimmed."""
    body = _screener_get(session, person_url)
    if not body:
        return None
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return None
    soup = BeautifulSoup(body, "html.parser")

    periods: list[str] = []
    rows: list[list] = []
    trades: list[list] = []
    for table in soup.find_all("table"):
        trs = table.find_all("tr")
        if len(trs) < 2:
            continue
        head = [c.get_text(" ", strip=True) for c in trs[0].find_all(["th", "td"])]
        if not head:
            continue
        kind = _table_heading(table)
        is_quarters = len(head) > 1 and not head[0] and all(
            _QUARTER_LABEL.match(h) for h in head[1:] if h)
        if is_quarters:
            if periods:
                continue                    # one shareholding table per page
            periods = head[1:]
            for tr in trs[1:]:
                cells = [c.get_text(" ", strip=True) for c in tr.find_all(["td", "th"])]
                if not cells or not cells[0]:
                    continue
                vals = cells[1:len(periods) + 1]
                vals += [""] * (len(periods) - len(vals))
                if any(v for v in vals):
                    rows.append([cells[0]] + vals)
        elif "quantity" in " ".join(head).lower():
            for tr in trs[1:]:
                cells = tr.find_all(["td", "th"])
                if len(cells) < 4:
                    continue
                link = cells[0].find("a")
                when = cells[0].find("span")
                trades.append([
                    link.get_text(" ", strip=True) if link
                    else cells[0].get_text(" ", strip=True),
                    when.get_text(" ", strip=True) if when else "",
                    cells[1].get_text(" ", strip=True),
                    cells[2].get_text(" ", strip=True),
                    cells[3].get_text(" ", strip=True),
                    kind or "Deal",
                ])

    if not rows and not trades:
        return None
    head_el = soup.find(["h1", "h2"])
    return {"name": head_el.get_text(strip=True) if head_el else "",
            "url": person_url, "periods": periods, "rows": rows,
            "trades": trades}


def merge_person_pages(pages: "list[dict]") -> "dict | None":
    """Fold every Screener person record for one investor into a single view.

    Necessary because Screener keys a person record on the exact spelling it
    was first given, so one firm can hold several. LRSD Securities has two —
    /people/186986/ 'Lrsd Securities Pvt Ltd' and /people/167947/ 'LRSD
    Securites Pvt LTD', the second a typo — and their contents are disjoint:
    12 holdings and 6 deals against 5 and 8. Reading either alone loses half
    the firm. That they are one firm is not guessed: 186986 shows Ecoline Exim
    at 1.10% in Sep 2025 and 167947 records the sale of it on 30 Sep 2025.

    Quarters are unioned and ordered chronologically, holdings are keyed by
    company so a gap on one record can be filled from another, and deals are
    concatenated newest first with exact duplicates of the same disclosure
    dropped."""
    pages = [p for p in pages if p]
    if not pages:
        return None

    periods = sorted({q for p in pages for q in p["periods"]}, key=_quarter_sort_key)
    holdings: dict[str, dict[str, str]] = {}
    for p in pages:
        for row in p["rows"]:
            company = row[0]
            cells = holdings.setdefault(company, {})
            for q, v in zip(p["periods"], row[1:]):
                if v and not cells.get(q):
                    cells[q] = v
    rows = [[c] + [holdings[c].get(q, "") for q in periods]
            for c in sorted(holdings, key=str.lower)]

    trades, seen = [], set()
    for p in pages:
        for t in p["trades"]:
            sig = tuple(t)
            if sig in seen:
                continue
            seen.add(sig)
            trades.append(t)
    trades.sort(key=lambda t: _deal_sort_key(t[1]), reverse=True)

    return {"name": pages[0]["name"],
            "url": pages[0]["url"],
            "sources": [[p["name"], p["url"]] for p in pages],
            "periods": periods, "rows": rows, "trades": trades}


def _split_investor_remap(profiles: "dict[str, dict]",
                          wanted: "dict[str, set[str]]") -> "dict[str, str]":
    """Fold together groups that are one investor split by a truncated name.

    Both guards are needed, and neither works alone.

    A shared Screener person id alone is NOT enough. Screener keeps one record
    per legal entity, so /people/153538/ is claimed by six Aditya Birla Sun
    Life schemes — Balanced, Banking, Digital, Equity, Large, Special
    Opportunities — which really are different investors and must stay apart.

    A shared name prefix alone is NOT enough either: plenty of unrelated funds
    share opening words.

    Requiring BOTH — one squashed name a strict prefix of the other AND at
    least one person id in common — merges 'BHARAT VENTURE OPPORTU' into
    'BHARAT VENTURE OPPORTUNITIES', the same name cut short by OCR, while
    leaving the Birla schemes alone because no scheme name is a prefix of
    another. The surviving label is the spelling with the most allocations
    behind it, then the least mangled."""
    ids = {label: {_person_id(u) for u in urls} for label, urls in wanted.items()}
    keys = {label: squash_investor_name(label) for label in ids}
    parent = {label: label for label in ids}

    def root(x):
        while parent[x] != x:
            x = parent[x]
        return x

    # Longest name first, so a chain of truncations lands on the fullest form.
    order = sorted(ids, key=lambda l: (-len(keys[l]), l))
    for a in order:
        ka = keys[a]
        for b in order:
            kb = keys[b]
            if b == a or not kb or len(kb) >= len(ka) or not ka.startswith(kb):
                continue
            if not (ids[a] & ids[b]) or root(a) == root(b):
                continue
            parent[root(b)] = root(a)

    groups: dict[str, list[str]] = {}
    for label in ids:
        groups.setdefault(root(label), []).append(label)

    remap: dict[str, str] = {}
    for members in groups.values():
        if len(members) < 2:
            continue
        keep = max(members, key=lambda m: (len(profiles[m]["anchors"]),
                                           -_label_penalty(m), -len(m)))
        for m in members:
            if m != keep:
                remap[m] = keep
    return remap


def build_investor_profiles_offline(merged: "dict[str, list[Anchor]]",
                                    winners: "list[IPO]",
                                    words: int = 2,
                                    smart: bool = True) -> "dict[str, dict]":
    """Group anchor allocations by investor, with no network access.

    Grouping goes through investor_labeller() — the same clustering the
    Investor Frequency sheet uses — so the two views agree. They must: one fund
    reaches the anchor sheet under several spellings, because the filings
    themselves are inconsistent and OCR adds its own noise. Saint Capital Fund
    arrives as both 'Saint Capital Fund' and 'SAINT CAPITAL FUND'; Finavenue
    under nine variants. Keying on the raw string splits one investor into
    several dropdown entries, each undercounting.

    'ipos' counts DISTINCT issuers, not rows — an investor allotted twice in one
    book is one IPO — which is the same thing the frequency sheet's IPOs column
    counts."""
    disp = {w.symbol: w.company for w in winners}
    label = investor_labeller(merged, words, smart)
    groups: dict[str, dict] = {}
    for key, anchors in merged.items():
        company = disp.get(key, key)
        for a in anchors:
            nm = str(a.name).strip()
            if not nm:
                continue
            gk = label(nm)
            if not gk:
                continue
            g = groups.setdefault(gk, {"anchors": [], "variants": {}})
            g["anchors"].append([company, a.amount, a.shares, a.price, nm])
            g["variants"][nm] = g["variants"].get(nm, 0) + 1

    # The group key IS the label the frequency sheet shows, so reuse it rather
    # than re-deriving one — re-deriving is how the two views drifted apart.
    profiles: dict[str, dict] = {}
    for gk, g in groups.items():
        profiles[gk] = {
            "anchors": g["anchors"],
            "screener": None,
            "variants": sorted(g["variants"], key=lambda n: -g["variants"][n]),
            "ipos": len({r[0] for r in g["anchors"]}),
        }
    return profiles


def build_investor_profiles(merged: "dict[str, list[Anchor]]",
                            winners: "list[IPO]",
                            words: int = 2,
                            smart: bool = True) -> "tuple[dict[str, dict], dict[str, str]]":
    """Anchor groupings from build_investor_profiles_offline, enriched with the
    holdings history Screener carries for each investor.

    Returns the profiles and the label remap that Screener's own identities
    justified, which the caller must apply to the frequency sheet so the two
    views keep grouping alike.

    Every step degrades quietly: no credentials, no network or no match simply
    means that investor shows anchor data alone."""
    profiles = build_investor_profiles_offline(merged, winners, words, smart)

    session, authed = screener_session()
    if not authed:
        print("  [screener] not logged in (SCREENER_USER / SCREENER_PASS unset, "
              "or Screener is refusing connections). Cached data will still be "
              "used; anything missing is simply absent from the tab.")

    # 1. Reverse index, built from the gainers themselves because Screener has
    #    no investor search. BOTH sources are read: the shareholder API lists
    #    only investors that still hold above the disclosure threshold, while
    #    the deals page lists everyone who has traded the stock, including
    #    those who have since sold out. One name can carry several person ids.
    index: dict[str, set[str]] = {}
    seen_names: dict[str, str] = {}
    canonical: dict[str, str] = {}      # person id -> the URL first seen for it
    resolved = 0
    for w in winners:
        cid = screener_company_id(session, w.company)
        if not cid:
            print(f"  [screener] no company match for {w.company!r}")
            continue
        resolved += 1
        for nm, url in (screener_people_for_company(session, cid)
                        + screener_people_in_trades(session, cid)):
            k = squash_investor_name(nm)
            if not k:
                continue
            url = canonical.setdefault(_person_id(url), url)
            index.setdefault(k, set()).add(url)
            seen_names.setdefault(k, nm)
    print(f"  [screener] resolved {resolved}/{len(winners)} companies, "
          f"{len(index)} distinct investor names indexed over "
          f"{len(canonical)} person pages.")

    # 2. Match our anchor names against it. Both sides are squashed to bare
    #    letters with legal suffixes peeled, which absorbs case, punctuation,
    #    'Pvt Ltd' against 'Private Limited' and the missing spaces OCR leaves
    #    behind. Beyond that, Screener's own records contain typos — LRSD is
    #    filed there once as 'Securities' and once as 'Securites' — so a near
    #    match is accepted when it is close and unambiguous.
    keys = list(index)

    def _lookup(name: str) -> "set[str]":
        k = squash_investor_name(name)
        if not k:
            return set()
        if k in index:
            return set(index[k])
        if len(k) < 8:
            return set()
        # One-sided prefix: Screener may carry a shorter or longer form.
        hits = {ik for ik in keys if ik.startswith(k) or k.startswith(ik)}
        if len(hits) != 1:
            # Spelling slips. Guarded hard: same opening, near-identical length,
            # very high similarity, and only one candidate.
            hits = {ik for ik in keys
                    if ik[:4] == k[:4] and abs(len(ik) - len(k)) <= 2
                    and difflib.SequenceMatcher(None, ik, k).ratio() >= 0.94}
        if len(hits) != 1:
            return set()
        return set(index[hits.pop()])

    #    Every filed spelling of the investor is tried and the results are
    #    UNIONED rather than stopping at the first hit, because different
    #    spellings resolve to different Screener records for the same firm.
    wanted: dict[str, set[str]] = {}
    for label, p in profiles.items():
        urls: set[str] = set()
        for variant in [label] + p["variants"]:
            urls |= _lookup(variant)
        if urls:
            wanted[label] = urls

    # 3. Groups that Screener proves are one investor, merged into one profile.
    remap = _split_investor_remap(profiles, wanted)
    for old, keep in remap.items():
        src, dst = profiles.pop(old), profiles[keep]
        dst["anchors"].extend(src["anchors"])
        dst["variants"] += [v for v in src["variants"] if v not in dst["variants"]]
        dst["ipos"] = len({r[0] for r in dst["anchors"]})
        wanted[keep] = wanted.get(keep, set()) | wanted.pop(old, set())
    if remap:
        print(f"  [screener] {len(remap)} truncated name(s) folded into the "
              f"investor they belong to, e.g. "
              f"{next(iter(remap))!r} -> {remap[next(iter(remap))]!r}.")

    # 4. One fetch per distinct person page, shared by every name pointing at it.
    by_url: dict[str, dict | None] = {}
    targets = sorted({u for s in wanted.values() for u in s})
    for i, url in enumerate(targets, 1):
        by_url[url] = screener_person_holdings(session, url)
        if i % 25 == 0:
            print(f"  [screener] fetched {i}/{len(targets)} investor pages ...")
    for label, urls in wanted.items():
        profiles[label]["screener"] = merge_person_pages(
            [by_url.get(u) for u in sorted(urls)])

    hit = sum(1 for p in profiles.values() if p["screener"])
    multi = sum(1 for p in profiles.values()
                if p["screener"] and len(p["screener"]["sources"]) > 1)
    print(f"  [screener] matched {hit}/{len(profiles)} anchor investors to a "
          f"Screener investor page ({multi} needed more than one record).")
    return profiles, remap


def _profiles_panel_html(profiles: "dict[str, dict]") -> str:
    """The Investor Profile tab: a picker plus a client-side rendered profile.

    The data is embedded as JSON so the page stays a single self-contained file
    like the other tabs — opening it never touches the network. `<` is escaped
    in the payload so no OCR'd name can break out of the script element."""
    if not profiles:
        return '<div class="empty">No anchor investors to profile.</div>'
    payload = json.dumps(profiles, ensure_ascii=False, default=str)
    payload = payload.replace("<", "\\u003c").replace("\u2028", "\\u2028")
    names = sorted(profiles, key=lambda n: (-profiles[n]["ipos"], n.lower()))
    opts = "".join(
        f'<option value="{_html.escape(n)}">{_html.escape(n)}'
        f' — {profiles[n]["ipos"]} IPO'
        f'{"s" if profiles[n]["ipos"] != 1 else ""}'
        f'{" · Screener" if profiles[n]["screener"] else ""}</option>'
        for n in names)
    return f"""<div class="picker">
  <label for="invFilter">Filter</label>
  <input id="invFilter" type="text" placeholder="type to narrow the list ..."
         oninput="filterInvestors()">
  <label for="invPick">Anchor investor</label>
  <select id="invPick" size="1" onchange="renderInvestor()">{opts}</select>
</div>
<div id="invBody"></div>
<script id="invData" type="application/json">{payload}</script>
<script>
const INV = JSON.parse(document.getElementById('invData').textContent);
const ALL_NAMES = Array.from(document.getElementById('invPick').options)
                       .map(o => ({{v: o.value, t: o.textContent}}));

function esc(s) {{
  return String(s).replace(/[&<>"']/g, c => (
    {{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[c]));
}}
function fmt(v, d) {{
  if (v === null || v === undefined || v === '') return '';
  const n = Number(v);
  if (!isFinite(n)) return esc(v);
  return n.toLocaleString('en-IN', {{minimumFractionDigits: d, maximumFractionDigits: d}});
}}
function filterInvestors() {{
  const q = document.getElementById('invFilter').value.toLowerCase();
  const sel = document.getElementById('invPick');
  const keep = ALL_NAMES.filter(o => o.t.toLowerCase().includes(q));
  sel.innerHTML = keep.map(o =>
    `<option value="${{esc(o.v)}}">${{esc(o.t)}}</option>`).join('');
  sel.size = 1;
  renderInvestor();
}}
function renderInvestor() {{
  const sel = document.getElementById('invPick');
  const box = document.getElementById('invBody');
  const name = sel.value;
  const p = INV[name];
  if (!p) {{ box.innerHTML = '<div class="empty">No investor selected.</div>'; return; }}

  let amt = 0, sh = 0;
  p.anchors.forEach(a => {{ amt += Number(a[1]) || 0; sh += Number(a[2]) || 0; }});
  const cards = `<div class="cards">
    <div class="card"><div class="k">Anchor IPOs</div><div class="v">${{p.ipos}}</div></div>
    <div class="card"><div class="k">Allocations</div><div class="v">${{p.anchors.length}}</div></div>
    <div class="card"><div class="k">Total anchor &#8377;</div><div class="v">${{amt ? fmt(amt, 0) : '&ndash;'}}</div></div>
    <div class="card"><div class="k">Shares allotted</div><div class="v">${{sh ? fmt(sh, 0) : '&ndash;'}}</div></div>
    <div class="card"><div class="k">Screener holdings</div><div class="v">${{p.screener ? p.screener.rows.length : '&ndash;'}}</div></div>
    <div class="card"><div class="k">Screener deals</div><div class="v">${{p.screener ? p.screener.trades.length : '&ndash;'}}</div></div>
  </div>`;

  const alias = (p.variants && p.variants.length > 1)
    ? `<div class="note">Filings spell this investor ${{p.variants.length}} ` +
      `different ways; all are counted here as one. Spellings seen: ` +
      p.variants.map(v => `<b>${{esc(v)}}</b>`).join(' &nbsp;&middot;&nbsp; ') +
      `</div>` : '';

  let anch = '<h3 class="sec">Anchor investments<span class="hint">from the ' +
             'issuers&rsquo; anchor allocation filings</span></h3>' +
    '<table><thead><tr><th>Company</th><th class="num">Amount Invested &#8377;</th>' +
    '<th class="num">Shares Allotted</th><th class="num">Price &#8377;</th>' +
    '<th>Name as filed</th>' +
    '</tr></thead><tbody>';
  p.anchors.slice().sort((a, b) => (Number(b[1]) || 0) - (Number(a[1]) || 0))
   .forEach(a => {{
    anch += `<tr><td>${{esc(a[0])}}</td><td class="num">${{fmt(a[1], 0)}}</td>` +
            `<td class="num">${{fmt(a[2], 0)}}</td><td class="num">${{fmt(a[3], 2)}}</td>` +
            `<td>${{esc(a[4] || '')}}</td></tr>`;
  }});
  anch += '</tbody></table>';

  let hold = '';
  if (p.screener) {{
    const s = p.screener;
    const srcs = s.sources || [[s.name, s.url]];
    const link = srcs.map(x => `<a class="ext" target="_blank" ` +
      `href="https://www.screener.in${{esc(x[1])}}">${{esc(x[0] || 'Screener')}} &#8599;</a>`).join(' &middot; ');
    const merged = srcs.length > 1
      ? `<div class="note">Screener holds ${{srcs.length}} separate records for ` +
        `this investor, one per spelling it was first filed under. They are ` +
        `combined below; the links go to each original.</div>`
      : '';
    if (s.rows && s.rows.length) {{
      hold += `<h3 class="sec">Holdings<span class="hint">` +
        `shareholding %, every quarter Screener reports &mdash; a blank cell ` +
        `means not a named holder that quarter &nbsp;${{link}}</span></h3>` +
        merged + '<div class="wide"><table class="frz"><thead><tr><th>Company</th>' +
        s.periods.map(q => `<th class="num">${{esc(q)}}</th>`).join('') +
        '</tr></thead><tbody>';
      s.rows.forEach(r => {{
        hold += `<tr><td>${{esc(r[0])}}</td>`;
        let prev = null;
        for (let i = 1; i <= s.periods.length; i++) {{
          const raw = r[i];
          const n = (raw === undefined || raw === '') ? null : Number(raw);
          let cls = 'num';
          if (n !== null && prev !== null) cls += (n > prev) ? ' up' : (n < prev ? ' dn' : '');
          hold += `<td class="${{cls}}">${{n === null ? '' : fmt(n, 2)}}</td>`;
          if (n !== null) prev = n;
        }}
        hold += '</tr>';
      }});
      hold += '</tbody></table></div>';
    }}
    if (s.trades && s.trades.length) {{
      hold += `<h3 class="sec">Trading history<span class="hint">` +
        `every bulk, block and insider deal Screener records for this ` +
        `investor, newest first` +
        `${{s.rows && s.rows.length ? '' : ' &nbsp;' + link}}</span></h3>` +
        ((s.rows && s.rows.length) ? '' : merged) +
        '<div class="wide"><table class="frz"><thead><tr><th>Company</th><th>Date</th>' +
        '<th>Action</th><th class="num">Quantity</th>' +
        '<th class="num">Price &#8377;</th>' +
        '<th>Disclosed as</th></tr></thead><tbody>';
      s.trades.forEach(t => {{
        const buy = /^b/i.test(t[2] || '');
        hold += `<tr><td>${{esc(t[0])}}</td><td>${{esc(t[1])}}</td>` +
          `<td class="${{buy ? 'up' : 'dn'}}">${{esc(t[2])}}</td>` +
          `<td class="num">${{esc(t[3])}}</td><td class="num">${{esc(t[4])}}</td>` +
          `<td>${{esc(t[5] || '')}}</td></tr>`;
      }});
      hold += '</tbody></table></div>';
    }}
    if (!hold) {{
      hold = '<h3 class="sec">Holdings &amp; trading history</h3>' +
        `<div class="note">Screener has a page for this investor but lists ` +
        `neither holdings nor deals on it. ${{link}}</div>`;
    }}
  }} else {{
    hold = '<h3 class="sec">Holdings &amp; trading history</h3>' +
      '<div class="note">No Screener investor page matched this name. ' +
      'Screener has no investor search, so an investor is only reachable when ' +
      'it is named as a shareholder of, or recorded dealing in, one of the ' +
      'companies on this list. One that has neither held above the disclosure ' +
      'threshold nor done a reported bulk, block or insider deal in any of ' +
      'them cannot be located at all.</div>';
  }}
  box.innerHTML = cards + alias + anch + hold;
}}
renderInvestor();
</script>"""


# ───────────────────────────── HTML output ─────────────────────────────────
# Same tab shell and palette as market_charts.html, so the two dashboards read
# as one family.
HTML_SHELL = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>__TITLE__</title>
<style>
  * { margin:0; padding:0; box-sizing:border-box; }
  body { font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
         background:#f5f5f5; color:#1f2937; }
  .tab-bar { display:flex; gap:4px; padding:12px 16px;
             background:linear-gradient(135deg,#1F4E79,#2E75B6);
             position:sticky; top:0; z-index:1000;
             box-shadow:0 2px 8px rgba(0,0,0,0.3); }
  .tab-btn { padding:10px 22px; border:none; border-radius:6px 6px 0 0;
             cursor:pointer; font-size:14px; font-weight:600; color:#b0c4de;
             background:rgba(255,255,255,0.1); transition:all 0.2s; }
  .tab-btn:hover { color:#fff; background:rgba(255,255,255,0.2); }
  .tab-btn.active { color:#fff; background:#e94560;
                    box-shadow:0 -2px 6px rgba(233,69,96,0.4); }
  .tab-panel { width:100%; height:calc(100vh - 60px); overflow:auto;
               padding:16px; }
  .meta { font-size:12px; color:#5b6b7c; margin-bottom:10px; }
  table { border-collapse:collapse; width:100%; background:#fff;
          box-shadow:0 1px 4px rgba(0,0,0,0.12); font-size:13px; }
  thead th { position:sticky; top:0; background:#1F4E79; color:#fff;
             font-weight:600; text-align:left; padding:9px 12px;
             white-space:nowrap; z-index:2; }
  thead th.num { text-align:right; }
  table.dt thead th { cursor:pointer; user-select:none; }
  table.dt thead th:hover { background:#2E75B6; }
  table.dt thead th.sorted[data-dir="asc"]::after { content:' \\25B2'; }
  table.dt thead th.sorted[data-dir="desc"]::after { content:' \\25BC'; }
  tbody td { padding:7px 12px; border-bottom:1px solid #e6e9ee;
             white-space:nowrap; }
  tbody td.wrap { white-space:normal; max-width:520px; overflow-wrap:anywhere;
                  vertical-align:top; }
  tbody tr:nth-child(even) { background:#fafbfd; }
  table.dt tbody tr:nth-child(even) { background:#fff; }
  table.dt tbody tr.alt { background:#fafbfd; }
  tbody tr:hover { background:#eaf2fb; }
  td.num { text-align:right; font-variant-numeric:tabular-nums; }
  tbody tr.grp td { border-top:2px solid #c7d4e4; }
  tbody tr.grp td:first-child { font-weight:600; color:#1F4E79; }
  td.rep { color:transparent; }
  .empty { padding:24px; color:#6b7280; font-style:italic; }
  .picker { display:flex; align-items:center; gap:10px; flex-wrap:wrap;
            background:#fff; padding:12px 14px; border-radius:8px;
            box-shadow:0 1px 4px rgba(0,0,0,0.12); margin-bottom:14px; }
  .picker label { font-size:13px; font-weight:600; color:#1F4E79; }
  .picker select, .picker input { font-size:13px; padding:7px 10px;
            border:1px solid #cbd5e1; border-radius:6px; background:#fff;
            color:#1f2937; min-width:260px; }
  .picker select { max-width:520px; }
  .picker .count { font-size:12px; color:#6b7280;
                   font-variant-numeric:tabular-nums; }
  .cards { display:flex; gap:10px; flex-wrap:wrap; margin-bottom:14px; }
  .card { background:#fff; border-radius:8px; padding:10px 16px;
          box-shadow:0 1px 4px rgba(0,0,0,0.12); min-width:150px; }
  .card .k { font-size:11px; text-transform:uppercase; letter-spacing:.04em;
             color:#6b7280; }
  .card .v { font-size:19px; font-weight:700; color:#1F4E79;
             font-variant-numeric:tabular-nums; }
  h3.sec { font-size:14px; font-weight:700; color:#1F4E79;
           margin:18px 0 8px; padding-bottom:5px;
           border-bottom:2px solid #e94560; }
  h3.sec .hint { font-weight:400; color:#6b7280; font-size:12px;
                 margin-left:8px; }
  .note { background:#fff8e1; border-left:3px solid #f0ad4e; padding:9px 12px;
          font-size:12px; color:#6b5a2b; border-radius:0 6px 6px 0;
          margin-bottom:12px; }
  td.up { color:#137333; } td.dn { color:#b3261e; }
  a.ext { color:#2E75B6; text-decoration:none; font-size:12px; }
  a.ext:hover { text-decoration:underline; }
  /* A long-standing holder can span 40+ quarters, far wider than the page, and
     a prolific one hundreds of deals. The box scrolls in BOTH directions so the
     sticky header and first column have something to stick to — sticky resolves
     against the nearest scrolling ancestor, and a box that only scrolls
     sideways gives the header nothing to hold. */
  .wide { overflow:auto; max-width:100%; max-height:70vh; }
  /* A whole-tab table gets the panel's height, less the tab bar, the caption
     and the filter box; a profile table is one section of many, so it keeps
     the 70vh above. */
  .wide.tall { max-height:calc(100vh - 205px); }
  table.frz thead th:first-child,
  table.frz tbody td:first-child { position:sticky; left:0;
                                   box-shadow:1px 0 0 #d7dee8; }
  /* Sticky cells slide over the ones behind them, so they need their own
     opaque background — the row's does not travel with them. */
  table.frz tbody td:first-child { background:#fff; z-index:1; }
  table.frz tbody tr:nth-child(even) td:first-child { background:#fafbfd; }
  table.frz tbody tr:hover td:first-child { background:#eaf2fb; }
  table.frz thead th:first-child { z-index:3; }
  /* table.dt stripes per issuer block with .alt, not nth-child, so its frozen
     column needs its own backgrounds to stay in step with the row. The extra
     `tr` is load-bearing: it outranks the nth-child rule above, which counts
     rows a filter has hidden and so drifts out of step with .alt. */
  table.dt.frz tbody tr td:first-child { background:#fff; }
  table.dt.frz tbody tr.alt td:first-child { background:#fafbfd; }
  table.dt.frz tbody tr:hover td:first-child { background:#eaf2fb; }
</style>
</head>
<body>
<div class="tab-bar">__BUTTONS__</div>
__PANELS__
<script>
function showTab(idx) {
  document.querySelectorAll('.tab-btn').forEach((b, i) => b.classList.toggle('active', i === idx));
  document.querySelectorAll('.tab-panel').forEach((p, i) => p.style.display = (i === idx ? 'block' : 'none'));
}

// A collapsed cell is blanked for READING only; its value lives in data-v so
// that sorting and filtering still see the whole column.
function cellVal(td) {
  return td.dataset.v !== undefined ? td.dataset.v : td.textContent;
}

// Re-blank the repeated issuer names and re-stripe, over VISIBLE rows only,
// after any sort or filter has changed which rows sit next to each other.
function redraw(tbl) {
  const ci = tbl.dataset.collapse === '' ? -1 : Number(tbl.dataset.collapse);
  let prev = null, n = 0;
  for (const r of tbl.tBodies[0].rows) {
    if (r.style.display === 'none') continue;
    r.classList.toggle('alt', n % 2 === 1);
    n++;
    if (ci < 0) continue;
    const td = r.cells[ci], v = cellVal(td), head = v !== prev;
    prev = v;
    r.classList.toggle('grp', head);
    td.classList.toggle('rep', !head);
    td.textContent = head ? v : '';
  }
}

function sortTable(th) {
  const tbl = th.closest('table'), tb = tbl.tBodies[0], i = Number(th.dataset.i);
  const dir = th.dataset.dir === 'asc' ? -1 : 1;
  for (const c of tbl.tHead.rows[0].cells) {
    c.classList.remove('sorted');
    delete c.dataset.dir;
  }
  th.classList.add('sorted');
  th.dataset.dir = dir === 1 ? 'asc' : 'desc';
  const rows = Array.from(tb.rows);
  const num = rows.every(r => r.cells[i].classList.contains('num')
                              || !cellVal(r.cells[i]).trim());
  rows.sort((a, b) => {
    const x = cellVal(a.cells[i]).trim(), y = cellVal(b.cells[i]).trim();
    // Blanks always sink, in both directions: an empty cell is missing data,
    // not the smallest value, and floating it to the top buries the answer.
    if (!x || !y) return !x && !y ? 0 : (!x ? 1 : -1);
    if (num) return dir * (parseFloat(x.replace(/,/g, ''))
                           - parseFloat(y.replace(/,/g, '')));
    return dir * x.localeCompare(y, undefined, {numeric: true,
                                                sensitivity: 'base'});
  });
  rows.forEach(r => tb.appendChild(r));
  redraw(tbl);
}

function filterTable(inp) {
  const tbl = document.getElementById(inp.dataset.for);
  const cols = inp.dataset.cols.split(',').map(Number);
  const q = inp.value.toLowerCase().trim();
  let shown = 0;
  for (const r of tbl.tBodies[0].rows) {
    const hit = !q || cols.some(i => cellVal(r.cells[i]).toLowerCase().includes(q));
    r.style.display = hit ? '' : 'none';
    if (hit) shown++;
  }
  const c = document.getElementById(inp.id + 'Count');
  if (c) c.textContent = shown + (shown === 1 ? ' row' : ' rows');
  redraw(tbl);
}

document.querySelectorAll('table.dt').forEach(redraw);
</script>
</body>
</html>
"""


def _df_to_html(df: pd.DataFrame, collapse_col: str | None = None,
                table_id: str = "", freeze: bool = False) -> str:
    """Render a DataFrame as a plain HTML table.

    Written by hand rather than via DataFrame.to_html because that renders NaN
    as the literal string 'nan' and right-alignment cannot be driven per cell.
    Every value is escaped: anchor names come out of OCR'd third-party PDFs and
    are never trusted as markup.

    `collapse_col` blanks a repeated value so it is printed once per run rather
    than on every row — the issuer's name against each of its fifty anchors is
    noise, and printing it once makes the groups legible. The blanking is
    display-only; the workbook keeps the value on every row so the sheet stays
    sortable and filterable."""
    if df is None or df.empty:
        return '<div class="empty">Nothing to show.</div>'
    head = "".join(
        f'<th data-i="{j}" onclick="sortTable(this)"'
        + (' class="num"' if pd.api.types.is_numeric_dtype(df[c]) else "")
        + f">{_html.escape(str(c))}</th>"
        for j, c in enumerate(df.columns))
    ci = list(df.columns).index(collapse_col) if collapse_col in df.columns else -1
    # A long text column left on one line pushes every column after it off the
    # screen, so it is wrapped inside its own width instead. Note is exempt:
    # it is the last column and wrapping it collapses it to one char per line.
    wide = {j for j, c in enumerate(df.columns)
            if not pd.api.types.is_numeric_dtype(df[c])
            and str(c).strip().lower() != "note"
            and df[c].map(lambda v: len(str(v))).max() > 60}
    body = []
    prev = object()
    for row in df.itertuples(index=False):
        cells = []
        new_group = ci >= 0 and row[ci] != prev
        if ci >= 0:
            prev = row[ci]
        for j, v in enumerate(row):
            if j == ci:
                # The value is carried in data-v even when blanked, so that
                # sorting and filtering see the whole column.
                esc = _html.escape(str(v))
                cells.append(f'<td data-v="{esc}">{esc}</td>' if new_group
                             else f'<td class="rep" data-v="{esc}"></td>')
            elif v is None or (not isinstance(v, str) and pd.isna(v)):
                cells.append("<td></td>")
            elif pd.api.types.is_number(v) and not isinstance(v, bool):
                # A nullable-integer column yields numpy ints, which are not
                # Python ints, so a plain isinstance check printed counts as
                # bare text and "101.00" for what is a whole number of IPOs.
                txt = f"{v:,.2f}" if isinstance(v, float) else f"{v:,}"
                cells.append(f'<td class="num">{txt}</td>')
            else:
                td = ' class="wrap"' if j in wide else ""
                cells.append(f"<td{td}>{_html.escape(str(v))}</td>")
        tr = '<tr class="grp">' if new_group and ci >= 0 else "<tr>"
        body.append(tr + "".join(cells) + "</tr>")
    # A leading collapse column is the row's label, so freeze it against a
    # sideways scroll the way the profile tables freeze theirs.
    frozen = freeze or ci == 0
    cls = "dt frz" if frozen else "dt"
    tbl = (f'<table class="{cls}" id="{table_id}" data-collapse="'
           + (str(ci) if ci >= 0 else "")
           + '"><thead><tr>' + head + "</tr></thead><tbody>"
           + "".join(body) + "</tbody></table>")
    # Sticky resolves against the nearest scrolling ancestor. Left to the tab
    # panel, whose padding the rows scroll underneath, a frozen cell pins one
    # padding-width in and the scrolled row shows through the gutter beside it.
    # Its own box has no padding, so the freeze lands flush.
    return f'<div class="wide tall">{tbl}</div>' if frozen else tbl


def _filter_box(box_id: str, table_id: str, df: pd.DataFrame,
                cols: "list[str]") -> str:
    """A text box that narrows `table_id` on the named columns."""
    idx = [str(list(df.columns).index(c)) for c in cols if c in df.columns]
    if not idx:
        return ""
    where = " or ".join(cols)
    return (f'<div class="picker"><label for="{box_id}">Filter</label>'
            f'<input id="{box_id}" type="text" data-for="{table_id}" '
            f'data-cols="{",".join(idx)}" oninput="filterTable(this)" '
            f'placeholder="type to narrow by {_html.escape(where)} ...">'
            f'<span class="count" id="{box_id}Count">{len(df):,} rows</span>'
            "</div>")


def write_html(tabs: "list[tuple]", out_html: Path) -> None:
    """Write the tabbed dashboard.

    Each tab is (label, sub-caption, body). The body is either a DataFrame —
    rendered as a table — or a string of pre-built HTML, which is how the
    investor-profile tab gets in. A 4th element, when present, names the column
    whose repeated values should be collapsed, a 5th the columns the tab's
    filter box should search, and a 6th freezes the first column even when the
    tab has no collapse column to imply it."""
    buttons, panels = [], []
    for i, tab in enumerate(tabs):
        label, note, body = tab[0], tab[1], tab[2]
        collapse = tab[3] if len(tab) > 3 else None
        filter_cols = tab[4] if len(tab) > 4 else None
        freeze = bool(tab[5]) if len(tab) > 5 else False
        cls = "tab-btn active" if i == 0 else "tab-btn"
        buttons.append(f'<button class="{cls}" onclick="showTab({i})">'
                       f'{_html.escape(label)}</button>')
        disp = "block" if i == 0 else "none"
        if isinstance(body, str):
            inner = body
        else:
            tid = f"tbl-{i}"
            inner = _df_to_html(body, collapse, tid, freeze)
            if filter_cols and body is not None and not body.empty:
                inner = _filter_box(f"flt{i}", tid, body, filter_cols) + inner
        panels.append(f'<div class="tab-panel" id="panel-{i}" style="display:{disp};">'
                      f'<div class="meta">{_html.escape(note)}</div>'
                      f'{inner}</div>')
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    page = (HTML_SHELL
            .replace("__TITLE__", f"IPO Listing Gainers — {stamp}")
            .replace("__BUTTONS__", "".join(buttons))
            .replace("__PANELS__", "\n".join(panels)))
    out_html.parent.mkdir(parents=True, exist_ok=True)
    out_html.write_text(page, encoding="utf-8")


# ─────────────────────────────── main ──────────────────────────────────────
def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--start", default="2025-01-01",
                    help="Window start (listing date >= this). Default 2025-01-01.")
    ap.add_argument("--end", default=date.today().isoformat(),
                    help="Window end (listing date <= this). Default today.")
    ap.add_argument("--threshold", type=float, default=50.0,
                    help="Min %% gain to qualify (listing-day OR 30d peak). Default 50.")
    ap.add_argument("--window-days", type=int, default=30,
                    help="'Within a month' window in calendar days. Default 30.")
    ap.add_argument("--workers", type=int, default=5,
                    help="Concurrent price fetches. Default 5.")
    ap.add_argument("--limit", type=int, default=0, help="Debug: first N IPOs only.")
    ap.add_argument("--bse-sme", dest="bse_sme",
                    action=argparse.BooleanOptionalAction, default=True,
                    help="Fold in dated BSE-SME-only back-history, which neither "
                         "exchange feed can supply. On by default.")
    ap.add_argument("--bse-recent", dest="bse_recent",
                    action=argparse.BooleanOptionalAction, default=True,
                    help="Fold in BSE's live public-issue feed, which catches an "
                         "issue that has listed too recently to be indexed "
                         "anywhere else. On by default; --no-bse-recent turns it "
                         "off. BSE publishes no dated history, so each run "
                         "snapshots the currently-open window into a ledger.")
    ap.add_argument("--anchor-dir", default=str(ANCHOR_DIR_DEFAULT),
                    help="Folder holding the 'Allocation to Anchor Investors' "
                         "filings (PDF or image, one per symbol), named with the "
                         f"symbol or company. Default {ANCHOR_DIR_DEFAULT}.")
    ap.add_argument("--name-key", choices=("smart", "words"), default="smart",
                    help="How to decide that two anchor name cells are the same "
                         "investor. 'smart' (default) canonicalises the name to "
                         "its significant tokens and merges a name into a fuller "
                         "one; 'words' is the legacy first-N-words key. Default "
                         "smart.")
    ap.add_argument("--freq-words", type=int, default=2,
                    help="How many leading words of an investor name to group on. "
                         "Only used with --name-key words. Default 2.")
    ap.add_argument("--no-ocr", action="store_true",
                    help="Disable the OCR fallback for scanned anchor filings.")
    ap.add_argument("--no-ipoplatform", action="store_true",
                    help="Do not use ipoplatform.com's published anchor books, "
                         "which otherwise replace the OCR reading for SME issues. "
                         "Falls back to reading the scanned filings, which is "
                         "markedly less accurate.")
    ap.add_argument("--no-fetch", action="store_true",
                    help="Do not download anchor filings from NSE; use only what "
                         "is already in --anchor-dir.")
    ap.add_argument("--rebuild-anchors", action="store_true",
                    help="Replace, rather than merge with, the anchor rows already "
                         "in the workbook for every symbol re-read from its filing. "
                         "Use after a parser fix, when the stored names are the "
                         "ones to be corrected. Symbols with no filing are kept.")
    ap.add_argument("--no-screener", action="store_true",
                    help="Skip the Screener.in lookup. The Investor Profile tab "
                         "then shows anchor allocations only, with no holdings "
                         "or trading history.")
    ap.add_argument("--xlsx", default=str(OUTPUT_DIR / "ipo_listing_gainers.xlsx"),
                    help="Four-sheet Excel output (All IPOs + Gainers + Anchor "
                         "Investors + Investor Frequency).")
    ap.add_argument("--html", default=str(OUTPUT_DIR / "ipo_listing_gainers.html"),
                    help="Tabbed HTML dashboard: the four sheets plus a "
                         "per-investor profile tab.")
    args = ap.parse_args()

    start = parse_date(args.start)
    end = parse_date(args.end)
    if start is None or end is None:
        print("Bad --start/--end date.", file=sys.stderr)
        return 2

    print(f"[1/3] Fetching IPO master (listings {start} .. {end}) ...")
    ipos = collect_nse_ipos(start, end)
    n_nse = len(ipos)

    def _norm_name(c: str) -> str:
        return re.sub(r"\s+", " ", c.lower()).replace(" limited", "").strip()

    have = {_norm_name(i.company) for i in ipos}
    print(f"      NSE feed: {n_nse} IPOs")
    if args.bse_sme:
        bsme = collect_bse_sme(start, end, have)
        ipos += bsme
        have |= {_norm_name(i.company) for i in bsme}
        print(f"      BSE SME back-history: {len(bsme)}")
    else:
        bsme = []
    if args.bse_recent:
        # Anything the dated BSE-SME source already covers must not arrive twice
        # under its scrip code.
        seen_codes = {i.quote_symbol for i in bsme if i.quote_symbol}
        bse = [i for i in collect_bse_recent(start, end, have)
               if not i.quote_symbol or i.quote_symbol not in seen_codes]
        ipos += bse
        print(f"      BSE live supplement: {len(bse)}")

    ipos.sort(key=lambda x: x.listing_date, reverse=True)
    if args.limit:
        ipos = ipos[:args.limit]

    print(f"[2/3] Pricing {len(ipos)} IPOs (workers={args.workers}) ...")
    done = 0
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = {ex.submit(compute_returns, ipo, args.window_days): ipo for ipo in ipos}
        for f in as_completed(futs):
            done += 1
            if done % 25 == 0 or done == len(ipos):
                print(f"      {done}/{len(ipos)}")

    thr = args.threshold

    def _peaked_on_listing_day(i: IPO) -> bool:
        """Its best close in the whole window was day one — it never traded
        higher afterwards. A pop, not growth, so it is not a gainer."""
        return i.peak_date is not None and i.peak_date == i.listing_date

    winners = [i for i in ipos
               if ((i.listing_pct is not None and i.listing_pct >= thr)
                   or (i.peak_pct is not None and i.peak_pct >= thr))
               and not _peaked_on_listing_day(i)]
    winners.sort(key=lambda i: max(i.listing_pct or -1e9, i.peak_pct or -1e9),
                 reverse=True)

    no_data = [i for i in ipos if i.listing_close is None]

    # ── Sheet 1: every IPO in the window; Sheet 2: the ones that qualified ──
    def _row(i: IPO) -> dict:
        reason = []
        if i.listing_pct is not None and i.listing_pct >= thr:
            reason.append("listing-day")
        if i.peak_pct is not None and i.peak_pct >= thr:
            reason.append("30d-peak")
        return {
            "Symbol": i.symbol,
            "Exchange": "BSE" if i.segment.startswith("BSE") else "NSE",
            "Company": i.company,
            "Segment": i.segment,
            "Listing Date": i.listing_date.isoformat(),
            "Issue ₹": round(i.issue_price, 2),
            "Listing Close ₹": round(i.listing_close, 2) if i.listing_close else None,
            "Listing-day %": round(i.listing_pct, 1) if i.listing_pct is not None else None,
            "30d Peak ₹": round(i.peak_close, 2) if i.peak_close else None,
            "30d Peak Date": i.peak_date.isoformat() if i.peak_date else None,
            "30d Peak %": round(i.peak_pct, 1) if i.peak_pct is not None else None,
            "Qualifies By": " + ".join(reason),
        }

    win_ids = {id(i) for i in winners}
    all_rows = []
    for i in sorted(ipos, key=lambda x: x.listing_date, reverse=True):
        r = _row(i)
        r["Qualified"] = "yes" if id(i) in win_ids else "no"
        # Why a row has no return: no price history, or a rights/FPO/re-listing
        # the pricing step deliberately refuses to score.
        note = i.note
        if not note and r["Qualified"] == "no" and r["Qualifies By"]:
            note = "peaked on listing day — no follow-through, excluded"
        r["Note"] = note
        all_rows.append(r)
    all_df = pd.DataFrame(all_rows)
    df = pd.DataFrame([_row(i) for i in winners])

    # ── Sheet 2: anchor investors ────────────────────────────────────────
    anchor_dir = Path(args.anchor_dir)
    if not args.no_fetch:
        sync_anchor_filings(anchor_dir, winners)

    print(f"\n[anchors] Reading anchor filings from {anchor_dir} ...")
    xlsx_path = Path(args.xlsx)
    existing = read_existing_anchors(xlsx_path)
    # --rebuild-anchors exists to force a re-read; honouring the stored
    # readings would make it a no-op for exactly the filings it targets.
    global _FROZEN_DISABLED
    _FROZEN_DISABLED = bool(args.rebuild_anchors)
    anchor_map, unmatched, _, verified = collect_anchor_investors(
        anchor_dir, winners, use_ocr=not args.no_ocr,
        use_ipoplatform=not args.no_ipoplatform,
        extra_keys=tuple(existing), universe=ipos)

    matched = sum(1 for w in winners if anchor_map.get(w.symbol))
    total_names = sum(len(v) for v in anchor_map.values())
    missing = [w.symbol for w in winners if not anchor_map.get(w.symbol)]
    print(f"          TOTAL: {matched}/{len(winners)} symbols with anchors "
          f"({total_names} names extracted).")
    if unmatched:
        print(f"          {len(unmatched)} file(s) could not be matched to a symbol: "
              f"{', '.join(unmatched)}")
    if missing:
        print(f"          {len(missing)} symbol(s) have NO anchor data in this run "
              "(any names already in the workbook are preserved):")
        by_sym = {w.symbol: w.company for w in winners}
        for sym in missing:
            print(f"            - {sym}  ({by_sym.get(sym, '')})")
        print(f"          Drop each one's 'Allocation to Anchor Investors' filing "
              f"into {anchor_dir}\n          (name the file with the symbol or "
              "company, e.g. MEESHO.pdf), then re-run.")

    if existing:
        print(f"  [anchor] merging with {sum(len(v) for v in existing.values())} "
              f"row(s) already in the workbook.")
    if verified:
        print(f"  [anchor] {len(verified)} filing(s) read to a table that "
              f"reconciles exactly; their stored rows are being replaced "
              f"rather than merged: {', '.join(sorted(verified))}")
    merged = build_merged_anchors(winners, anchor_map, existing,
                                  rebuild=args.rebuild_anchors, universe=ipos,
                                  verified=verified)
    anchors_df = anchors_to_frame(merged, winners, universe=ipos)
    smart_key = args.name_key == "smart"

    # Profiles first: the Screener stage can prove that two groups are one
    # investor, and the frequency sheet has to be built with that knowledge or
    # the two views disagree on how many investors there are.
    html_path = Path(args.html)
    if args.no_screener:
        profiles = build_investor_profiles_offline(merged, winners,
                                                   words=args.freq_words,
                                                   smart=smart_key)
        remap: dict[str, str] = {}
    else:
        print("\n[screener] building investor profiles ...")
        profiles, remap = build_investor_profiles(merged, winners,
                                                  words=args.freq_words,
                                                  smart=smart_key)

    # The frequency sheet needs three things main already knows: which issues
    # qualified, each key's company name (its identity across the two sources),
    # and how many issues each investor anchored in total.
    qualified: dict[str, bool] = {}
    company_of: dict[str, str] = {}
    for i in ipos:
        won = id(i) in win_ids
        for key in (i.symbol, i.company):
            qualified[key] = won
            company_of[key] = i.company

    _lab = investor_labeller(merged, args.freq_words, smart_key)

    def _folded(name: str) -> str:
        k = _lab(name)
        return remap.get(k, k) if k else ""

    participation: dict[str, set] = {}
    if not args.no_ipoplatform:
        print("\n[ipoplatform] counting total anchor participations "
              f"{start} .. {end} ...")
        participation = ipl_participation(merged, _folded,
                                          start.isoformat(), end.isoformat())

    freq_df = build_investor_frequency(merged, words=args.freq_words,
                                       smart=smart_key, remap=remap,
                                       qualified=qualified,
                                       company_of=company_of,
                                       participation=participation or None)

    tracked_df = None
    if not args.no_ipoplatform:
        print(f"\n[ipoplatform] tracked watchlist participation {start} .. {end} ...")
        tracked_df = build_tracked_investors(ipl_anchor_scan(ipos),
                                             {i.symbol for i in winners})

    write_workbook(all_df, df, merged, anchors_df, freq_df, xlsx_path,
                   tracked_df=tracked_df)

    key_note = (("Spelling and suffix variants of one investor share a row; "
                 "different schemes of the same house do not."
                 if smart_key else
                 f"Grouped on the first {args.freq_words} word(s) of each "
                 "investor name.")
                + f" 'Total IPOs' is every issue the investor anchored "
                  f"{start} .. {end} per IPOPlatform, whatever it did next; "
                  "'Qualified IPOs' is how many of those cleared "
                  f"{thr:.0f}%, and 'Hit Rate %' is the ratio. A blank total "
                  "means IPOPlatform does not list that investor, so the rate "
                  "is unknown rather than perfect.")

    tabs = [
        ("All IPOs",
         f"Every IPO listed {start} .. {end} that the feeds report, NSE and "
         f"BSE SME alike — {len(all_df)} in all, of which {len(winners)} "
         f"cleared {thr:.0f}%. Same two returns as the Gainers tab; "
         "'Qualified' says whether it made the cut, and 'Note' says why a row "
         "has no return or why it was excluded.",
         all_df, None, ["Symbol", "Company", "Exchange", "Segment"], True),
        ("Gainers",
         f"IPOs listed {start} .. {end} that gained >= {thr:.0f}% on listing day "
         f"or within {args.window_days} days, EXCLUDING any whose best close was "
         "listing day itself — those never traded higher afterwards, so the "
         "gain is a pop, not growth.",
         df, None, ["Symbol", "Company", "Exchange", "Segment"]),
        ("Anchor Investors",
         f"{len(anchors_df)} anchor allocations across "
         f"{anchors_df[ANCHOR_COLUMNS[0]].nunique() if not anchors_df.empty else 0} "
         "issues. Amount = shares x allocation price. The issuer is printed "
         "once per block.", anchors_df, ANCHOR_COLUMNS[0], [ANCHOR_COLUMNS[0]]),
        ("Investor Frequency", key_note, freq_df, None, ["Investor", "Symbols"],
         True),
    ]
    if tracked_df is not None:
        tabs.append((
            "Tracked Investors",
            f"Every name on the watchlist and how many IPOs it anchored "
            f"{start} .. {end}, read from the anchor book of EVERY issue in the "
            "window rather than from our own filings, which exist only for "
            f"issues that cleared {thr:.0f}%. 'Total IPOs' therefore counts "
            "participation whatever the price did next; 'Qualified IPOs' is how "
            "many of those cleared the bar. A zero means the investor anchored "
            "nothing in the window, not that it is unknown.",
            tracked_df, None, ["Investor", "Symbols", "Qualified Symbols"], True))
    tabs.append(
        ("Investor Profile",
         f"{len(profiles)} anchor investors, grouped exactly as the Investor "
         "Frequency tab groups them, so the two agree. Pick one for its anchor "
         "allocations and, where Screener names it, its holdings across every "
         "company quarter by quarter.",
         _profiles_panel_html(profiles)),
    )
    write_html(tabs, html_path)

    print(f"\n[3/3] {len(winners)} of {len(all_df)} IPOs gained >= {thr:.0f}% "
          f"(listing-day OR within {args.window_days} days).")
    if no_data:
        print(f"      ({len(no_data)} IPOs had no usable price data and were skipped)")
    print(f"      Excel -> {xlsx_path}  (sheets: {', '.join(MANAGED_SHEETS)})")
    print(f"      HTML  -> {html_path}\n")
    if freq_df is not None and not freq_df.empty:
        top = freq_df[freq_df["Qualified IPOs"] >= 2]
        rated = freq_df["Hit Rate %"].notna().sum()
        print(f"      Investor Frequency: {len(freq_df)} unique investors "
              f"({args.name_key} name key), {len(top)} cleared the threshold in "
              f">=2 IPOs, {rated} have a hit rate.")
    if not df.empty:
        with pd.option_context("display.max_rows", None,
                               "display.width", 200,
                               "display.max_colwidth", 40):
            print(df.to_string(index=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
